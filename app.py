from flask import Flask, jsonify, request, render_template_string
import json, os, threading
from datetime import datetime

CONFIG_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
DATA_FILE     = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')
HISTORIAL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'historial.json')

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {'password': '1212'}

def save_config(cfg):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

app = Flask(__name__)
lock = threading.Lock()

_db = {'transactions': [], 'tx_id_counter': 0, 'tarjetas': {}, 'tarjetas_conf': []}

# Cargar datos guardados al iniciar
if os.path.exists(DATA_FILE):
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as _f:
            _loaded = json.load(_f)
            _db.update(_loaded)
    except Exception as _e:
        print(f'[data] Error cargando data.json: {_e}')

# Estado compartido entre todos los dispositivos
_state = {
    'hora_fin': '05:30',
    'premio': '',
    'winner_show': False,
    'winner_ts': 0,
    'cartel_show': False,
    'cartel_ts': 0,
    'cartel_data': {},
}

def load_data():
    return _db

def save_data(data):
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        print(f'[data] Error guardando: {e}')

def load_historial():
    if os.path.exists(HISTORIAL_FILE):
        try:
            with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_historial(h):
    try:
        with open(HISTORIAL_FILE, 'w', encoding='utf-8') as f:
            json.dump(h, f, ensure_ascii=False)
    except Exception as e:
        print(f'[historial] Error guardando: {e}')

HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ranking VIP</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Oswald:wght@400;500;600;700&family=Rajdhani:wght@400;500;600&display=swap');
:root{
  --gold:#c9a227;--gold-light:#e8c84a;--gold-dim:#7a6010;
  --black:#080808;--surface:#111;--border:#2a2a2a;
  --text:#f0ece0;--text-dim:#555;--danger:#a83030;
  --white:#fff;--green:#2ecc71;
  --surface-gold:#0d0b00;
}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--black);color:var(--text);font-family:'Rajdhani',sans-serif;min-height:100vh;}

/* TABS */
.tabs-bar{display:flex;background:#0a0a0a;border-bottom:1px solid #222;position:sticky;top:0;z-index:100;}
.tab-btn{flex:1;padding:13px 6px;text-align:center;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:var(--text-dim);background:none;border:none;border-bottom:2px solid transparent;transition:all .2s;}
.tab-btn:hover{color:#999;}
.tab-btn.active{color:var(--gold);border-bottom:2px solid var(--gold);}
.tab-btn .dot{display:inline-block;width:5px;height:5px;border-radius:50%;background:var(--gold);margin-left:5px;vertical-align:middle;animation:blink 2s infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}
body.modo-presentacion .tab-btn-caja{display:none;}
body.modo-presentacion .status-bar{display:none;}
body.modo-presentacion .config-panel{display:none;}
body.modo-presentacion .tabs-bar{display:none;}

.screen{display:none;padding:22px 26px 70px;}
.screen.active{display:block;}

/* CONFIG PANEL */
.config-panel{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:20px;}
.config-title{color:var(--gold-dim);font-size:10px;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;}
.config-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}
.config-row:last-child{margin-bottom:0;}
.config-label{color:var(--text-dim);font-size:12px;white-space:nowrap;}
.config-input{background:#0d0d0d;border:1px solid var(--border);border-radius:6px;color:var(--text);padding:8px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;}
.config-input:focus{outline:none;border-color:var(--gold);}
.config-input.wide{flex:1;min-width:180px;}
.config-input.narrow{width:110px;}
.btn-pres{background:var(--gold);color:#000;border:none;border-radius:6px;padding:9px 20px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;white-space:nowrap;transition:background .15s;}
.btn-pres:hover{background:var(--gold-light);}
.btn-reset{background:transparent;color:#555;border:1px solid #222;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;cursor:pointer;white-space:nowrap;transition:all .15s;}
.btn-reset:hover{border-color:var(--danger);color:#cc4444;}

/* PANTALLA HEADER */
.pres-header-wrap{position:relative;margin-bottom:28px;}
.pres-clock{position:absolute;top:0;left:0;text-align:left;}
.pres-clock-hora{font-family:'Oswald',sans-serif;font-size:13px;color:var(--text-dim);letter-spacing:1px;text-transform:uppercase;}
.pres-clock-time{font-family:'Oswald',sans-serif;font-size:32px;color:#ffffff;font-weight:700;line-height:1.1;}
.pres-clock-fin{font-size:11px;color:var(--gold-dim);letter-spacing:1px;margin-top:6px;text-transform:uppercase;}
.pres-clock-fin-val{font-family:'Oswald',sans-serif;font-size:26px;color:var(--gold);font-weight:700;line-height:1.1;display:block;}
.pres-header{text-align:center;padding-top:8px;}
.pres-logo{font-family:'Oswald',sans-serif;font-size:52px;font-weight:700;color:var(--white);letter-spacing:10px;text-transform:uppercase;display:block;width:100%;}
.pres-logo .vip{color:var(--gold);}
.pres-logo .club{font-size:26px;font-weight:600;color:#888;letter-spacing:10px;display:block;margin-top:2px;text-transform:uppercase;width:100%;}
.pres-line{height:1px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:12px auto;max-width:100%;}
.live-badge{display:inline-flex;align-items:center;gap:6px;border:1px solid #2a2a2a;border-radius:20px;padding:4px 14px;font-size:11px;color:#777;letter-spacing:1px;margin-top:6px;}
.live-dot{width:6px;height:6px;border-radius:50%;background:#3a9a5a;animation:blink 1.5s infinite;}

/* RANKING */
.ranking-wrap{max-width:100%;margin:0;padding:0 10px;}
.rank-header{display:grid;grid-template-columns:100px 1fr 130px 180px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:8px 8px 0 0;padding:12px 30px;margin-bottom:3px;}
.rank-header span{font-family:'Oswald',sans-serif;font-size:13px;font-weight:500;letter-spacing:2px;text-transform:uppercase;color:var(--gold-dim);}
.rank-header .col-r{text-align:right;}
.rank-rows{display:flex;flex-direction:column;gap:4px;}
.rank-row{display:grid;grid-template-columns:100px 1fr 130px 180px;align-items:center;background:var(--surface);border:1px solid #1e1e1e;border-radius:6px;padding:20px 30px;transition:border-color .3s, background .4s, box-shadow .4s;}
.rank-row.rank-1{background:var(--surface-gold);border-color:var(--gold-dim);}
/* Nueva tarjeta que entra */
.rank-row.nueva{animation:entradaFila .55s cubic-bezier(.22,1,.36,1) both;}
@keyframes entradaFila{from{opacity:0;transform:translateX(-32px)}to{opacity:1;transform:none}}
/* Highlight flash al subir al #1 */
@keyframes crownGlow{
  0%  {box-shadow:0 0 0px rgba(201,162,39,0);}
  40% {box-shadow:0 0 32px rgba(201,162,39,.55);}
  100%{box-shadow:none;}
}
.rank-row.rank-1.ascendio{animation:crownGlow .9s ease both;}
/* Guante que golpea hacia la derecha */
@keyframes gloveKnockout{
  0%  {left:-60px;opacity:1;transform:scale(1.2) rotate(-15deg);}
  35% {left:60%;opacity:1;transform:scale(1.3) rotate(-10deg);}
  50% {left:70%;opacity:0.8;transform:scale(1.1) rotate(-5deg);}
  70% {left:110%;opacity:0;transform:scale(0.8) rotate(0deg);}
  100%{left:110%;opacity:0;}
}
/* Punch hit */
@keyframes punchHit{
  0%  {transform:translateX(-60px) translateY(22px) rotate(-5deg);filter:brightness(2);}
  30% {transform:translateX(8px) translateY(0) rotate(1deg);filter:brightness(1.5);}
  60% {transform:translateX(-3px) rotate(-0.5deg);}
  100%{transform:none;filter:none;}
}
.col-puesto{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:#444;}
.rank-row.rank-1 .col-puesto{color:var(--gold);font-size:34px;}
.rank-row.rank-2 .col-puesto{color:#aaa;}
.rank-row.rank-3 .col-puesto{color:#8a6a40;}
.col-nombre{font-family:'Oswald',sans-serif;font-size:30px;font-weight:600;color:var(--white);}
.rank-row.rank-1 .col-nombre{font-size:36px;}
.col-mesa{font-family:'Oswald',sans-serif;font-size:24px;font-weight:700;color:#e8e8e8;letter-spacing:1px;}
.rank-row.rank-1 .col-mesa{font-size:28px;color:#ffffff;}
.col-total{font-family:'Oswald',sans-serif;font-size:30px;font-weight:700;color:var(--gold);text-align:right;}
.rank-row.rank-1 .col-total{font-size:38px;}
.miles-lbl{font-size:0.45em;opacity:0.55;letter-spacing:2px;margin-left:5px;vertical-align:middle;font-weight:600;}

.btn-exit-float{display:none;position:fixed;top:12px;left:16px;z-index:9999;background:transparent;color:#1c1c1c;border:none;font-size:22px;font-family:'Oswald',sans-serif;font-weight:300;cursor:pointer;padding:4px 8px;transition:color .3s;line-height:1;}
.btn-exit-float:hover{color:#555;}
body.modo-presentacion .btn-exit-float{display:block;}
.empty-msg{text-align:center;color:#222;font-size:15px;padding:70px 20px;letter-spacing:2px;font-family:'Oswald',sans-serif;}

/* PREMIO */
.premio-wrap{text-align:center;margin-top:36px;padding-bottom:20px;}
.premio-box{display:inline-block;background:var(--gold);color:#000;font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;letter-spacing:.5px;padding:14px 44px;border-radius:6px;border:none;}
.premio-box:empty{display:none;}
@keyframes bottleBounce{
  0%,100%{transform:translateY(0) rotate(-5deg) scale(1);}
  20%{transform:translateY(-18px) rotate(5deg) scale(1.05);}
  40%{transform:translateY(-6px) rotate(-3deg) scale(0.98);}
  60%{transform:translateY(-14px) rotate(4deg) scale(1.03);}
  80%{transform:translateY(-4px) rotate(-2deg) scale(0.99);}
}

/* CAJAS */
.caja-header-row{display:flex;align-items:center;gap:12px;margin-bottom:20px;}
.caja-badge{background:var(--gold);color:#000;border-radius:6px;padding:5px 14px;font-family:'Oswald',sans-serif;font-size:13px;letter-spacing:2px;font-weight:600;}
.caja-title{font-family:'Oswald',sans-serif;font-size:20px;color:var(--white);letter-spacing:1px;}
.modo-tabs{display:flex;gap:8px;margin-bottom:16px;}
.modo-tab{flex:1;padding:10px;text-align:center;background:var(--surface);border:1px solid var(--border);border-radius:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:600;letter-spacing:1px;color:var(--text-dim);transition:all .2s;}
.modo-tab.active{background:#0f0e05;border-color:var(--gold-dim);color:var(--gold);}
.modo-content{display:none;}
.modo-content.active{display:block;}
.scan-hint{display:flex;align-items:center;gap:10px;background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:16px;color:var(--text-dim);font-size:13px;letter-spacing:1px;}
.scan-hint.esperando{border-color:var(--gold-dim);color:var(--gold);}
.scan-icon{font-size:20px;opacity:.4;}
.scan-hint.esperando .scan-icon{opacity:1;animation:blink .6s infinite;}
.tarjeta-card{background:#0a1200;border:1px solid #1a3a00;border-radius:10px;padding:16px 20px;margin-bottom:16px;display:none;}
.tarjeta-card.visible{display:block;}
.tarjeta-card.sin-saldo{background:#120000;border-color:#3a0000;}
.tarjeta-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;}
.tarjeta-mesa-label{color:#3a6a00;font-size:11px;letter-spacing:2px;text-transform:uppercase;}
.tarjeta-card.sin-saldo .tarjeta-mesa-label{color:#6a2000;}
.tarjeta-mesa-num{font-family:'Oswald',sans-serif;font-size:32px;color:#ffffff;font-weight:700;}
.tarjeta-card.sin-saldo .tarjeta-mesa-num{color:#ff8888;}
.tarjeta-saldo-wrap{text-align:right;}
.tarjeta-saldo-label{color:#3a6a00;font-size:10px;letter-spacing:1px;text-transform:uppercase;}
.tarjeta-saldo{font-family:'Oswald',sans-serif;font-size:26px;color:var(--green);font-weight:700;}
.tarjeta-card.sin-saldo .tarjeta-saldo{color:var(--danger);}
.tarjeta-bar-wrap{height:4px;background:#1a1a1a;border-radius:2px;}
.tarjeta-bar{height:4px;background:var(--green);border-radius:2px;transition:width .5s ease;}
.tarjeta-card.sin-saldo .tarjeta-bar{background:var(--danger);}
.tarjeta-nombre{margin-top:10px;font-size:13px;color:#3a5a00;}
.tarjeta-nombre span{color:#aaddaa;font-weight:600;}
.form-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:20px;margin-bottom:18px;}
.field-label{color:#888;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;display:block;font-weight:600;}
.field-input{width:100%;background:#0d0d0d;border:1px solid var(--border);border-radius:7px;color:var(--text);padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:500;transition:border-color .2s;margin-bottom:14px;}
.field-input:focus{outline:none;border-color:var(--gold);}
.field-input::placeholder{color:#333;}
.field-input.amount-input{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:var(--gold);padding:13px 16px;letter-spacing:1px;}
.field-input.amount-input::placeholder{color:#222;font-size:20px;}
.hint-miles{color:#3a3a00;font-size:13px;margin-top:-10px;margin-bottom:10px;letter-spacing:1px;}
.hint-miles.ok{color:#3a6a00;font-family:'Oswald',sans-serif;font-size:15px;font-weight:600;}
.btn-row{display:flex;gap:10px;}
.btn-add{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 22px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;text-transform:uppercase;cursor:pointer;flex:1;transition:background .15s,transform .1s;}
.btn-add:hover{background:var(--gold-light);}
.btn-add:active{transform:scale(.97);}
.btn-add:disabled{background:#2a2a2a;color:#555;cursor:not-allowed;}
.section-label{color:#2a2a2a;font-size:10px;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;}
.tx-list{display:flex;flex-direction:column;gap:6px;max-height:240px;overflow-y:auto;}
.tx-list::-webkit-scrollbar{width:3px;}
.tx-list::-webkit-scrollbar-thumb{background:#2a2a2a;border-radius:3px;}
.tx-item{display:flex;justify-content:space-between;align-items:center;background:var(--surface);border:1px solid #1a1a1a;border-radius:7px;padding:10px 14px;}
.tx-info{flex:1;}
.tx-name{font-size:14px;font-weight:500;color:var(--text);}
.tx-meta{font-size:11px;color:#3a3a3a;margin-top:1px;}
.tx-right{display:flex;align-items:center;gap:10px;}
.tx-amount{font-family:'Oswald',sans-serif;font-size:16px;color:var(--gold);}
.btn-del{background:none;border:1px solid #2a1a1a;color:#4a2a2a;border-radius:5px;padding:4px 8px;font-size:12px;cursor:pointer;transition:all .15s;}
.btn-del:hover{border-color:var(--danger);color:#cc4444;}
.btn-edit{background:none;border:1px solid #1a2a1a;color:#2a4a2a;border-radius:5px;padding:4px 8px;font-size:12px;cursor:pointer;transition:all .15s;margin-right:4px;}
.btn-edit:hover{border-color:#3a6a10;color:#6a9a30;}
.caja-total-bar{background:var(--surface);border:1px solid #1e1a00;border-radius:8px;padding:13px 18px;display:flex;justify-content:space-between;align-items:center;margin-top:14px;}
.caja-total-label{color:#444;font-size:11px;letter-spacing:2px;text-transform:uppercase;}
.caja-total-val{font-family:'Oswald',sans-serif;font-size:24px;color:var(--gold);}
.no-tx{color:#1e1e1e;font-size:13px;padding:14px 0;letter-spacing:1px;}

/* CONFIGURACION TARJETAS */
.conf-header{margin-bottom:24px;}
.conf-title{font-family:'Oswald',sans-serif;font-size:22px;color:var(--white);letter-spacing:1px;margin-bottom:6px;}
.conf-sub{color:var(--text-dim);font-size:13px;line-height:1.7;}
.tarjetas-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:10px;margin-bottom:24px;}
.tarjeta-conf{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 16px;}
.tarjeta-conf.configurada{border-color:#2a3a00;background:#0a0f00;}
.tc-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;}
.tc-num{font-family:'Oswald',sans-serif;font-size:16px;color:var(--gold-dim);font-weight:600;}
.tarjeta-conf.configurada .tc-num{color:var(--gold);}
.tc-status{font-size:10px;letter-spacing:1px;color:#2a2a2a;text-transform:uppercase;}
.tarjeta-conf.configurada .tc-status{color:#3a6a00;}
.tc-btns{display:flex;gap:6px;align-items:center;}
.tc-scan-btn{background:#1a1a1a;color:#555;border:1px solid #2a2a2a;border-radius:5px;padding:4px 10px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;letter-spacing:1px;transition:all .15s;}
.tc-scan-btn:hover{border-color:var(--gold-dim);color:var(--gold);}
.tc-scan-btn.activo{border-color:var(--gold);color:var(--gold);animation:blink .6s infinite;}
.tc-confirm-btn{background:var(--gold);color:#000;border:none;border-radius:5px;padding:4px 10px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;transition:background .15s;}
.tc-confirm-btn:hover{background:var(--gold-light);}
.tc-clear-btn{background:none;color:#333;border:1px solid #1e1e1e;border-radius:5px;padding:4px 8px;font-size:11px;cursor:pointer;transition:all .15s;}
.tc-clear-btn:hover{border-color:var(--danger);color:#cc4444;}
.tc-field{display:flex;flex-direction:column;gap:4px;margin-bottom:8px;}
.tc-label{font-size:10px;letter-spacing:1px;color:#888;text-transform:uppercase;font-weight:600;}
.tc-input{background:#0d0d0d;border:1px solid var(--border);border-radius:5px;color:var(--text);padding:7px 10px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:500;width:100%;}
.tc-input:focus{outline:none;border-color:var(--gold);}
.tc-input::placeholder{color:#333;}
.tc-code{font-size:10px;color:#2a2a2a;margin-top:4px;font-family:monospace;}
.tarjeta-conf.configurada .tc-code{color:#2a4a00;}
.tc-saldo-bar{margin-top:8px;}
.tc-saldo-info{display:flex;justify-content:space-between;font-size:11px;margin-bottom:4px;}
.tc-saldo-used{color:var(--danger);}
.tc-saldo-left{color:var(--green);}
.saldo-bajo-warn{background:#1a0800;border:1px solid #6a2a00;border-radius:7px;padding:8px 12px;font-size:12px;color:#ff7733;letter-spacing:1px;margin-top:6px;display:flex;align-items:center;gap:7px;animation:warnPulse 1.5s ease-in-out infinite alternate;}
@keyframes warnPulse{0%{border-color:#6a2a00;color:#ff7733;}100%{border-color:#ff7733;color:#ffaa66;}}
.tc-bar-wrap{height:3px;background:#1a1a1a;border-radius:2px;}
.tc-bar-fill{height:3px;background:var(--green);border-radius:2px;transition:width .5s;}
.conf-actions{display:flex;gap:10px;flex-wrap:wrap;}
.btn-guardar-conf{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 30px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;}
.btn-guardar-conf:hover{background:var(--gold-light);}
.btn-limpiar-conf{background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;}
.btn-limpiar-conf:hover{border-color:var(--danger);color:#cc4444;}

/* STATUS BAR */
.status-bar{position:fixed;bottom:0;left:0;right:0;background:#050505;border-top:1px solid #141414;padding:6px 18px;display:flex;justify-content:space-between;align-items:center;font-size:11px;color:#2a2a2a;letter-spacing:1px;z-index:200;}
.status-ok{color:#3a6a3a;}
.status-err{color:var(--danger);}
.toast{position:fixed;bottom:40px;left:50%;transform:translateX(-50%);background:#1a2a00;border:1px solid #3a6a00;color:var(--green);padding:10px 24px;border-radius:8px;font-size:14px;font-weight:600;z-index:9999;opacity:0;transition:opacity .3s;pointer-events:none;white-space:nowrap;}
.toast.show{opacity:1;}
.toast.error{background:#2a0000;border-color:#6a0000;color:#ff6666;}

/* ══════════════════════════════════════════
   STATS
══════════════════════════════════════════ */
.stats-header{margin-bottom:24px;}
.stats-title{font-family:'Oswald',sans-serif;font-size:22px;color:var(--white);letter-spacing:1px;margin-bottom:4px;}
.stats-sub{font-size:13px;color:var(--text-dim);letter-spacing:1px;}
.kpi-row{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:12px;margin-bottom:24px;}
.kpi-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px 20px;}
.kpi-label{font-size:10px;color:var(--text-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;}
.kpi-val{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:var(--gold);}
.charts-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px;}
@media(max-width:700px){.charts-row{grid-template-columns:1fr;}}
.chart-box{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px 20px;}
.chart-title{font-family:'Oswald',sans-serif;font-size:14px;color:var(--gold-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:14px;}
.chart-legend{display:flex;flex-wrap:wrap;gap:10px;margin-top:12px;}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--text-dim);}
.legend-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0;}
.stats-section-title{font-family:'Oswald',sans-serif;font-size:14px;color:var(--gold-dim);letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;padding-bottom:6px;border-bottom:1px solid #1e1e1e;}
/* Top clientes */
.top-list{display:flex;flex-direction:column;gap:8px;margin-bottom:24px;}
.top-item{display:flex;align-items:center;gap:12px;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:12px 16px;}
.top-pos{font-family:'Oswald',sans-serif;font-size:20px;font-weight:700;color:#333;width:30px;text-align:center;flex-shrink:0;}
.top-item.pos-1 .top-pos{color:var(--gold);}
.top-item.pos-2 .top-pos{color:#aaa;}
.top-item.pos-3 .top-pos{color:#8a6a40;}
.top-name{font-family:'Oswald',sans-serif;font-size:18px;color:var(--white);flex:1;}
.top-mesa{font-size:12px;color:var(--text-dim);}
.top-bar-wrap{flex:2;height:6px;background:#1a1a1a;border-radius:3px;}
.top-bar-fill{height:6px;background:var(--gold);border-radius:3px;transition:width .6s ease;}
.top-amount{font-family:'Oswald',sans-serif;font-size:18px;color:var(--gold);text-align:right;min-width:90px;}
/* Detalle cajas */
.cajas-detail-row{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:24px;}
@media(max-width:600px){.cajas-detail-row{grid-template-columns:1fr;}}
.caja-stat-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}
.caja-stat-badge{background:var(--gold);color:#000;font-family:'Oswald',sans-serif;font-size:11px;letter-spacing:2px;font-weight:700;padding:3px 10px;border-radius:4px;display:inline-block;margin-bottom:10px;}
.caja-stat-total{font-family:'Oswald',sans-serif;font-size:26px;color:var(--gold);font-weight:700;margin-bottom:4px;}
.caja-stat-ops{font-size:12px;color:var(--text-dim);}
.caja-stat-list{margin-top:10px;display:flex;flex-direction:column;gap:4px;max-height:140px;overflow-y:auto;}
.caja-stat-item{display:flex;justify-content:space-between;font-size:13px;color:var(--text-dim);padding:4px 0;border-bottom:1px solid #1a1a1a;}
.caja-stat-item:last-child{border-bottom:none;}
.caja-stat-item span:last-child{color:var(--gold);}

/* ══════════════════════════════════════════
   PANEL DE PERSONALIZACION
══════════════════════════════════════════ */
.custom-section{margin-bottom:28px;}
.custom-section-title{font-family:'Oswald',sans-serif;font-size:16px;color:var(--gold);letter-spacing:2px;text-transform:uppercase;margin-bottom:14px;padding-bottom:6px;border-bottom:1px solid #1e1e1e;}
.color-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;margin-bottom:10px;}
.color-item{display:flex;align-items:center;gap:10px;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:10px 14px;}
.color-swatch{width:36px;height:36px;border-radius:6px;border:2px solid #333;cursor:pointer;flex-shrink:0;position:relative;overflow:hidden;}
.color-swatch input[type=color]{position:absolute;inset:-4px;width:calc(100%+8px);height:calc(100%+8px);border:none;cursor:pointer;opacity:0;}
.color-label{font-size:12px;color:var(--text-dim);letter-spacing:1px;flex:1;}
.color-hex{font-size:11px;color:#444;font-family:monospace;}
.custom-text-row{display:flex;gap:12px;align-items:center;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:12px 16px;margin-bottom:10px;}
.custom-text-label{font-size:12px;color:var(--text-dim);letter-spacing:1px;white-space:nowrap;min-width:130px;}
.custom-text-input{flex:1;background:#0d0d0d;border:1px solid var(--border);border-radius:6px;color:var(--text);padding:9px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;}
.custom-text-input:focus{outline:none;border-color:var(--gold);}
.btn-custom-save{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 32px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;margin-right:10px;}
.btn-custom-save:hover{background:var(--gold-light);}
.btn-custom-reset{background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;}
.btn-custom-reset:hover{border-color:var(--danger);color:#cc4444;}
.preview-bar{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin-bottom:20px;text-align:center;}
.preview-label{font-size:10px;color:var(--text-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;}
.preview-logo{font-family:'Oswald',sans-serif;font-size:32px;font-weight:700;letter-spacing:6px;color:var(--white);}
.preview-logo .vip-prev{color:var(--gold);}
.preview-logo .club-prev{font-size:13px;font-weight:400;color:#5a5a5a;letter-spacing:5px;display:block;margin-top:2px;}

/* ══════════════════════════════════════════
   ANIMACION GANADOR
══════════════════════════════════════════ */
#winner-overlay{
  display:none;
  position:fixed;inset:0;z-index:10000;
  background:rgba(0,0,0,0.97);
  flex-direction:column;align-items:center;justify-content:center;
  overflow:hidden;
}
#winner-overlay.show{display:flex;}

/* Confetti particles */
.confetti-wrap{position:absolute;inset:0;pointer-events:none;overflow:hidden;}
.confetti-piece{
  position:absolute;top:-20px;
  width:10px;height:16px;
  border-radius:2px;
  animation:confettiFall linear infinite;
}
@keyframes confettiFall{
  0%  {transform:translateY(-20px) rotate(0deg);opacity:1;}
  100%{transform:translateY(110vh) rotate(720deg);opacity:.2;}
}

/* Destellos radiales */
.winner-rays{
  position:absolute;inset:0;
  background:conic-gradient(from 0deg, transparent 0deg, rgba(201,162,39,0.04) 10deg, transparent 20deg,
    transparent 40deg, rgba(201,162,39,0.04) 50deg, transparent 60deg,
    transparent 80deg, rgba(201,162,39,0.04) 90deg, transparent 100deg,
    transparent 120deg, rgba(201,162,39,0.04) 130deg, transparent 140deg,
    transparent 160deg, rgba(201,162,39,0.04) 170deg, transparent 180deg,
    transparent 200deg, rgba(201,162,39,0.04) 210deg, transparent 220deg,
    transparent 240deg, rgba(201,162,39,0.04) 250deg, transparent 260deg,
    transparent 280deg, rgba(201,162,39,0.04) 290deg, transparent 300deg,
    transparent 320deg, rgba(201,162,39,0.04) 330deg, transparent 340deg,
    transparent 360deg);
  animation:raysRotate 12s linear infinite;
  pointer-events:none;
}
@keyframes raysRotate{from{transform:rotate(0deg)}to{transform:rotate(360deg)}}

.winner-content{
  position:relative;z-index:2;
  text-align:center;
  animation:winnerEntrada 1s cubic-bezier(.22,1,.36,1) both;
}
@keyframes winnerEntrada{
  0%  {opacity:0;transform:scale(.5) translateY(60px);}
  60% {transform:scale(1.05) translateY(-10px);}
  100%{opacity:1;transform:scale(1) translateY(0);}
}

.winner-corona{font-size:80px;animation:coronaPulse 1.5s ease-in-out infinite alternate;display:block;margin-bottom:10px;}
@keyframes coronaPulse{from{transform:scale(1) rotate(-5deg)}to{transform:scale(1.15) rotate(5deg)}}


.winner-titulo{
  font-family:'Oswald',sans-serif;
  font-size:22px;font-weight:500;
  letter-spacing:10px;text-transform:uppercase;
  color:var(--gold-dim);margin-bottom:6px;
}
.winner-nombre{
  font-family:'Oswald',sans-serif;
  font-size:clamp(52px,8vw,110px);
  font-weight:700;
  color:#fff;
  letter-spacing:4px;
  text-transform:uppercase;
  line-height:1;
  margin-bottom:10px;
  text-shadow:0 0 60px rgba(201,162,39,.5);
  animation:nombreGlow 2s ease-in-out infinite alternate;
}
@keyframes nombreGlow{
  from{text-shadow:0 0 40px rgba(201,162,39,.3);}
  to  {text-shadow:0 0 80px rgba(201,162,39,.8), 0 0 120px rgba(201,162,39,.3);}
}
.winner-line{
  height:2px;
  background:linear-gradient(to right,transparent,var(--gold),transparent);
  margin:16px auto;width:80%;
  animation:lineExpand 1s ease both;animation-delay:.5s;
  transform-origin:center;
}
@keyframes lineExpand{from{transform:scaleX(0)}to{transform:scaleX(1)}}

.winner-info-row{
  display:flex;gap:60px;justify-content:center;align-items:center;
  margin:20px 0;flex-wrap:wrap;
}
.winner-info-block{text-align:center;}
.winner-info-label{
  font-size:12px;letter-spacing:3px;text-transform:uppercase;
  color:var(--gold-dim);margin-bottom:4px;
}
.winner-info-val{
  font-family:'Oswald',sans-serif;
  font-size:clamp(28px,4vw,52px);
  font-weight:700;color:var(--gold);
}

.winner-mensaje{
  font-family:'Rajdhani',sans-serif;
  font-size:clamp(20px,3vw,36px);
  font-weight:600;
  color:#fff;
  margin-top:24px;
  padding:16px 40px;
  border:2px solid var(--gold);
  border-radius:8px;
  background:rgba(201,162,39,0.08);
  max-width:90vw;
  line-height:1.3;
  animation:mensajeAppear 1s ease both;animation-delay:.8s;
  opacity:0;
}
@keyframes mensajeAppear{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:none}}

.winner-close{
  position:fixed;top:20px;right:28px;z-index:10001;
  background:transparent;color:#333;border:none;
  font-size:28px;cursor:pointer;font-family:'Oswald',sans-serif;
  transition:color .2s;line-height:1;
}
.winner-close:hover{color:#888;}

/* Boton manual ganador */
.btn-show-winner{
  background:linear-gradient(135deg,#c9a227,#e8c84a);
  color:#000;border:none;border-radius:7px;
  padding:10px 24px;
  font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;
  letter-spacing:1px;cursor:pointer;
  transition:all .2s;
  box-shadow:0 4px 20px rgba(201,162,39,.3);
}
.btn-show-winner:hover{transform:translateY(-1px);box-shadow:0 6px 28px rgba(201,162,39,.4);}

@keyframes shimmerGold{0%{background-position:200% center}100%{background-position:-200% center}}
@keyframes shimmerRed{0%{background-position:200% center}100%{background-position:-200% center}}
@keyframes emojiGlow{from{filter:drop-shadow(0 0 20px rgba(201,162,39,0.4))}to{filter:drop-shadow(0 0 50px rgba(201,162,39,0.9))}}

/* ══ TEMA JAGGER 12 AÑOS (B&W + PODIO DORADO/PLATEADO/BRONCE) ══ */
body.tema-jagger12 .rank-row{border-color:#1e1e1e;}
body.tema-jagger12 .rank-row.rank-1{background:#1a1200;border-color:#c9a227;}
body.tema-jagger12 .rank-row.rank-2{background:#0e0e0e;border-color:#777;}
body.tema-jagger12 .rank-row.rank-3{background:#0e0800;border-color:#7a4a20;}
body.tema-jagger12 .rank-row.rank-1 .col-puesto{color:#c9a227;font-size:34px;}
body.tema-jagger12 .rank-row.rank-2 .col-puesto{color:#aaaaaa;}
body.tema-jagger12 .rank-row.rank-3 .col-puesto{color:#cd7f32;}
body.tema-jagger12 .rank-row.rank-1 .col-total{color:#e8c84a;}
body.tema-jagger12 .rank-row.rank-2 .col-total{color:#cccccc;}
body.tema-jagger12 .rank-row.rank-3 .col-total{color:#cd7f32;}
body.tema-jagger12 .col-total{color:#ccc;}
body.tema-jagger12 .pres-line{background:linear-gradient(to right,transparent,#c9a227,transparent);}
body.tema-jagger12 .live-dot{background:#c9a227;}
body.tema-jagger12 .col-nombre{color:#fff !important;}
body.tema-jagger12 .col-mesa{color:#ddd !important;}

/* ══ TEMA VELADA DE BOXEO ══ */
body.tema-jagger12boxeo .rank-row{border-color:#2a0000;}
body.tema-jagger12boxeo .rank-row.rank-1{background:#1a0000;border-color:#880000;}
body.tema-jagger12boxeo .pres-line{background:linear-gradient(to right,transparent,#ff2222,transparent);}
body.tema-jagger12boxeo .live-dot{background:#ff2222;}
body.tema-jagger12boxeo .live-badge{border-color:#3a0000;color:#aa3333;}
body.tema-jagger12boxeo .col-nombre{color:#fff0f0 !important;}
body.tema-jagger12boxeo .col-mesa{color:#ffaaaa !important;}

/* ══ TEMA A TOUCH OF PINK ══ */
body.tema-touchofpink .rank-header{background:#2d0022;border-color:#6a2050;}
body.tema-touchofpink .rank-header span{color:#f472b6;text-shadow:0 0 8px rgba(244,114,182,0.4);}
body.tema-touchofpink .rank-row{border-color:#8a3070;background:#3d002c;}
body.tema-touchofpink .rank-row.rank-1{background:#5a0042;border-color:#f472b6;box-shadow:0 0 20px rgba(244,114,182,.2);}
body.tema-touchofpink .rank-row.rank-1 .col-puesto{color:#f472b6;font-size:34px;}
body.tema-touchofpink .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink .col-total{color:#fbb6ce;}
body.tema-touchofpink .col-nombre{color:#ffffff !important;}
body.tema-touchofpink .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink .col-puesto{color:#eeaad8;}
body.tema-touchofpink .pres-line{background:linear-gradient(to right,transparent,#f472b6,transparent);}
body.tema-touchofpink .live-dot{background:#f472b6;}
body.tema-touchofpink .live-badge{border-color:#8a3070;color:#f472b6;}
body.tema-touchofpink .rank-row.rank-1.ascendio{animation:crownGlow .9s ease both;}
/* A Touch of Pink — modo rose medio */
body.tema-touchofpink.pink-claro{--black:#3a0028;--surface:#580040;--border:#a04080;--gold:#f472b6;--gold-light:#fbb6ce;--gold-dim:#e896cc;--text:#ffe8f5;--text-dim:#ddaacc;--white:#ffffff;}
body.tema-touchofpink.pink-claro .rank-header{background:#4a0035;border-color:#a04080;}
body.tema-touchofpink.pink-claro .rank-header span{color:#f472b6;text-shadow:0 0 8px rgba(244,114,182,0.4);}
body.tema-touchofpink.pink-claro .rank-row{background:#580040;border-color:#a04080;}
body.tema-touchofpink.pink-claro .rank-row.rank-1{background:#7a0058;border-color:#f472b6;box-shadow:0 0 24px rgba(244,114,182,.3);}
body.tema-touchofpink.pink-claro .rank-row.rank-1 .col-puesto{color:#f472b6;}
body.tema-touchofpink.pink-claro .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink.pink-claro .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink.pink-claro .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink.pink-claro .col-total{color:#fbb6ce;}
body.tema-touchofpink.pink-claro .col-nombre{color:#ffffff !important;}
body.tema-touchofpink.pink-claro .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink.pink-claro .col-puesto{color:#eeaad8;}

/* ══ CARTEL OVERLAY ══ */
#cartel-overlay.show{display:flex !important;}

</style>
</head>
<body>

<div class="tabs-bar">
  <button class="tab-btn active" onclick="showTab('pantalla')" id="tbtn-pantalla">Pantalla <span class="dot"></span></button>
  <button class="tab-btn tab-btn-caja" onclick="showTab('caja1')" id="tbtn-caja1">Abajo</button>
  <button class="tab-btn tab-btn-caja" onclick="showTab('caja2')" id="tbtn-caja2">Extendido</button>
  <button class="tab-btn tab-btn-caja" onclick="showTab('caja3')" id="tbtn-caja3">VIP</button>
  <button class="tab-btn tab-btn-caja" onclick="showTab('config')" id="tbtn-config">Tarjetas</button>
  <button class="tab-btn tab-btn-caja" onclick="showTab('stats')" id="tbtn-stats">📊 Stats</button>
  <button class="tab-btn tab-btn-caja" onclick="showTab('custom')" id="tbtn-custom">Diseño</button>
</div>

<!-- PANTALLA -->
<div id="tab-pantalla" class="screen active">
  <div class="config-panel">
    <div class="config-title">Configuracion de pantalla</div>
    <div class="config-row">
      <span class="config-label">Premio:</span>
      <input class="config-input wide" id="msg-input" type="text" placeholder="Ej: El ganador se lleva una botella gratis" oninput="updateMsg()" />
      <select class="config-input" id="premio-size" onchange="updatePremioSize()" title="Tamaño del texto del premio" style="width:auto;padding:8px 10px;">
        <option value="16px">Chico</option>
        <option value="22px" selected>Normal</option>
        <option value="30px">Grande</option>
        <option value="42px">Muy grande</option>
        <option value="56px">Enorme</option>
      </select>
      <button class="btn-pres" onclick="activarPresentacion()">Modo Presentacion</button>
      <button class="btn-reset" onclick="resetNoche()">Resetear noche</button>
      <button onclick="cerrarNoche()" style="background:transparent;color:#3a9a5a;border:1px solid #1a3a2a;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;white-space:nowrap;transition:all .15s;" onmouseover="this.style.borderColor='#3a9a5a'" onmouseout="this.style.borderColor='#1a3a2a'">💾 Cerrar noche</button>

    </div>
    <div class="config-row">
      <span class="config-label">Hora de finalizacion:</span>
      <input class="config-input narrow" id="hora-fin-input" type="time" value="05:30" oninput="updateHoraFin()" />
      <button class="btn-show-winner" onclick="mostrarGanadorManual()">🏆 Mostrar Ganador Ahora</button>
    </div>
  </div>

  <div class="pres-header-wrap">
    <!-- Decoraciones temáticas -->
    <div id="tema-overlay" style="position:fixed;inset:0;pointer-events:none;z-index:0;overflow:hidden;opacity:0;transition:opacity 1s;"></div>
    <!-- Efectos de fondo permanentes -->
    <div id="efectos-overlay" style="position:fixed;inset:0;pointer-events:none;z-index:0;overflow:hidden;"></div>
    <div class="pres-clock">
      <div class="pres-clock-hora">Hora</div>
      <div class="pres-clock-time" id="clock-display">00:00</div>
      <div class="pres-clock-fin">Finaliza</div>
      <span class="pres-clock-fin-val" id="clock-fin">05:30</span>
    </div>
    <div class="pres-header">
      <div class="pres-logo" id="main-logo">RANKING <span class="vip" id="logo-vip">VIP</span></div><div id="tema-tagline" style="font-family:'Rajdhani',sans-serif;font-size:28px;font-weight:600;color:#555;letter-spacing:5px;text-transform:uppercase;text-align:center;margin-top:2px;min-height:0;">JAGGER CLUB</div>
      <div class="pres-line"></div>
      <div class="live-badge"><span class="live-dot"></span> EN VIVO</div>
    </div>
  </div>

  <div class="ranking-wrap">
    <div class="rank-header" id="rank-header" style="display:none">
      <span>PUESTO</span><span>NOMBRE</span><span>MESA</span><span class="col-r">TOTAL</span>
    </div>
    <div class="rank-rows" id="rank-rows"></div>
    <div id="empty-msg" class="empty-msg">Aun no hay consumos registrados</div>
  </div>
  <div class="premio-wrap"><div class="premio-box" id="premio-box"></div></div>
  <div id="credit-pres" style="text-align:center;margin-top:18px;color:#2a2a2a;font-family:'Rajdhani',sans-serif;font-size:13px;letter-spacing:2px;">Made by Santino Sucatti</div>
</div>

<button class="btn-exit-float" onclick="salirPresentacion()">&lt;</button>

<!-- CAJAS -->
<div id="tab-caja1" class="screen"><div id="caja-inner-1"></div></div>
<div id="tab-caja2" class="screen"><div id="caja-inner-2"></div></div>
<div id="tab-caja3" class="screen"><div id="caja-inner-3"></div></div>

<!-- CONFIGURACION TARJETAS -->
<div id="tab-config" class="screen">
  <div class="conf-header">
    <div class="conf-title">Configuracion de Tarjetas</div>
    <div class="conf-sub">
      La mesa se asigna automaticamente igual al numero de tarjeta (Tarjeta 1 = Mesa 1).<br>
      Solo tenes que vincular la tarjeta fisica y poner el saldo inicial como monto completo (ej: 50000).<br>
      Presiona LEER en el slot y pasa la tarjeta por el lector.
    </div>
  </div>
  <div class="tarjetas-grid" id="tarjetas-grid"></div>
  <div style="background:#1a0e00;border:2px solid #c9a227;border-radius:10px;padding:14px 20px;margin-bottom:18px;display:flex;align-items:center;gap:14px;">
    <span style="font-size:28px;">⚠️</span>
    <div>
      <div style="font-family:'Oswald',sans-serif;font-size:15px;color:#c9a227;letter-spacing:1px;font-weight:700;margin-bottom:3px;">MUY IMPORTANTE</div>
      <div style="font-size:13px;color:#e8c84a;line-height:1.6;">Después de <strong>cualquier cambio</strong> (vincular tarjeta, cambiar saldo, limpiar slot) siempre hacé click en <strong>"Guardar configuración"</strong>. Si no guardás, los cambios se pierden al recargar la página.</div>
    </div>
  </div>
  <div style="margin-bottom:16px;">
    <button onclick="abrirCartelModal()" style="background:linear-gradient(135deg,#c9a227,#e8c84a);color:#000;border:none;border-radius:7px;padding:12px 28px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;letter-spacing:2px;cursor:pointer;transition:all .15s;" onmouseover="this.style.opacity='0.85'" onmouseout="this.style.opacity='1'">📣 CARTEL</button>
  </div>
  <div class="conf-actions">
    <button class="btn-guardar-conf" onclick="guardarTarjetas()">💾 Guardar configuracion</button>
    <button class="btn-limpiar-conf" onclick="limpiarTarjetas()">Limpiar todo</button>
  </div>
</div>

<!-- ══ PANEL PERSONALIZACION ══ -->
<div id="tab-custom" class="screen">
  <div class="conf-header">
    <div class="conf-title">🎨 Personalización</div>
    <div class="conf-sub">Cambia colores, textos y nombres del club. Los cambios se aplican en tiempo real.</div>
  </div>

  <!-- Temas especiales -->
  <div class="custom-section">
    <div class="custom-section-title">🎉 Temas de noche especial</div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:10px;margin-bottom:14px;">
      <button onclick="aplicarTema('default')" style="background:#111;border:1px solid #2a2a2a;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#888;font-family:'Rajdhani',sans-serif;font-weight:600;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#555'" onmouseout="this.style.borderColor='#2a2a2a'">
        <div style="font-size:28px;margin-bottom:6px;">⬛</div>DEFAULT
      </button>
      <button onclick="aplicarTema('jagger12')" style="background:#000;border:2px solid #333;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#fff;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#888'" onmouseout="this.style.borderColor='#333'">
        <div style="font-size:28px;margin-bottom:6px;">🥂</div>JAGGER 12 AÑOS
      </button>
      <button onclick="aplicarTema('jagger12boxeo')" style="background:linear-gradient(135deg,#0a0000,#1a0505);border:2px solid #4a0000;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#ff3333;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#880000'" onmouseout="this.style.borderColor='#4a0000'">
        <div style="font-size:28px;margin-bottom:6px;">🥊</div>VELADA BOXEO
      </button>
      <button onclick="aplicarTema('touchofpink')" style="background:linear-gradient(135deg,#140010,#2a0020);border:2px solid #9d174d;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#f472b6;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#f472b6'" onmouseout="this.style.borderColor='#9d174d'">
        <div style="font-size:28px;margin-bottom:6px;">🌸</div>TURNS PINK
      </button>
    </div>
    <!-- Toggle decoraciones para temas que las tienen -->
    <div id="tema-deco-toggle" style="display:none;background:#0a0a0a;border:1px solid #222;border-radius:8px;padding:12px 16px;margin-top:8px;align-items:center;gap:12px;flex-wrap:wrap;">
      <label id="deco-main-label" style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
        <input type="checkbox" id="toggle-deco" onchange="toggleDecoActual(this.checked)" style="width:16px;height:16px;accent-color:#c9a227;" checked />
        <span id="toggle-deco-label">Activar decoraciones animadas</span>
      </label>
      <div id="show-12-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-12" onchange="mostrar12Fondo=this.checked;reiniciarDeco12();showToast(this.checked?'12 de fondo activado':'12 de fondo desactivado');" style="width:16px;height:16px;accent-color:#c9a227;" checked />
          <span>Mostrar "12" de fondo</span>
        </label>
      </div>

      <div id="falling-gloves-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-falling-gloves" onchange="fallingGlovesActivos=this.checked;if(!this.checked){const w=document.getElementById('boxing-particles');if(w){[...w.children].filter(el=>el.textContent==='🥊').forEach(g=>g.remove());}}else{iniciarLluviaGuantes();}showToast(this.checked?'🥊 Guantes cayendo activados':'Guantes cayendo desactivados');" style="width:16px;height:16px;accent-color:#ff2222;" checked />
          <span>🥊 Guantes cayendo (lluvia de guantes)</span>
        </label>
      </div>

      <div id="pink-petalos-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="checkbox" id="toggle-pink-petalos" onchange="pinkPetalosActivos=this.checked;if(this.checked){iniciarPetalos();}else{const w=document.getElementById('petalos-wrap');if(w)w.innerHTML='';}" style="width:16px;height:16px;accent-color:#f472b6;" checked />
          <span>🌸 Pétalos cayendo</span>
        </label>
      </div>
      <div id="pink-modo-toggle" style="display:none;flex-direction:column;gap:6px;">
        <div style="font-family:'Rajdhani',sans-serif;font-size:11px;color:#cc88bb;letter-spacing:1px;text-transform:uppercase;margin-bottom:2px;">Fondo</div>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="radio" name="pink-modo" value="oscuro" checked style="accent-color:#f472b6;" onchange="document.body.classList.remove('pink-claro');pinkModoClaro=false;" />
          <span>🌙 Rosa oscuro</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="radio" name="pink-modo" value="claro" style="accent-color:#f472b6;" onchange="document.body.classList.add('pink-claro');pinkModoClaro=true;" />
          <span>🌸 Rosa medio</span>
        </label>
      </div>
      <div id="ko-anim-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-ko-anim" onchange="koAnimActiva=this.checked;showToast(this.checked?'🥊 Animación KO activada':'Animación KO desactivada');" style="width:16px;height:16px;accent-color:#ff2222;" checked />
          <span>🥊 Animación KO al cambiar de posición</span>
        </label>
      </div>
    </div>
  </div>

  <!-- Efectos de fondo -->
  <div class="custom-section">
    <div class="custom-section-title">✨ Efectos de fondo (modo presentación)</div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:8px;margin-bottom:10px;" id="efectos-grid">
      <button onclick="aplicarEfecto('ninguno')" id="efecto-btn-ninguno" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;">
        <div style="font-size:22px;margin-bottom:4px;">⬛</div>NINGUNO
      </button>
      <button onclick="aplicarEfecto('burbujas')" id="efecto-btn-burbujas" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;">
        <div style="font-size:22px;margin-bottom:4px;">🫧</div>BURBUJAS
      </button>
      <button onclick="aplicarEfecto('estrellas')" id="efecto-btn-estrellas" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;">
        <div style="font-size:22px;margin-bottom:4px;">⭐</div>ESTRELLAS
      </button>

    </div>
    <div class="preview-bar">
      <div class="preview-label">Vista previa del logo</div>
      <div class="preview-logo" id="prev-logo">RANKING <span class="vip-prev" id="prev-vip">VIP</span><span class="club-prev" id="prev-club">JAGGER CLUB</span></div>
    </div>
  </div>

  <!-- Textos -->
  <div class="custom-section">
    <div class="custom-section-title">Textos</div>
    <div class="custom-text-row">
      <span class="custom-text-label">Texto junto al nombre (ej: VIP, GOLD, PLUS...)</span>
      <input class="custom-text-input" id="ct-vip" type="text" placeholder="VIP" value="VIP" oninput="previewTextos()" />
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Mensaje ganador</span>
      <input class="custom-text-input" id="ct-winner-msg" type="text" placeholder="¡EL GANADOR DE LA NOCHE!" value="¡EL GANADOR DE LA NOCHE!" />
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Subtítulo ganador</span>
      <input class="custom-text-input" id="ct-winner-sub" type="text" placeholder="Ej: ¡Se lleva la botella!" value="" />
    </div>
    <div style="border-top:1px solid #1a1a1a;margin-top:14px;padding-top:14px;">
      <div style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Texto debajo de "RANKING VIP"</div>
      <div class="custom-text-row">
        <span class="custom-text-label">Texto</span>
        <input class="custom-text-input" id="ct-tagline" type="text" placeholder="JAGGER CLUB" value="JAGGER CLUB" oninput="previewTagline()" />
      </div>
      <div class="custom-text-row" style="align-items:center;">
        <span class="custom-text-label">Color</span>
        <input type="color" id="ct-tagline-color" value="#555555" oninput="previewTagline()" style="width:38px;height:28px;border:none;background:none;cursor:pointer;padding:0;flex-shrink:0;" />
      </div>
      <div class="custom-text-row" style="align-items:center;gap:8px;">
        <span class="custom-text-label">Brillo glow</span>
        <input type="range" id="ct-tagline-brightness" min="0" max="1" step="0.05" value="0" oninput="previewTagline();document.getElementById('ct-tagline-brightness-val').textContent=parseFloat(this.value).toFixed(2)" style="flex:1;" />
        <span id="ct-tagline-brightness-val" style="font-size:11px;color:#555;width:32px;text-align:right;flex-shrink:0;">0.00</span>
      </div>
      <div class="custom-text-row" style="align-items:center;">
        <span class="custom-text-label">Fuente</span>
        <select id="ct-tagline-font" onchange="previewTagline()" style="flex:1;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:6px;color:#f0ece0;padding:7px 8px;font-size:13px;outline:none;">
          <option value="'Rajdhani',sans-serif">Rajdhani</option>
          <option value="'Oswald',sans-serif">Oswald</option>
          <option value="Impact,sans-serif">Impact</option>
          <option value="Arial,sans-serif">Arial</option>
        </select>
      </div>
    </div>
  </div>

  <!-- Opciones del ganador -->
  <div class="custom-section">
    <div class="custom-section-title">🏆 Opciones del ganador</div>
    <div style="display:flex;flex-direction:column;gap:8px;">
      <div style="background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:14px 16px;">
        <div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#666;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Efecto de partículas</div>
        <div style="display:flex;flex-direction:column;gap:8px;">
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="confetti" checked style="accent-color:#c9a227;" onchange="tipoParticula='confetti';confettiGanadorActivo=true;" />
            <span>🎊 Confetti de colores</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="billetes" style="accent-color:#c9a227;" onchange="tipoParticula='billetes';confettiGanadorActivo=true;" />
            <span>💵 Lluvia de billetes</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="ninguno" style="accent-color:#c9a227;" onchange="tipoParticula='ninguno';confettiGanadorActivo=false;" />
            <span>Sin partículas</span>
          </label>
        </div>
      </div>
    </div>
  </div>

  <!-- Colores -->
  <div class="custom-section">
    <div class="custom-section-title">Colores</div>
    <div class="color-grid" id="color-grid"></div>
  </div>

  <div style="display:flex;gap:10px;flex-wrap:wrap;">
    <button class="btn-custom-save" onclick="aplicarPersonalizacion()">✓ Aplicar cambios</button>
    <button class="btn-custom-reset" onclick="resetPersonalizacion()">Restaurar defaults</button>
  </div>

  <!-- Contraseña -->
  <div class="custom-section" style="margin-top:18px;">
    <div class="custom-section-title">🔒 Contraseña de acceso</div>
    <div style="background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:14px 16px;display:flex;flex-direction:column;gap:10px;">
      <input id="pwd-actual" type="password" placeholder="PIN actual (4 dígitos)" class="custom-text-input" maxlength="4" inputmode="numeric" pattern="[0-9]*" />
      <input id="pwd-nueva" type="password" placeholder="Nuevo PIN (4 dígitos)" class="custom-text-input" maxlength="4" inputmode="numeric" pattern="[0-9]*" />
      <input id="pwd-confirm" type="password" placeholder="Confirmar nuevo PIN" class="custom-text-input" maxlength="4" inputmode="numeric" pattern="[0-9]*" />
      <button onclick="cambiarPassword()" style="background:#c9a227;color:#000;border:none;border-radius:6px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;" onmouseover="this.style.background='#e8c84a'" onmouseout="this.style.background='#c9a227'">Cambiar contraseña</button>
      <div id="pwd-msg" style="font-family:'Rajdhani',sans-serif;font-size:13px;letter-spacing:1px;min-height:18px;"></div>
    </div>
  </div>
</div>

<!-- ══ PANEL STATS ══ -->
<div id="tab-stats" class="screen">
  <div class="stats-header" style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:10px;">
    <div>
      <div class="stats-title">Estadísticas de la noche</div>
      <div class="stats-sub" id="stats-sub">—</div>
    </div>
    <a href="/historial" target="_blank" style="background:#c9a227;color:#000;border:none;border-radius:7px;padding:10px 22px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:2px;cursor:pointer;text-decoration:none;white-space:nowrap;flex-shrink:0;margin-top:4px;">📊 VER HISTORIAL</a>
  </div>

  <!-- KPIs -->
  <div class="kpi-row">
    <div class="kpi-card">
      <div class="kpi-label">Total de la noche</div>
      <div class="kpi-val" id="kpi-total">$0</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Operaciones</div>
      <div class="kpi-val" id="kpi-ops">0</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Promedio por operación</div>
      <div class="kpi-val" id="kpi-avg">$0</div>
    </div>
  </div>

  <!-- Fila de gráficos -->
  <div class="charts-row">
    <div class="chart-box">
      <div class="chart-title">Facturación por caja</div>
      <canvas id="chart-cajas" height="180"></canvas>
      <div class="chart-legend" id="legend-cajas"></div>
    </div>
    <div class="chart-box">
      <div class="chart-title">Consumo por hora</div>
      <canvas id="chart-horas" height="180"></canvas>
    </div>
  </div>

  <!-- Top clientes -->
  <div class="stats-section-title">Top clientes de la noche</div>
  <div id="stats-top-clientes"></div>

  <!-- Detalle por caja -->
  <div class="stats-section-title">Detalle por caja</div>
  <div class="cajas-detail-row" id="cajas-detail"></div>
</div>

<!-- ══ OVERLAY GANADOR ══ -->
<div id="winner-overlay">
  <div class="winner-rays"></div>
  <div class="confetti-wrap" id="confetti-wrap"></div>
  <button class="winner-close" onclick="cerrarGanador()">✕</button>
  <div class="winner-content">
    <span class="winner-corona" id="winner-corona">👑</span>
    <div class="winner-titulo" id="winner-titulo">GANADOR DE LA NOCHE</div>
    <div class="winner-nombre" id="winner-nombre">—</div>
    <div class="winner-line"></div>
    <div class="winner-info-row">
      <div class="winner-info-block">
        <div class="winner-info-label">Mesa</div>
        <div class="winner-info-val" id="winner-mesa">—</div>
      </div>
      <div class="winner-info-block">
        <div class="winner-info-label">Total consumido</div>
        <div class="winner-info-val" id="winner-total">—</div>
      </div>
    </div>
    <div class="winner-mensaje" id="winner-mensaje"></div>
  </div>
</div>



<!-- ══ OVERLAY CARTEL ══ -->
<div id="cartel-overlay" style="display:none;position:fixed;inset:0;z-index:11000;background:rgba(0,0,0,0.98);flex-direction:column;align-items:center;justify-content:center;overflow:hidden;">
  <div id="cartel-tema-bg" style="position:absolute;inset:0;pointer-events:none;z-index:0;"></div>
  <!-- Efecto de rayos de fondo -->
  <div id="cartel-rays" style="position:absolute;inset:0;pointer-events:none;z-index:1;overflow:hidden;"></div>
  <button onclick="cerrarCartel()" style="position:fixed;top:20px;right:28px;z-index:11001;background:transparent;color:#333;border:none;font-size:28px;cursor:pointer;font-family:'Oswald',sans-serif;transition:color .2s;line-height:1;" onmouseover="this.style.color='#888'" onmouseout="this.style.color='#333'">✕</button>
  <div id="cartel-content" style="position:relative;z-index:2;text-align:center;max-width:90vw;width:100%;padding:0 24px;">
    <!-- Emoji animado grande -->
    <div id="cartel-emoji-big" style="font-size:110px;margin-bottom:6px;display:block;animation:bottleBounce 1.2s cubic-bezier(.36,.07,.19,.97) infinite,emojiGlow 2s ease-in-out infinite alternate;filter:drop-shadow(0 0 30px rgba(201,162,39,0.6));"></div>
    <div id="cartel-nombre-display" style="font-family:'Oswald',sans-serif;font-size:clamp(48px,8vw,100px);font-weight:700;color:#fff;letter-spacing:4px;text-transform:uppercase;line-height:1;text-shadow:0 0 40px rgba(201,162,39,.5);"></div>
    <div id="cartel-mesa-display" style="font-family:'Oswald',sans-serif;font-size:clamp(20px,3vw,36px);color:#888;letter-spacing:3px;margin-top:8px;"></div>
    <div style="height:2px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:18px auto;max-width:600px;width:80%;"></div>
    <div id="cartel-frase-display" style="font-family:'Rajdhani',sans-serif;font-size:clamp(26px,4.5vw,56px);font-weight:700;color:#fff;letter-spacing:2px;text-transform:uppercase;line-height:1.3;padding:18px 36px;border:2px solid var(--gold);border-radius:8px;background:rgba(201,162,39,0.1);max-width:85vw;display:inline-block;text-shadow:0 0 20px rgba(201,162,39,0.3);box-shadow:0 0 40px rgba(201,162,39,0.15),inset 0 0 40px rgba(201,162,39,0.05);"></div>
  </div>
</div>

<!-- ══ MODAL CARTEL EDITOR ══ -->
<div id="cartel-modal" style="display:none;position:fixed;inset:0;z-index:10500;background:rgba(0,0,0,0.92);align-items:center;justify-content:center;">
  <div style="background:#111;border:1px solid #2a2a2a;border-radius:14px;padding:28px 32px;width:min(480px,92vw);max-height:90vh;overflow-y:auto;">
    <div style="font-family:'Oswald',sans-serif;font-size:22px;color:#fff;letter-spacing:1px;margin-bottom:20px;">📣 Configurar Cartel</div>
    <div style="margin-bottom:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Nombre (opcional)</label>
      <input id="cartel-nombre" type="text" placeholder="Ej: MATI" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:600;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Mesa (opcional)</label>
      <input id="cartel-mesa" type="text" placeholder="Ej: 5" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:16px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Frase del cartel</label>
      <input id="cartel-frase" type="text" placeholder="Ej: SACÓ UN NUVO CON BENGALAS" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:18px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:5px;">Emojis <span style="color:#333;font-size:9px;letter-spacing:1px;">(hasta 3 · vacío = auto)</span></label>
      <div style="margin-bottom:8px;display:flex;gap:8px;align-items:center;">
        <input id="cartel-emoji-input" type="text" placeholder="🍾" maxlength="24" style="width:120px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:10px 10px;font-size:22px;text-align:center;outline:none;letter-spacing:4px;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" oninput="limitarEmojis(this)" />
        <button onclick="document.getElementById('cartel-emoji-input').value=''" style="background:none;border:1px solid #222;border-radius:5px;color:#555;padding:8px 10px;cursor:pointer;font-size:13px;font-family:'Rajdhani',sans-serif;" title="Limpiar">✕</button>
      </div>
      <div style="display:flex;gap:6px;flex-wrap:wrap;">
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🍾</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🎆</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🥊</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🔥</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">💎</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🎉</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">⭐</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🏆</span>
      </div>
    </div>
    <div style="display:flex;gap:10px;flex-wrap:wrap;">
      <button onclick="mostrarCartel()" style="flex:1;background:#c9a227;color:#000;border:none;border-radius:7px;padding:13px 20px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;" onmouseover="this.style.background='#e8c84a'" onmouseout="this.style.background='#c9a227'">📣 MOSTRAR</button>
      <button onclick="cerrarCartelModal()" style="background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:13px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#222';this.style.color='#555'">Cancelar</button>
    </div>
  </div>
</div>

<div class="status-bar">
  <span id="status-txt">Conectando...</span>
  <span id="last-update"></span>
</div>
<div class="toast" id="toast"></div>

<!-- ══ MODAL EDITAR TX ══ -->
<div id="edit-tx-modal" style="display:none;position:fixed;inset:0;z-index:10800;background:rgba(0,0,0,0.88);align-items:center;justify-content:center;">
  <div style="background:#111;border:1px solid #2a2a2a;border-radius:12px;padding:28px 32px;width:min(400px,92vw);">
    <div style="font-family:'Oswald',sans-serif;font-size:18px;color:#fff;letter-spacing:1px;margin-bottom:18px;">✎ Editar operación</div>
    <input type="hidden" id="edit-tx-id" />
    <label class="field-label">Nombre</label>
    <input class="field-input" id="edit-tx-name" type="text" autocomplete="off" />
    <label class="field-label">Monto ($)</label>
    <input class="field-input amount-input" id="edit-tx-amount" type="number" min="1" step="100" />
    <div style="display:flex;gap:10px;margin-top:6px;">
      <button onclick="confirmarEditTx()" style="flex:1;background:#c9a227;color:#000;border:none;border-radius:7px;padding:12px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;">Guardar</button>
      <button onclick="document.getElementById('edit-tx-modal').style.display='none'" style="background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;">Cancelar</button>
    </div>
  </div>
</div>

<!-- ══ MODAL LOGIN ══ -->
<div id="login-modal" style="display:flex;position:fixed;inset:0;z-index:99999;background:#000;align-items:center;justify-content:center;flex-direction:column;">
  <div style="background:#111;border:1px solid #2a2a2a;border-radius:14px;padding:36px 32px;width:min(340px,90vw);text-align:center;">
    <div style="font-family:'Oswald',sans-serif;font-size:30px;color:#c9a227;letter-spacing:4px;margin-bottom:4px;">RANKING VIP</div>
    <div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#333;letter-spacing:4px;text-transform:uppercase;margin-bottom:28px;">JAGGER CLUB</div>
    <!-- PIN dots -->
    <div style="display:flex;justify-content:center;gap:14px;margin-bottom:24px;">
      <div class="pin-dot" id="pin-d0"></div>
      <div class="pin-dot" id="pin-d1"></div>
      <div class="pin-dot" id="pin-d2"></div>
      <div class="pin-dot" id="pin-d3"></div>
    </div>
    <div id="login-error" style="color:#a83030;font-family:'Rajdhani',sans-serif;font-size:13px;letter-spacing:1px;margin-bottom:16px;min-height:18px;"></div>
    <!-- Keypad -->
    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;">
      <button class="pin-key" onclick="pinPress('1')">1</button>
      <button class="pin-key" onclick="pinPress('2')">2</button>
      <button class="pin-key" onclick="pinPress('3')">3</button>
      <button class="pin-key" onclick="pinPress('4')">4</button>
      <button class="pin-key" onclick="pinPress('5')">5</button>
      <button class="pin-key" onclick="pinPress('6')">6</button>
      <button class="pin-key" onclick="pinPress('7')">7</button>
      <button class="pin-key" onclick="pinPress('8')">8</button>
      <button class="pin-key" onclick="pinPress('9')">9</button>
      <button class="pin-key" onclick="pinBack()" style="color:#888;">⌫</button>
      <button class="pin-key" onclick="pinPress('0')">0</button>
      <button class="pin-key" onclick="checkLogin()" style="background:#1a1a1a;color:#c9a227;">✓</button>
    </div>
  </div>
</div>
<style>
.pin-dot{width:16px;height:16px;border-radius:50%;border:2px solid #333;background:transparent;transition:all .15s;}
.pin-dot.filled{background:#c9a227;border-color:#c9a227;}
.pin-key{background:#0d0d0d;border:1px solid #2a2a2a;border-radius:9px;color:#f0ece0;padding:16px 10px;font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:600;cursor:pointer;transition:all .1s;letter-spacing:0;}
.pin-key:hover{background:#1a1a1a;border-color:#c9a227;color:#c9a227;}
.pin-key:active{transform:scale(.93);}
</style>

<script>
let txData = [];
let tarjetasData = {};
let confTarjetas = Array.from({length:30}, (_,i) => ({slot:i+1, codigo:'', saldo_inicial:''}));
let scanSlotActivo = null;
let cajaFocus = 0;
let globalBuffer = '';
let globalTimeout = null;
let horaFin = '05:30';
let ganadorMostrado = false;

// ══════════════════════════════════════════
//  PERSONALIZACION
// ══════════════════════════════════════════
const COLOR_DEFS = [
  {key:'--gold',       label:'Dorado principal — logo VIP, números del ranking, totales, botones principales y barra del ganador', default:'#c9a227'},
  {key:'--gold-light', label:'Dorado claro — color al pasar el mouse por encima de botones dorados', default:'#e8c84a'},
  {key:'--gold-dim',   label:'Dorado oscuro — títulos de sección, línea decorativa del header, texto secundario dorado', default:'#7a6010'},
  {key:'--black',      label:'Fondo principal — color de fondo de TODA la pantalla', default:'#080808'},
  {key:'--surface',    label:'Superficies — fondo de paneles, tarjetas, filas del ranking y formularios', default:'#111111'},
  {key:'--border',     label:'Bordes — líneas que rodean los paneles, tarjetas y separadores', default:'#2a2a2a'},
  {key:'--text',       label:'Texto principal — todo el texto de contenido, etiquetas e instrucciones', default:'#f0ece0'},
  {key:'--text-dim',   label:'Texto secundario — hints, subtítulos, labels de campos y texto apagado', default:'#555555'},
  {key:'--white',      label:'Blanco — nombres de clientes en el ranking y encabezados principales', default:'#ffffff'},
  {key:'--green',      label:'Verde — saldo disponible de tarjetas e indicador de conexión activa', default:'#2ecc71'},
  {key:'--danger',     label:'Rojo — errores, saldo insuficiente, botones de eliminar y resetear', default:'#a83030'},
];

let customColors = {};
COLOR_DEFS.forEach(c => customColors[c.key] = c.default);

// ══════════════════════════════════════════
//  TEMAS FESTIVOS
// ══════════════════════════════════════════
const TEMAS = {
  default: {
    colors: {}, // sin cambios, usa defaults
    overlay: '',
    bodyClass: ''
  },
  fullblack: {
    colors: {'--black':'#000000','--surface':'#080808','--border':'#1a1a1a','--gold':'#c9a227','--text':'#cccccc'},
    overlay: '',
    bodyClass: 'tema-fullblack'
  },
  navidad: {
    colors: {'--black':'#050f05','--surface':'#091209','--border':'#1a3a1a','--gold':'#e8c84a','--text':'#f0f0e0','--white':'#ffffff'},
    bodyClass: 'tema-navidad',
    overlay: `
      <!-- Estrellas de nieve -->
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;">
        <defs>
          <radialGradient id="vign" cx="50%" cy="50%" r="70%">
            <stop offset="0%" stop-color="transparent"/>
            <stop offset="100%" stop-color="#020a02" stop-opacity="0.7"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#vign)"/>
      </svg>
      <!-- Copos de nieve animados -->
      <div id="copos-wrap" style="position:absolute;inset:0;overflow:hidden;"></div>
      <!-- Árbol esquina izquierda -->
      <svg width="180" height="280" viewBox="0 0 180 280" style="position:absolute;bottom:0;left:0;opacity:.18;" xmlns="http://www.w3.org/2000/svg">
        <polygon points="90,10 160,120 120,120 150,200 100,200 110,260 80,260 70,200 30,200 60,120 20,120" fill="#2d7a2d"/>
        <rect x="75" y="255" width="30" height="25" fill="#5c3a1a" rx="3"/>
        <circle cx="90" cy="8" r="8" fill="#ffd700"/>
        <circle cx="55" cy="100" r="5" fill="#ff4444"/><circle cx="120" cy="130" r="5" fill="#4488ff"/>
        <circle cx="75" cy="160" r="5" fill="#ffaa00"/><circle cx="105" cy="180" r="5" fill="#ff4444"/>
        <circle cx="60" cy="195" r="4" fill="#4488ff"/><circle cx="130" cy="155" r="4" fill="#ffaa00"/>
        <circle cx="45" cy="130" r="4" fill="#ff4444"/><circle cx="90" cy="90" r="4" fill="#4488ff"/>
      </svg>
      <!-- Árbol esquina derecha -->
      <svg width="140" height="220" viewBox="0 0 180 280" style="position:absolute;bottom:0;right:0;opacity:.14;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
        <polygon points="90,10 160,120 120,120 150,200 100,200 110,260 80,260 70,200 30,200 60,120 20,120" fill="#2d7a2d"/>
        <rect x="75" y="255" width="30" height="25" fill="#5c3a1a" rx="3"/>
        <circle cx="90" cy="8" r="8" fill="#ffd700"/>
        <circle cx="55" cy="100" r="5" fill="#ff4444"/><circle cx="120" cy="130" r="5" fill="#4488ff"/>
        <circle cx="75" cy="160" r="5" fill="#ffaa00"/><circle cx="105" cy="180" r="5" fill="#ff4444"/>
      </svg>
      <!-- Guirnalda superior -->
      <svg width="100%" height="60" style="position:absolute;top:0;left:0;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="none">
        <path d="M0,15 Q12,28 24,15 Q36,2 48,15 Q60,28 72,15 Q84,2 96,15 Q108,28 120,15 Q132,2 144,15 Q156,28 168,15 Q180,2 192,15 Q204,28 216,15 Q228,2 240,15 Q252,28 264,15 Q276,2 288,15 Q300,28 312,15 Q324,2 336,15 Q348,28 360,15 Q372,2 384,15 Q396,28 408,15 Q420,2 432,15 Q444,28 456,15 Q468,2 480,15 Q492,28 504,15 Q516,2 528,15 Q540,28 552,15 Q564,2 576,15 Q588,28 600,15" stroke="#2d7a2d" stroke-width="4" fill="none" opacity="0.5"/>
        <circle cx="24" cy="15" r="4" fill="#ff4444" opacity="0.7"/><circle cx="72" cy="15" r="4" fill="#ffd700" opacity="0.7"/>
        <circle cx="120" cy="15" r="4" fill="#4488ff" opacity="0.7"/><circle cx="168" cy="15" r="4" fill="#ff4444" opacity="0.7"/>
        <circle cx="216" cy="15" r="4" fill="#ffd700" opacity="0.7"/><circle cx="264" cy="15" r="4" fill="#4488ff" opacity="0.7"/>
        <circle cx="312" cy="15" r="4" fill="#ff4444" opacity="0.7"/><circle cx="360" cy="15" r="4" fill="#ffd700" opacity="0.7"/>
        <circle cx="408" cy="15" r="4" fill="#4488ff" opacity="0.7"/><circle cx="456" cy="15" r="4" fill="#ff4444" opacity="0.7"/>
        <circle cx="504" cy="15" r="4" fill="#ffd700" opacity="0.7"/><circle cx="552" cy="15" r="4" fill="#4488ff" opacity="0.7"/>
      </svg>`
  },
  anonuevo: {
    colors: {'--black':'#000510','--surface':'#050a18','--border':'#1a1a4a','--gold':'#ffe066','--text':'#e8e8ff','--white':'#ffffff'},
    bodyClass: 'tema-anonuevo',
    overlay: `
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;">
        <defs>
          <radialGradient id="sky" cx="50%" cy="30%" r="80%">
            <stop offset="0%" stop-color="#050a28" stop-opacity="0.6"/>
            <stop offset="100%" stop-color="#000510" stop-opacity="0.9"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#sky)"/>
        <!-- Estrellas fijas -->
        <circle cx="5%" cy="8%" r="1.5" fill="white" opacity="0.8"/>
        <circle cx="12%" cy="20%" r="1" fill="white" opacity="0.6"/>
        <circle cx="20%" cy="5%" r="2" fill="white" opacity="0.9"/>
        <circle cx="32%" cy="15%" r="1" fill="white" opacity="0.7"/>
        <circle cx="45%" cy="3%" r="1.5" fill="white" opacity="0.8"/>
        <circle cx="55%" cy="18%" r="1" fill="white" opacity="0.6"/>
        <circle cx="67%" cy="7%" r="2" fill="white" opacity="0.9"/>
        <circle cx="78%" cy="12%" r="1" fill="white" opacity="0.7"/>
        <circle cx="88%" cy="4%" r="1.5" fill="white" opacity="0.8"/>
        <circle cx="95%" cy="22%" r="1" fill="white" opacity="0.6"/>
        <circle cx="15%" cy="35%" r="1" fill="#ffe066" opacity="0.5"/>
        <circle cx="40%" cy="28%" r="1.5" fill="#ffe066" opacity="0.4"/>
        <circle cx="72%" cy="30%" r="1" fill="#ffe066" opacity="0.5"/>
        <circle cx="90%" cy="38%" r="1.5" fill="#ffe066" opacity="0.4"/>
      </svg>
      <!-- Fuegos artificiales animados -->
      <div id="fuegos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
      <!-- Luna / reloj medianoche -->
      <svg width="90" height="90" viewBox="0 0 90 90" style="position:absolute;top:20px;right:30px;opacity:.25;" xmlns="http://www.w3.org/2000/svg">
        <circle cx="45" cy="45" r="40" fill="none" stroke="#ffe066" stroke-width="2"/>
        <circle cx="45" cy="45" r="35" fill="none" stroke="#ffe066" stroke-width="0.5" opacity="0.5"/>
        <line x1="45" y1="10" x2="45" y2="45" stroke="#ffe066" stroke-width="2" stroke-linecap="round"/>
        <line x1="45" y1="45" x2="70" y2="45" stroke="#ffe066" stroke-width="1.5" stroke-linecap="round"/>
        <circle cx="45" cy="45" r="3" fill="#ffe066"/>
        <text x="45" y="78" text-anchor="middle" font-size="8" fill="#ffe066" font-family="Oswald">FELIZ AÑO</text>
      </svg>`
  },
  touchofpink: {
    colors: {
      '--black':      '#080005',
      '--surface':    '#140010',
      '--border':     '#3d1035',
      '--gold':       '#f472b6',
      '--gold-light': '#fbb6ce',
      '--gold-dim':   '#9d174d',
      '--text':       '#ffe0f0',
      '--white':      '#ffffff'
    },
    bodyClass: 'tema-touchofpink',
    overlay: `
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;pointer-events:none;">
        <defs>
          <radialGradient id="pinkglow" cx="50%" cy="50%" r="70%">
            <stop offset="0%" stop-color="#3d0028" stop-opacity="0.45"/>
            <stop offset="100%" stop-color="#080005" stop-opacity="0.95"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#pinkglow)"/>
      </svg>
      <!-- Rosa izquierda -->
      <svg width="220" height="230" viewBox="0 0 220 230" style="position:absolute;bottom:0;left:0;opacity:.22;pointer-events:none;" xmlns="http://www.w3.org/2000/svg">
        <g transform="translate(75,155)">
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(0)"   fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(45)"  fill="#ec4899"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(90)"  fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(135)" fill="#ec4899"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(180)" fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(225)" fill="#ec4899"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(270)" fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(315)" fill="#ec4899"/>
          <circle cx="0" cy="0" r="16" fill="#fbb6ce"/>
        </g>
        <line x1="75" y1="155" x2="50" y2="230" stroke="#4a7a40" stroke-width="3" opacity="0.5"/>
        <ellipse cx="64" cy="196" rx="14" ry="7" fill="#4a7a40" transform="rotate(-30,64,196)" opacity="0.4"/>
      </svg>
      <!-- Rosa derecha (más pequeña) -->
      <svg width="170" height="190" viewBox="0 0 220 230" style="position:absolute;bottom:0;right:0;opacity:.16;pointer-events:none;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
        <g transform="translate(75,170)">
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(0)"   fill="#fbb6ce"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(60)"  fill="#f9a8d4"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(120)" fill="#fbb6ce"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(180)" fill="#f9a8d4"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(240)" fill="#fbb6ce"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(300)" fill="#f9a8d4"/>
          <circle cx="0" cy="0" r="13" fill="#ffe0f0"/>
        </g>
        <line x1="75" y1="170" x2="95" y2="230" stroke="#4a7a40" stroke-width="2.5" opacity="0.4"/>
      </svg>
      <!-- Dress code watermark -->
      <div style="position:absolute;bottom:16px;left:50%;transform:translateX(-50%);font-family:'Oswald',sans-serif;font-size:11px;letter-spacing:7px;color:rgba(244,114,182,0.3);text-transform:uppercase;white-space:nowrap;pointer-events:none;">DRESS CODE · TURNS PINK</div>
      <!-- Pétalos animados -->
      <div id="petalos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>`
  },
  halloween: {
    colors: {'--black':'#050200','--surface':'#0a0500','--border':'#3a1a00','--gold':'#ff8c00','--gold-light':'#ffaa00','--gold-dim':'#7a4000','--text':'#f0d0a0','--white':'#fff0e0'},
    bodyClass: 'tema-halloween',
    overlay: `
      <!-- Niebla de fondo -->
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;">
        <defs>
          <radialGradient id="fog" cx="50%" cy="100%" r="80%">
            <stop offset="0%" stop-color="#1a0a00" stop-opacity="0.5"/>
            <stop offset="100%" stop-color="#050200" stop-opacity="0.9"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#fog)"/>
        <!-- Luna llena -->
        <circle cx="85%" cy="12%" r="55" fill="#ff8c00" opacity="0.08"/>
        <circle cx="85%" cy="12%" r="48" fill="#ffaa00" opacity="0.06"/>
        <circle cx="85%" cy="12%" r="40" fill="#fff0c0" opacity="0.1"/>
      </svg>
      <!-- Telarañas SVG esquinas -->
      <svg width="260" height="200" viewBox="0 0 260 200" style="position:absolute;top:0;left:0;opacity:.35;" xmlns="http://www.w3.org/2000/svg">
        <g stroke="#888" stroke-width="0.8" fill="none" opacity="0.9">
          <!-- Radio lines from top-left corner -->
          <line x1="0" y1="0" x2="180" y2="0"/>
          <line x1="0" y1="0" x2="150" y2="50"/>
          <line x1="0" y1="0" x2="110" y2="90"/>
          <line x1="0" y1="0" x2="60" y2="130"/>
          <line x1="0" y1="0" x2="0" y2="180"/>
          <!-- Arcos concéntricos -->
          <path d="M40,0 Q20,20 0,40"/>
          <path d="M90,0 Q55,35 20,55 Q0,65 0,90"/>
          <path d="M140,0 Q95,45 55,75 Q25,100 0,140"/>
          <path d="M190,0 Q140,50 95,95 Q50,135 0,175"/>
        </g>
        <!-- Araña -->
        <g transform="translate(62,62)">
          <circle cx="0" cy="0" r="8" fill="#222" stroke="#555" stroke-width="0.5"/>
          <circle cx="0" cy="0" r="4" fill="#111"/>
          <circle cx="3" cy="-2" r="1.5" fill="#ff0000" opacity="0.6"/>
          <circle cx="-1" cy="-2" r="1.5" fill="#ff0000" opacity="0.6"/>
          <!-- Patas -->
          <line x1="-8" y1="-3" x2="-18" y2="-10" stroke="#444" stroke-width="1"/>
          <line x1="-8" y1="0" x2="-18" y2="0" stroke="#444" stroke-width="1"/>
          <line x1="-8" y1="3" x2="-18" y2="10" stroke="#444" stroke-width="1"/>
          <line x1="8" y1="-3" x2="18" y2="-10" stroke="#444" stroke-width="1"/>
          <line x1="8" y1="0" x2="18" y2="0" stroke="#444" stroke-width="1"/>
          <line x1="8" y1="3" x2="18" y2="10" stroke="#444" stroke-width="1"/>
          <!-- Hilo -->
          <line x1="0" y1="-8" x2="0" y2="-40" stroke="#555" stroke-width="0.6"/>
        </g>
      </svg>
      <!-- Telaraña esquina derecha (espejada) -->
      <svg width="220" height="180" viewBox="0 0 260 200" style="position:absolute;top:0;right:0;opacity:.3;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
        <g stroke="#777" stroke-width="0.8" fill="none">
          <line x1="0" y1="0" x2="180" y2="0"/>
          <line x1="0" y1="0" x2="150" y2="50"/>
          <line x1="0" y1="0" x2="110" y2="90"/>
          <line x1="0" y1="0" x2="60" y2="130"/>
          <line x1="0" y1="0" x2="0" y2="180"/>
          <path d="M40,0 Q20,20 0,40"/>
          <path d="M90,0 Q55,35 20,55 Q0,65 0,90"/>
          <path d="M140,0 Q95,45 55,75 Q25,100 0,140"/>
        </g>
      </svg>
      <!-- Murciélagos animados -->
      <div id="murcielagos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
      <!-- Calabazas decorativas abajo -->
      <svg width="100%" height="80" style="position:absolute;bottom:0;left:0;" xmlns="http://www.w3.org/2000/svg" viewBox="0 0 800 80" preserveAspectRatio="xMidYMax meet">
        <!-- Calabaza izq -->
        <g transform="translate(30,10)" opacity="0.25">
          <ellipse cx="25" cy="40" rx="22" ry="28" fill="#c84b00"/>
          <ellipse cx="25" cy="40" rx="16" ry="28" fill="#e05500"/>
          <ellipse cx="25" cy="40" rx="10" ry="28" fill="#c84b00"/>
          <rect x="22" y="8" width="6" height="12" rx="3" fill="#3a6a00"/>
          <!-- Cara -->
          <polygon points="17,32 14,38 20,38" fill="#050200"/>
          <polygon points="33,32 30,38 36,38" fill="#050200"/>
          <path d="M16,46 Q25,55 34,46" stroke="#050200" stroke-width="2" fill="none"/>
          <!-- Ojos brillan naranja -->
          <polygon points="17,32 14,38 20,38" fill="#ff8c00" opacity="0.5"/>
          <polygon points="33,32 30,38 36,38" fill="#ff8c00" opacity="0.5"/>
        </g>
        <!-- Calabaza derecha -->
        <g transform="translate(730,5)" opacity="0.2">
          <ellipse cx="25" cy="45" rx="26" ry="32" fill="#c84b00"/>
          <ellipse cx="25" cy="45" rx="18" ry="32" fill="#e05500"/>
          <ellipse cx="25" cy="45" rx="10" ry="32" fill="#c84b00"/>
          <rect x="22" y="10" width="6" height="14" rx="3" fill="#3a6a00"/>
          <polygon points="17,36 13,44 21,44" fill="#050200"/>
          <polygon points="33,36 29,44 37,44" fill="#050200"/>
          <path d="M15,52 Q25,63 35,52" stroke="#050200" stroke-width="2" fill="none"/>
        </g>
      </svg>`
  }
};

function aplicarTema(nombre) {
  const tema = TEMAS[nombre];
  if (!tema) return;

  // Limpiar clases anteriores
  document.body.classList.remove('tema-fullblack','tema-navidad','tema-anonuevo','tema-halloween','tema-touchofpink','pink-claro');
  pinkModoClaro = false;

  // Restaurar colores default primero
  if (nombre === 'default') {
    COLOR_DEFS.forEach(c => document.documentElement.style.setProperty(c.key, c.default));
    document.documentElement.style.setProperty('--surface-gold', '#0d0b00');
    customColors = {};
    COLOR_DEFS.forEach(c => customColors[c.key] = c.default);
  } else {
    // Aplicar colores del tema
    COLOR_DEFS.forEach(c => {
      const v = tema.colors[c.key] || c.default;
      document.documentElement.style.setProperty(c.key, v);
      customColors[c.key] = v;
    });
    if (tema.colors['--surface']) {
      document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(tema.colors['--surface']));
    }
    if (tema.bodyClass) document.body.classList.add(tema.bodyClass);
  }

  // Overlay decorativo
  const overlay = document.getElementById('tema-overlay');
  overlay.innerHTML = tema.overlay || '';
  overlay.style.opacity = tema.overlay ? '1' : '0';

  // Iniciar animaciones específicas
  if (nombre === 'navidad') iniciarCopos();
  if (nombre === 'anonuevo') iniciarFuegos();
  if (nombre === 'halloween') iniciarMurcielagos();
  if (nombre === 'touchofpink') iniciarPetalos();

  buildColorGrid();
  showToast('Tema ' + nombre.toUpperCase() + ' aplicado');

  try {
    localStorage.setItem('rankingVIP_tema', nombre);
  } catch(e) {}
}

function iniciarCopos() {
  const wrap = document.getElementById('copos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  const copoSVGs = ['❄','❅','❆','✻','✼'];
  for (let i = 0; i < 40; i++) {
    const el = document.createElement('div');
    const sym = copoSVGs[Math.floor(Math.random() * copoSVGs.length)];
    el.textContent = sym;
    el.style.cssText = `position:absolute;top:-30px;left:${Math.random()*100}vw;font-size:${10+Math.random()*18}px;color:rgba(255,255,255,${0.3+Math.random()*0.4});animation:copoFall ${5+Math.random()*8}s linear ${Math.random()*8}s infinite;pointer-events:none;`;
    wrap.appendChild(el);
  }
  // Agregar keyframes si no existen
  if (!document.getElementById('kf-copo')) {
    const s = document.createElement('style');
    s.id = 'kf-copo';
    s.textContent = '@keyframes copoFall{0%{transform:translateY(-30px) rotate(0deg);opacity:1}100%{transform:translateY(105vh) rotate(360deg);opacity:0}}';
    document.head.appendChild(s);
  }
}

function iniciarFuegos() {
  const wrap = document.getElementById('fuegos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  const colores = ['#ffe066','#ff4488','#44aaff','#ff8844','#aaffaa','#ff44ff'];
  function lanzarFuego() {
    if (!document.getElementById('fuegos-wrap')) return;
    const x = 10 + Math.random() * 80;
    const y = 5 + Math.random() * 50;
    const color = colores[Math.floor(Math.random() * colores.length)];
    const burst = document.createElement('div');
    burst.style.cssText = `position:absolute;left:${x}%;top:${y}%;pointer-events:none;`;
    for (let i = 0; i < 14; i++) {
      const spark = document.createElement('div');
      const angle = (i / 14) * 360;
      const dist = 30 + Math.random() * 40;
      spark.style.cssText = `position:absolute;width:3px;height:3px;border-radius:50%;background:${color};box-shadow:0 0 4px ${color};animation:spark ${0.8+Math.random()*0.4}s ease-out forwards;--dx:${Math.cos(angle*Math.PI/180)*dist}px;--dy:${Math.sin(angle*Math.PI/180)*dist}px;`;
      burst.appendChild(spark);
    }
    wrap.appendChild(burst);
    setTimeout(() => burst.remove(), 1400);
    setTimeout(lanzarFuego, 800 + Math.random() * 2000);
  }
  if (!document.getElementById('kf-spark')) {
    const s = document.createElement('style');
    s.id = 'kf-spark';
    s.textContent = '@keyframes spark{0%{opacity:1;transform:translate(0,0)}100%{opacity:0;transform:translate(var(--dx),var(--dy))}}';
    document.head.appendChild(s);
  }
  lanzarFuego(); lanzarFuego(); lanzarFuego();
}

function iniciarMurcielagos() {
  const wrap = document.getElementById('murcielagos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  if (!document.getElementById('kf-bat')) {
    const s = document.createElement('style');
    s.id = 'kf-bat';
    s.textContent = `@keyframes batFly{0%{transform:translateX(-80px) translateY(0)}100%{transform:translateX(110vw) translateY(var(--dy))}}
    @keyframes batWing{0%,100%{transform:scaleY(1)}50%{transform:scaleY(-0.3)}}`;
    document.head.appendChild(s);
  }
  function lanzarMurcie() {
    if (!document.getElementById('murcielagos-wrap')) return;
    const el = document.createElement('div');
    const y = 5 + Math.random() * 60;
    const dur = 6 + Math.random() * 8;
    const dy = (Math.random() - 0.5) * 200;
    const sz = 18 + Math.random() * 20;
    el.style.cssText = `position:absolute;top:${y}%;left:-80px;font-size:${sz}px;animation:batFly ${dur}s linear forwards;--dy:${dy}px;opacity:0.5;`;
    el.textContent = '🦇';
    wrap.appendChild(el);
    setTimeout(() => el.remove(), dur * 1000);
    setTimeout(lanzarMurcie, 2000 + Math.random() * 5000);
  }
  lanzarMurcie(); setTimeout(lanzarMurcie, 2000); setTimeout(lanzarMurcie, 4000);
}

function iniciarPetalos() {
  const wrap = document.getElementById('petalos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  const shapes = ['🌸','🌸','🌺','🌷','💮'];
  for (let i = 0; i < 35; i++) {
    const el = document.createElement('div');
    const sym = shapes[Math.floor(Math.random() * shapes.length)];
    el.textContent = sym;
    el.style.cssText = `position:absolute;top:-40px;left:${Math.random()*100}vw;font-size:${12+Math.random()*16}px;opacity:${0.25+Math.random()*0.45};animation:copoFall ${6+Math.random()*9}s linear ${Math.random()*8}s infinite;pointer-events:none;`;
    wrap.appendChild(el);
  }
  if (!document.getElementById('kf-copo')) {
    const s = document.createElement('style');
    s.id = 'kf-copo';
    s.textContent = '@keyframes copoFall{0%{transform:translateY(-40px) rotate(0deg);opacity:1}100%{transform:translateY(105vh) rotate(360deg);opacity:0}}';
    document.head.appendChild(s);
  }
}

// cargarTemaGuardado redefined below

function buildColorGrid() {
  const grid = document.getElementById('color-grid');
  grid.innerHTML = COLOR_DEFS.map(c => {
    const val = customColors[c.key];
    const hexId = 'hex-' + c.key.replace(/--/g,'').replace(/-/g,'_');
    const inputId = 'inp-' + c.key.replace(/--/g,'').replace(/-/g,'_');
    const swatchId = 'sw-' + c.key.replace(/--/g,'').replace(/-/g,'_');
    return `<div class="color-item">
      <div class="color-swatch" id="${swatchId}" style="background:${val}">
        <input type="color" id="${inputId}" value="${val}" data-key="${c.key}" oninput="onColorChange(this)" />
      </div>
      <div style="flex:1">
        <div class="color-label">${c.label}</div>
        <div class="color-hex" id="${hexId}">${val}</div>
      </div>
      <button onclick="resetColorSingle('${c.key}')" title="Restablecer este color" style="background:none;border:1px solid #2a2a2a;color:#444;border-radius:5px;padding:4px 8px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;letter-spacing:1px;transition:all .15s;white-space:nowrap;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#2a2a2a';this.style.color='#444'">↺ Default</button>
    </div>`;
  }).join('');
}

function resetColorSingle(key) {
  const def = COLOR_DEFS.find(c => c.key === key);
  if (!def) return;
  customColors[key] = def.default;
  document.documentElement.style.setProperty(key, def.default);
  // Update swatch bg
  const swatchId = 'sw-' + key.replace(/--/g,'').replace(/-/g,'_');
  const sw = document.getElementById(swatchId);
  if (sw) sw.style.background = def.default;
  // Update input value
  const inputId = 'inp-' + key.replace(/--/g,'').replace(/-/g,'_');
  const inp = document.getElementById(inputId);
  if (inp) inp.value = def.default;
  // Update hex label
  const hexId = 'hex-' + key.replace(/--/g,'').replace(/-/g,'_');
  const hexEl = document.getElementById(hexId);
  if (hexEl) hexEl.textContent = def.default;
  // Special: update preview gold
  if (key === '--gold') document.getElementById('prev-vip').style.color = def.default;
  if (key === '--surface') {
    const goldTint = blendSurfaceGold(def.default);
    document.documentElement.style.setProperty('--surface-gold', goldTint);
  }
  showToast('Color restablecido al default');
}

function blendSurfaceGold(surfaceHex) {
  // Mix the surface color with a subtle golden tint for rank-1
  try {
    const r = parseInt(surfaceHex.slice(1,3),16);
    const g = parseInt(surfaceHex.slice(3,5),16);
    const b = parseInt(surfaceHex.slice(5,7),16);
    // Add slight golden warmth: shift toward gold (#c9a227)
    const nr = Math.min(255, Math.round(r * 0.85 + 0xc9 * 0.15));
    const ng = Math.min(255, Math.round(g * 0.88 + 0xa2 * 0.12));
    const nb = Math.min(255, Math.round(b * 0.95 + 0x27 * 0.05));
    return '#' + nr.toString(16).padStart(2,'0') + ng.toString(16).padStart(2,'0') + nb.toString(16).padStart(2,'0');
  } catch(e) { return surfaceHex; }
}

function onColorChange(el) {
  const key = el.dataset.key;
  const val = el.value;
  customColors[key] = val;
  el.parentElement.style.background = val;
  const hexId = 'hex-' + key.replace(/--/g,'').replace(/-/g,'_');
  const hexEl = document.getElementById(hexId);
  if (hexEl) hexEl.textContent = val;
  // Aplicar en tiempo real
  document.documentElement.style.setProperty(key, val);
  // When surface changes, also update the gold-tinted surface for rank-1
  if (key === '--surface') {
    const goldTint = blendSurfaceGold(val);
    document.documentElement.style.setProperty('--surface-gold', goldTint);
  }
  // Update preview gold
  if (key === '--gold') {
    document.getElementById('prev-vip').style.color = val;
  }
}

function previewTextos() {
  const vip  = document.getElementById('ct-vip').value  || 'VIP';
  document.getElementById('prev-vip').textContent  = vip;
  previewTagline();
}
function hexToRgb(hex){
  const h=hex.replace('#','');
  return {r:parseInt(h.slice(0,2),16),g:parseInt(h.slice(2,4),16),b:parseInt(h.slice(4,6),16)};
}
function buildTaglineShadow(color, glow) {
  const g=parseFloat(glow||0);
  if(g<=0) return 'none';
  const {r,g:gr,b}=hexToRgb(color||'#555555');
  return `0 0 14px rgba(${r},${gr},${b},${g}), 0 0 30px rgba(${r},${gr},${b},${(g*0.6).toFixed(2)}), 0 0 60px rgba(${r},${gr},${b},${(g*0.3).toFixed(2)})`;
}
function applyTaglineStyle(el, text, color, glow, font) {
  if (!el) return;
  if (temaActual === 'touchofpink') {
    const kids = el.querySelectorAll('div');
    kids.forEach(d => d.style.fontFamily = font);
    return;
  }
  el.textContent = text;
  el.style.color = color;
  el.style.filter = 'none';
  el.style.textShadow = buildTaglineShadow(color, glow);
  el.style.fontFamily = font;
  el.style.fontSize = '28px';
  el.style.fontWeight = '600';
  el.style.letterSpacing = '5px';
  el.style.textTransform = 'uppercase';
  el.style.textAlign = 'center';
}
function previewTagline() {
  const text = (document.getElementById('ct-tagline')||{}).value || 'JAGGER CLUB';
  const color = (document.getElementById('ct-tagline-color')||{}).value || '#555555';
  const glow = (document.getElementById('ct-tagline-brightness')||{}).value || '0';
  const font = (document.getElementById('ct-tagline-font')||{}).value || "'Rajdhani',sans-serif";
  const pv = document.getElementById('prev-club');
  if (pv) { pv.textContent=text; pv.style.color=color; pv.style.textShadow=buildTaglineShadow(color,glow); pv.style.fontFamily=font; }
  applyTaglineStyle(document.getElementById('tema-tagline'), text, color, glow, font);
}

function aplicarPersonalizacion() {
  // Colores
  COLOR_DEFS.forEach(c => {
    document.documentElement.style.setProperty(c.key, customColors[c.key]);
  });
  // Sync surface-gold tint
  if (customColors['--surface']) {
    document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(customColors['--surface']));
  }
  // Textos logo
  const vip  = document.getElementById('ct-vip').value  || 'VIP';
  document.getElementById('logo-vip').textContent  = vip;
  // Tagline
  const taglineText = document.getElementById('ct-tagline').value || 'JAGGER CLUB';
  const taglineColor = document.getElementById('ct-tagline-color').value;
  const taglineGlow = document.getElementById('ct-tagline-brightness').value;
  const taglineFont = document.getElementById('ct-tagline-font').value;
  applyTaglineStyle(document.getElementById('tema-tagline'), taglineText, taglineColor, taglineGlow, taglineFont);
  // Guardar en localStorage para persistir
  try {
    localStorage.setItem('rankingVIP_colors', JSON.stringify(customColors));
    localStorage.setItem('rankingVIP_vip', vip);
    const wm = document.getElementById('ct-winner-msg').value;
    const ws = document.getElementById('ct-winner-sub').value;
    localStorage.setItem('rankingVIP_wmsg', wm);
    localStorage.setItem('rankingVIP_wsub', ws);
    localStorage.setItem('rankingVIP_tagline_text', taglineText);
    localStorage.setItem('rankingVIP_tagline_color', taglineColor);
    localStorage.setItem('rankingVIP_tagline_glow', taglineGlow);
    localStorage.setItem('rankingVIP_tagline_font', taglineFont);
  } catch(e){}
  showToast('Personalización aplicada');
}

function resetPersonalizacion() {
  if (!confirm('Restaurar todos los colores y textos al default?')) return;
  COLOR_DEFS.forEach(c => {
    customColors[c.key] = c.default;
    document.documentElement.style.setProperty(c.key, c.default);
  });
  document.documentElement.style.setProperty('--surface-gold', '#0d0b00');
  document.getElementById('ct-vip').value  = 'VIP';
  document.getElementById('logo-vip').textContent  = 'VIP';
  document.getElementById('ct-tagline').value = 'JAGGER CLUB';
  document.getElementById('ct-tagline-color').value = '#555555';
  document.getElementById('ct-tagline-brightness').value = '0';
  document.getElementById('ct-tagline-brightness-val').textContent = '0.00';
  applyTaglineStyle(document.getElementById('tema-tagline'),'JAGGER CLUB','#555','0',"'Rajdhani',sans-serif");
  try { localStorage.removeItem('rankingVIP_colors'); localStorage.removeItem('rankingVIP_club'); localStorage.removeItem('rankingVIP_vip'); localStorage.removeItem('rankingVIP_tagline_text'); localStorage.removeItem('rankingVIP_tagline_color'); localStorage.removeItem('rankingVIP_tagline_glow'); localStorage.removeItem('rankingVIP_tagline_font'); } catch(e){}
  buildColorGrid();
  previewTextos();
  showToast('Colores restaurados');
}

function cargarPersonalizacionGuardada() {
  try {
    const savedColors = localStorage.getItem('rankingVIP_colors');
    if (savedColors) {
      customColors = JSON.parse(savedColors);
      COLOR_DEFS.forEach(c => {
        if (customColors[c.key]) document.documentElement.style.setProperty(c.key, customColors[c.key]);
      });
    }
    const club = localStorage.getItem('rankingVIP_club');
    const vip  = localStorage.getItem('rankingVIP_vip');
    if (vip)  { document.getElementById('logo-vip').textContent  = vip;  document.getElementById('ct-vip').value  = vip;  document.getElementById('prev-vip').textContent  = vip; }
    const wm = localStorage.getItem('rankingVIP_wmsg');
    const ws = localStorage.getItem('rankingVIP_wsub');
    if (wm) document.getElementById('ct-winner-msg').value = wm;
    if (ws) document.getElementById('ct-winner-sub').value = ws;
    const tlText  = localStorage.getItem('rankingVIP_tagline_text');
    const tlColor = localStorage.getItem('rankingVIP_tagline_color');
    const tlGlow  = localStorage.getItem('rankingVIP_tagline_glow');
    const tlFont  = localStorage.getItem('rankingVIP_tagline_font');
    if (tlText||tlColor||tlGlow||tlFont) {
      const inp    = document.getElementById('ct-tagline');
      const colInp = document.getElementById('ct-tagline-color');
      const brInp  = document.getElementById('ct-tagline-brightness');
      const brVal  = document.getElementById('ct-tagline-brightness-val');
      const fntInp = document.getElementById('ct-tagline-font');
      if (tlText  && inp)    inp.value    = tlText;
      if (tlColor && colInp) colInp.value = tlColor;
      if (tlGlow  && brInp)  { brInp.value=tlGlow; if(brVal) brVal.textContent=parseFloat(tlGlow).toFixed(2); }
      if (tlFont  && fntInp) fntInp.value = tlFont;
      applyTaglineStyle(
        document.getElementById('tema-tagline'),
        tlText  || 'JAGGER CLUB',
        tlColor || '#555555',
        tlGlow  || '0',
        tlFont  || "'Rajdhani',sans-serif"
      );
    }
  } catch(e){}
}

// ══════════════════════════════════════════
//  ANIMACION GANADOR
// ══════════════════════════════════════════
function generarParticulas(wrapId) {
  const wrap = document.getElementById(wrapId || 'confetti-wrap');
  wrap.innerHTML = '';
  if (tipoParticula === 'ninguno') return;
  if (tipoParticula === 'billetes') {
    for (let i = 0; i < 55; i++) {
      const el = document.createElement('span');
      el.style.position = 'absolute';
      el.style.top = '-60px';
      el.style.left = (Math.random() * 100) + 'vw';
      el.style.fontSize = (20 + Math.random() * 18) + 'px';
      el.style.animationName = 'confettiFall';
      el.style.animationDuration = (4 + Math.random() * 4) + 's';
      el.style.animationDelay = (Math.random() * 5) + 's';
      el.style.animationTimingFunction = 'linear';
      el.style.animationIterationCount = 'infinite';
      el.textContent = '💵';
      wrap.appendChild(el);
    }
  } else {
    const colors = temaActual === 'touchofpink'
      ? ['#f472b6','#fbb6ce','#ffffff','#f9a8d4','#ec4899','#ffffff','#fce7f3','#ff69b4','#fff0f5']
      : ['#c9a227','#e8c84a','#fff','#f0ece0','#2ecc71','#e74c3c','#3498db','#9b59b6','#ff9f43'];
    for (let i = 0; i < 120; i++) {
      const el = document.createElement('div');
      el.className = 'confetti-piece';
      el.style.left = Math.random() * 100 + 'vw';
      el.style.background = colors[Math.floor(Math.random() * colors.length)];
      el.style.width = (6 + Math.random() * 10) + 'px';
      el.style.height = (10 + Math.random() * 16) + 'px';
      el.style.animationDuration = (3 + Math.random() * 5) + 's';
      el.style.animationDelay = (Math.random() * 4) + 's';
      el.style.borderRadius = Math.random() > 0.5 ? '50%' : '2px';
      wrap.appendChild(el);
    }
  }
}


function mostrarGanador() {
  const totals = {}, mesas = {};
  txData.forEach(t => {
    totals[t.name] = (totals[t.name]||0) + t.amount;
    if (t.mesa && !mesas[t.name]) mesas[t.name] = t.mesa;
  });
  const nombres = Object.keys(totals).sort((a,b) => totals[b]-totals[a]);
  if (!nombres.length) { showToast('No hay consumos registrados aun', true); return; }

  const ganador = nombres[0];
  const mesa = mesas[ganador] || '—';
  const total = totals[ganador];

  const isBoxeo = temaActual === 'jagger12boxeo';
  const wmsg = (document.getElementById('ct-winner-msg').value || '¡EL GANADOR DE LA NOCHE!').toUpperCase();
  const wsub = document.getElementById('ct-winner-sub').value || '';
  const premio = document.getElementById('msg-input').value.trim();

  // Icono: boxeo = guantes + trofeo, default = corona
  const coronaEl = document.getElementById('winner-corona');
  if (coronaEl) coronaEl.textContent = isBoxeo ? '🥊🏆🥊' : '👑';

  const nombreEl = document.getElementById('winner-nombre');
  const totalEl  = document.getElementById('winner-total');

  document.getElementById('winner-titulo').textContent = wmsg;
  document.getElementById('winner-mesa').textContent = mesa;
  nombreEl.style.borderRight = 'none';
  nombreEl.textContent = ganador.toUpperCase();
  totalEl.textContent  = fmt(total);

  const msgEl = document.getElementById('winner-mensaje');
  const textoFinal = wsub || premio || '';
  msgEl.textContent = textoFinal;
  msgEl.style.display = textoFinal ? 'block' : 'none';

  if (confettiGanadorActivo) generarParticulas();
  document.getElementById('winner-overlay').classList.add('show');

}

function mostrarGanadorManual() {
  ganadorMostrado = true;
  mostrarGanador();
  fetch('/api/winner/show',{method:'POST'});
}

function cerrarGanador() {
  document.getElementById('winner-overlay').classList.remove('show');
  fetch('/api/winner/hide',{method:'POST'});
}

// ══════════════════════════════════════════
//  RELOJ Y DETECCION DE HORA LIMITE
// ══════════════════════════════════════════
function tickClock() {
  const now = new Date();
  const h = String(now.getHours()).padStart(2,'0');
  const m = String(now.getMinutes()).padStart(2,'0');
  const el = document.getElementById('clock-display');
  if (el) el.textContent = h + ':' + m;

  // Detectar hora fin
  if (horaFin && !ganadorMostrado && txData.length > 0) {
    const [fh, fm] = horaFin.split(':').map(Number);
    if (now.getHours() === fh && now.getMinutes() === fm && now.getSeconds() < 30) {
      ganadorMostrado = true;
      mostrarGanador();
      fetch('/api/winner/show',{method:'POST'});
    }
  }
}
setInterval(tickClock, 1000);
tickClock();



// ══════════════════════════════════════════
//  LECTOR DE TARJETA
// ══════════════════════════════════════════
let pendingCodigos = {};
let lastKeyTime = 0;
const READER_SPEED_MS = 80;

document.addEventListener('keydown', (e) => {
  ['caja1','caja2','caja3'].forEach((s,i) => {
    const tab = document.getElementById('tab-'+s);
    if (tab && tab.classList.contains('active')) cajaFocus = i+1;
  });
  const enConfig = document.getElementById('tab-config') && document.getElementById('tab-config').classList.contains('active');
  const active = document.activeElement;

  const now = Date.now();
  const timeSinceLast = now - lastKeyTime;
  const isReaderSpeed = globalBuffer.length > 0 && timeSinceLast < READER_SPEED_MS;

  if (active && active.tagName === 'INPUT' && !active.classList.contains('tc-input')) {
    if (!isReaderSpeed && e.key !== 'Enter') return;
  }

  if (e.key === 'Enter') {
    if (globalBuffer.length > 1) {
      const codigo = globalBuffer.trim();
      if (enConfig && scanSlotActivo !== null) {
        pendingCodigos[scanSlotActivo] = codigo;
        const btn = document.getElementById('scan-btn-'+scanSlotActivo);
        if (btn) { btn.textContent='LEER'; btn.classList.remove('activo'); }
        const codeEl = document.getElementById('tc-code-'+scanSlotActivo);
        if (codeEl) codeEl.textContent = 'Leida: ' + codigo;
        const confirmBtn = document.getElementById('confirm-btn-'+scanSlotActivo);
        if (confirmBtn) confirmBtn.style.display = 'inline-block';
        const slotNum = confTarjetas[scanSlotActivo] ? confTarjetas[scanSlotActivo].slot : '?';
        showToast('Tarjeta leida con exito para Mesa ' + slotNum + ' — presiona CONFIRMAR');
        scanSlotActivo = null;
      } else if (cajaFocus > 0) {
        procesarTarjetaEnCaja(codigo, cajaFocus);
      }
    }
    globalBuffer = '';
    lastKeyTime = 0;
    clearTimeout(globalTimeout);
  } else if (e.key.length === 1) {
    globalBuffer += e.key;
    lastKeyTime = now;
    clearTimeout(globalTimeout);
    globalTimeout = setTimeout(() => { globalBuffer = ''; lastKeyTime = 0; }, 500);
  }
});

function confirmarVinculo(idx) {
  const codigo = pendingCodigos[idx];
  if (!codigo) { showToast('Primero pasa la tarjeta por el lector', true); return; }
  // Verificar duplicado: misma tarjeta ya registrada en otro slot
  const dupIdx = confTarjetas.findIndex((t, i) => i !== idx && t.codigo === codigo);
  if (dupIdx !== -1) {
    showToast('⚠ Esta tarjeta ya está registrada en Mesa ' + confTarjetas[dupIdx].slot + '. No se puede vincular dos veces.', true);
    delete pendingCodigos[idx];
    const codeEl2 = document.getElementById('tc-code-'+idx);
    if (codeEl2) codeEl2.textContent = 'Presiona LEER y pasa la tarjeta';
    const confirmBtn2 = document.getElementById('confirm-btn-'+idx);
    if (confirmBtn2) confirmBtn2.style.display = 'none';
    const scanBtn2 = document.getElementById('scan-btn-'+idx);
    if (scanBtn2) { scanBtn2.textContent = 'LEER'; scanBtn2.classList.remove('activo'); }
    return;
  }
  confTarjetas[idx].codigo = codigo;
  delete pendingCodigos[idx];
  const codeEl = document.getElementById('tc-code-'+idx);
  if (codeEl) codeEl.textContent = 'Cod: ' + codigo;
  const tcEl = document.getElementById('tc-'+idx);
  if (tcEl) tcEl.classList.add('configurada');
  const confirmBtn = document.getElementById('confirm-btn-'+idx);
  if (confirmBtn) confirmBtn.style.display = 'none';
  const statusEl = document.getElementById('tc-status-'+idx);
  if (statusEl) { statusEl.textContent = 'Vinculada'; statusEl.style.color = '#3a9a5a'; }
  showToast('Tarjeta vinculada a Mesa ' + confTarjetas[idx].slot);
}

function procesarTarjetaEnCaja(codigo, caja) {
  const idx = confTarjetas.findIndex(t => t.codigo === codigo);
  if (idx === -1) { showToast('Tarjeta no configurada — vinculala primero en la pestana Tarjetas', true); return; }
  const conf = confTarjetas[idx];
  const mesa = String(conf.slot);
  const saldoInfo = tarjetasData[codigo];
  const saldoActual = (saldoInfo && saldoInfo.saldo_actual !== undefined) ? saldoInfo.saldo_actual : parseFloat(conf.saldo_inicial||0);
  const saldoInicial = parseFloat(conf.saldo_inicial||0);
  const nombre = saldoInfo ? (saldoInfo.nombre||conf.nombre_cliente||'') : (conf.nombre_cliente||'');
  window['cajaState'+caja] = { codigo, mesa, slot: conf.slot, saldo_inicial: saldoInicial, saldo_actual: saldoActual, nombre };
  setModo(caja, 'tarjeta');
  renderCajaInner(caja);
  showToast('Tarjeta leida con exito para Mesa ' + mesa + ' — Saldo: '+fmt(saldoActual));
  setTimeout(() => {
    const el = document.getElementById(nombre ? 'amount'+caja : 'name'+caja);
    if (el) el.focus();
  }, 100);
}

function setModo(caja, modo) {
  ['tarjeta','manual','recargar'].forEach(m => {
    const tab = document.getElementById('modo-tab-'+m+'-'+caja);
    const content = document.getElementById('modo-content-'+m+'-'+caja);
    if (tab) tab.classList.toggle('active', m===modo);
    if (content) content.classList.toggle('active', m===modo);
  });
}

function showTab(id) {
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.screen').forEach(s => s.classList.remove('active'));
  document.getElementById('tbtn-'+id).classList.add('active');
  document.getElementById('tab-'+id).classList.add('active');
  if (id === 'config') renderConfigTarjetas();
  if (id === 'custom') buildColorGrid();
  if (id === 'stats') { setTimeout(renderStats, 50); } // timeout para que el canvas tenga ancho
}
function activarPresentacion() { document.body.classList.add('modo-presentacion'); showTab('pantalla'); }
function salirPresentacion() { document.body.classList.remove('modo-presentacion'); }
function updateMsg() {
  const val = document.getElementById('msg-input').value.trim();
  document.getElementById('premio-box').textContent = val;
  fetch('/api/state',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({premio:val})});
}
function updatePremioSize() {
  const size = document.getElementById('premio-size').value;
  document.getElementById('premio-box').style.fontSize = size;
}

function updateHoraFin() {
  horaFin = document.getElementById('hora-fin-input').value || '05:30';
  ganadorMostrado = false;
  const el = document.getElementById('clock-fin');
  if (el) el.textContent = horaFin;
  fetch('/api/state',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({hora_fin:horaFin})});
}function fmt(n) { return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0}); }
function fmtLabel(n){
  if(n>=1000000){const m=n/1000000;return(m%1===0?m:m.toFixed(1))+(m<2?' MILLÓN':' MILLONES');}
  if(n>=1000){const k=n/1000;return(k%1===0?k:k.toFixed(1))+(k<2?' MIL':' MILES');}
  return '';
}
function fmtDisplay(n){
  const base='$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0});
  const lbl=fmtLabel(n);
  return lbl?base+'<span class="miles-lbl">'+lbl+'</span>':base;
}
function showToast(msg, error=false) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className='toast show'+(error?' error':'');
  setTimeout(()=>t.className='toast', 2800);
}

async function cerrarNoche() {
  if (!confirm('¿Cerrar la noche y guardar en el historial?\nLos datos quedan registrados para siempre.')) return;
  try {
    const r = await fetch('/api/cerrar_noche', {method:'POST'});
    const d = await r.json();
    if (d.ok) {
      showToast('✓ Noche guardada en el historial');
      if (confirm('¿Exportar la noche a Excel ahora?')) window.open('/api/export/excel','_blank');
    } else { showToast(d.error||'Error al cerrar la noche', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

async function resetNoche() {
  if (!confirm('Resetear todos los consumos? Los saldos de tarjetas vuelven al valor inicial.')) return;
  await fetch('/api/reset',{method:'POST'});
  ganadorMostrado = false;
  knownNames.clear();
  for(let c=1;c<=3;c++){window['cajaState'+c]=null;renderCajaInner(c);}
  await loadData();
  showToast('Noche reseteada');
}

function renderCajaInner(caja) {
  const container = document.getElementById('caja-inner-'+caja);
  const ta = window['cajaState'+caja];
  let tarjetaHTML = '';
  if (ta) {
    const pct = ta.saldo_inicial>0 ? Math.max(0,Math.round((ta.saldo_actual/ta.saldo_inicial)*100)) : 0;
    const sinSaldo = ta.saldo_actual <= 0;
    const pctSaldo = ta.saldo_inicial > 0 ? ta.saldo_actual / ta.saldo_inicial : 1;
    const warnSaldo = !sinSaldo && pctSaldo <= 0.2;
    tarjetaHTML = `<div class="tarjeta-card visible ${sinSaldo?'sin-saldo':''}">
      <div class="tarjeta-top">
        <div><div class="tarjeta-mesa-label">Mesa</div><div class="tarjeta-mesa-num">${ta.mesa}</div></div>
        <div class="tarjeta-saldo-wrap"><div class="tarjeta-saldo-label">Saldo disponible</div><div class="tarjeta-saldo">${fmt(ta.saldo_actual)}</div></div>
      </div>
      <div class="tarjeta-bar-wrap"><div class="tarjeta-bar" style="width:${pct}%"></div></div>
      ${ta.nombre?`<div class="tarjeta-nombre">Cliente: <span>${ta.nombre}</span></div>`:''}
      ${warnSaldo?`<div class="saldo-bajo-warn">⚠ Saldo bajo — queda ${Math.round(pctSaldo*100)}%</div>`:''}
      ${sinSaldo?`<div class="saldo-bajo-warn" style="border-color:#a83030;color:#ff4444;animation:none;">✕ Sin saldo disponible</div>`:''}
    </div>`;
  }

  container.innerHTML = `
    <div class="caja-header-row">
      <span class="caja-badge">CAJA ${caja}</span>
      <span class="caja-title">Registrar consumo</span>
      <button onclick="abrirCartelModal()" style="margin-left:auto;background:linear-gradient(135deg,#c9a227,#e8c84a);color:#000;border:none;border-radius:6px;padding:7px 16px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;letter-spacing:2px;cursor:pointer;transition:all .15s;white-space:nowrap;" onmouseover="this.style.opacity='0.85'" onmouseout="this.style.opacity='1'">📣 CARTEL</button>
    </div>
    <div style="margin-bottom:12px;">
      <button onclick="this.nextElementSibling.style.display=this.nextElementSibling.style.display==='none'?'block':'none';this.textContent=this.nextElementSibling.style.display==='none'?'▸ Instrucciones rápidas':'▾ Instrucciones rápidas';" style="background:none;border:1px solid #2a2a2a;border-radius:6px;color:#888;font-size:11px;letter-spacing:1px;padding:4px 12px;cursor:pointer;">▸ Instrucciones rápidas</button>
      <div style="display:none;margin-top:8px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:8px;padding:12px 16px;font-size:12px;line-height:1.8;color:#aaa;">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;">
          <div>
            <div style="color:#e8c84a;font-weight:700;letter-spacing:1px;margin-bottom:4px;">▤ COBRAR CON TARJETA</div>
            <div>1. Pasá la tarjeta por el lector</div>
            <div>2. Verificá el saldo disponible</div>
            <div>3. Ingresá el nombre del cliente</div>
            <div>4. Ingresá el monto a cobrar</div>
            <div>5. Presioná <strong style="color:#fff;">+ Agregar</strong></div>
            <div style="color:#888;margin-top:4px;font-size:11px;">El sistema descuenta el monto del saldo de la tarjeta automáticamente.</div>
          </div>
          <div>
            <div style="color:#e8c84a;font-weight:700;letter-spacing:1px;margin-bottom:4px;">✎ COBRAR SIN TARJETA</div>
            <div>1. Seleccioná la pestaña <strong style="color:#fff;">Sin tarjeta</strong></div>
            <div>2. Ingresá el nombre del cliente</div>
            <div>3. Ingresá el número de mesa</div>
            <div>4. Ingresá el monto consumido</div>
            <div>5. Presioná <strong style="color:#fff;">+ Agregar</strong></div>
            <div style="color:#888;margin-top:4px;font-size:11px;">Usá este modo si el cliente no tiene tarjeta VIP o el lector no funciona.</div>
          </div>
          <div>
            <div style="color:#3a9a5a;font-weight:700;letter-spacing:1px;margin-bottom:4px;">⊕ CARGAR SALDO</div>
            <div>1. Seleccioná la pestaña <strong style="color:#fff;">Recargar</strong></div>
            <div>2. Pasá la tarjeta por el lector</div>
            <div>3. Verificá que sea la tarjeta correcta</div>
            <div>4. Ingresá el monto a agregar</div>
            <div>5. Presioná <strong style="color:#fff;">⊕ Recargar saldo</strong></div>
            <div style="color:#888;margin-top:4px;font-size:11px;">El saldo se suma al disponible. Guardá la config. de tarjetas luego.</div>
          </div>
        </div>
      </div>
    </div>
    <div class="modo-tabs">
      <div class="modo-tab ${!ta?'active':''}" id="modo-tab-tarjeta-${caja}" onclick="setModo(${caja},'tarjeta')">▤ Con tarjeta</div>
      <div class="modo-tab" id="modo-tab-manual-${caja}" onclick="setModo(${caja},'manual')">✎ Sin tarjeta</div>
      <div class="modo-tab" id="modo-tab-recargar-${caja}" onclick="setModo(${caja},'recargar')" style="color:#3a9a5a;">⊕ Recargar</div>
    </div>
    <div class="modo-content active" id="modo-content-tarjeta-${caja}">
      <div class="scan-hint ${!ta?'esperando':''}">
        <span class="scan-icon">▤</span>
        ${ta ? 'Tarjeta activa — pasa otra para cambiar' : 'Pasa la tarjeta por el lector para continuar'}
      </div>
      ${tarjetaHTML}
      <div class="form-card" ${!ta?'style="opacity:.35;pointer-events:none"':''}>
        <label class="field-label">Nombre del cliente</label>
        <input class="field-input" id="name${caja}" type="text" placeholder="Nombre..." autocomplete="off" list="nl${caja}" value="${ta&&ta.nombre?ta.nombre:''}" />
        <datalist id="nl${caja}"></datalist>
        <label class="field-label">Monto ($)</label>
        <input class="field-input amount-input" id="amount${caja}" type="number" min="0" step="100" placeholder="0" />
        <div class="hint-miles" id="hint${caja}">Ingresá el monto completo</div>
        <div class="btn-row" style="margin-top:4px">
          <button class="btn-add" id="btnadd${caja}" onclick="addTx(${caja},true)" ${!ta?'disabled':''}>+ Agregar</button>
        </div>
      </div>
    </div>
    <div class="modo-content" id="modo-content-manual-${caja}">
      <div class="form-card">
        <label class="field-label">Nombre del cliente</label>
        <input class="field-input" id="mname${caja}" type="text" placeholder="Nombre..." autocomplete="off" list="mnl${caja}" />
        <datalist id="mnl${caja}"></datalist>
        <label class="field-label">Mesa (opcional)</label>
        <input class="field-input" id="mmesa${caja}" type="text" placeholder="Ej: 5" />
        <label class="field-label">Monto ($)</label>
        <input class="field-input amount-input" id="mamount${caja}" type="number" min="0" step="100" placeholder="0" />
        <div class="hint-miles" id="mhint${caja}">Ingresá el monto completo</div>
        <div class="btn-row" style="margin-top:4px">
          <button class="btn-add" onclick="addTx(${caja},false)">+ Agregar</button>
        </div>
      </div>
    </div>
    <div class="modo-content" id="modo-content-recargar-${caja}">
      <div class="scan-hint ${!ta?'esperando':''}">
        <span class="scan-icon">▤</span>
        ${ta ? 'Tarjeta activa — pasa otra para cambiar' : 'Pasa la tarjeta para recargar'}
      </div>
      ${tarjetaHTML}
      <div class="form-card" ${!ta?'style="opacity:.35;pointer-events:none"':''}>
        <label class="field-label" style="color:#3a9a5a;">Monto a recargar ($)</label>
        <input class="field-input amount-input" id="recarga${caja}" type="number" min="0" step="100" placeholder="0" style="color:#3a9a5a;" />
        <div class="hint-miles" id="rhint${caja}">Ingresá el monto a agregar</div>
        <div class="btn-row" style="margin-top:4px">
          <button class="btn-add" style="background:#2a6a3a;border-color:#3a9a5a;" onclick="recargarTarjeta(${caja})" ${!ta?'disabled':''}>⊕ Recargar saldo</button>
        </div>
      </div>
    </div>
    <div class="section-label" style="margin-top:16px">Ultimas operaciones</div>
    <div id="txlist${caja}" class="tx-list"></div>
    <div class="caja-total-bar">
      <span class="caja-total-label">Total Caja ${caja}</span>
      <span class="caja-total-val" id="ctotal${caja}">$0</span>
    </div>`;

  const amEl = document.getElementById('amount'+caja);
  const hintEl = document.getElementById('hint'+caja);
  if (amEl && hintEl) {
    amEl.addEventListener('input', () => {
      const v = parseFloat(amEl.value);
      if (v>0){
        const label = v>=1000000 ? ' · '+( v/1000000).toFixed(v%1000000===0?0:1)+' millón' :
                      v>=1000 ? ' · '+(v/1000).toFixed(v%1000===0?0:1)+' mil' :
                      v>=100 ? ' · '+Math.round(v/100)+' cientos' : '';
        hintEl.textContent='= '+fmt(v)+label;
        hintEl.className='hint-miles ok';
      } else {
        hintEl.textContent='Ingresá el monto completo';
        hintEl.className='hint-miles';
      }
    });
    // Enter en monto con tarjeta → agregar
    amEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); addTx(caja, true); }
    });
  }
  // Enter en nombre con tarjeta → pasar al foco al monto
  const nameEl = document.getElementById('name'+caja);
  if (nameEl) {
    nameEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); if (amEl) amEl.focus(); }
    });
  }

  const mamEl = document.getElementById('mamount'+caja);
  const mhintEl = document.getElementById('mhint'+caja);
  if (mamEl && mhintEl) {
    mamEl.addEventListener('input', () => {
      const v = parseFloat(mamEl.value);
      if (v>0){
        const label = v>=1000000 ? ' · '+(v/1000000).toFixed(v%1000000===0?0:1)+' millón' :
                      v>=1000 ? ' · '+(v/1000).toFixed(v%1000===0?0:1)+' mil' :
                      v>=100 ? ' · '+Math.round(v/100)+' cientos' : '';
        mhintEl.textContent='= '+fmt(v)+label;
        mhintEl.className='hint-miles ok';
      } else {
        mhintEl.textContent='Ingresá el monto completo';
        mhintEl.className='hint-miles';
      }
    });
    // Enter en monto manual → agregar
    mamEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); addTx(caja, false); }
    });
  }
  // Enter en nombre manual → foco a mesa; Enter en mesa → foco a monto
  const mnameEl = document.getElementById('mname'+caja);
  const mmesaEl = document.getElementById('mmesa'+caja);
  if (mnameEl) {
    mnameEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); if (mmesaEl) mmesaEl.focus(); }
    });
  }
  if (mmesaEl) {
    mmesaEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); if (mamEl) mamEl.focus(); }
    });
  }
  // Hint monto recarga
  const reEl = document.getElementById('recarga'+caja);
  const rhintEl = document.getElementById('rhint'+caja);
  if (reEl && rhintEl) {
    reEl.addEventListener('input', () => {
      const v = parseFloat(reEl.value);
      if (v>0){ const lbl=fmtLabel(v); rhintEl.textContent='= '+fmt(v)+(lbl?' · '+lbl:''); rhintEl.className='hint-miles ok'; }
      else { rhintEl.textContent='Ingresá el monto a agregar'; rhintEl.className='hint-miles'; }
    });
    reEl.addEventListener('keydown', e=>{ if(e.key==='Enter'){e.preventDefault();recargarTarjeta(caja);} });
  }
  updateNamelists();
  renderCajaList(caja);
}

async function recargarTarjeta(caja) {
  const ta = window['cajaState'+caja];
  if (!ta) { showToast('Pasa una tarjeta primero', true); return; }
  const monto = parseFloat(document.getElementById('recarga'+caja).value);
  if (!monto || monto <= 0) { showToast('Ingresá un monto válido', true); return; }
  if (!confirm(`¿Agregar ${fmt(monto)} al saldo de Mesa ${ta.mesa}?`)) return;
  try {
    const r = await fetch('/api/tarjetas/recargar', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({codigo:ta.codigo, monto})});
    const d = await r.json();
    if (d.ok) {
      // Actualizar estado local ANTES de cualquier render
      ta.saldo_actual = d.nuevo_saldo;
      window['cajaState'+caja] = ta;
      // Sincronizar tarjetasData localmente para que el próximo render sea correcto
      if (!tarjetasData[ta.codigo]) tarjetasData[ta.codigo] = {};
      tarjetasData[ta.codigo].saldo_actual = d.nuevo_saldo;
      document.getElementById('recarga'+caja).value = '';
      showToast(`✓ Recargado ${fmt(monto)} — Nuevo saldo: ${fmt(d.nuevo_saldo)}`);
      await loadData();
      // Forzar el saldo correcto del server (autoritativo) post-loadData
      if (window['cajaState'+caja]) window['cajaState'+caja].saldo_actual = d.nuevo_saldo;
      if (!tarjetasData[ta.codigo]) tarjetasData[ta.codigo] = {};
      tarjetasData[ta.codigo].saldo_actual = d.nuevo_saldo;
      renderCajaInner(caja);
      setModo(caja, 'recargar');
    } else { showToast(d.error||'Error', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

async function addTx(caja, conTarjeta) {
  let name, amount, mesa, tarjeta_codigo;
  if (conTarjeta) {
    const ta = window['cajaState'+caja];
    if (!ta) { showToast('Pasa una tarjeta primero', true); return; }
    name = document.getElementById('name'+caja).value.trim();
    amount = parseFloat(document.getElementById('amount'+caja).value);
    mesa = ta.mesa;
    tarjeta_codigo = ta.codigo;
    if (!name) { document.getElementById('name'+caja).focus(); showToast('Falta el nombre', true); return; }
    if (!amount || amount<=0) { document.getElementById('amount'+caja).focus(); showToast('Falta el monto', true); return; }
    if (amount > ta.saldo_actual) { showToast('Saldo insuficiente! Disponible: '+fmt(ta.saldo_actual), true); return; }
    try {
      const res = await fetch('/api/tx',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,amount,caja,mesa,tarjeta_codigo,client_time:new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',hour12:false})})});
      if (res.ok) {
        ta.saldo_actual = ta.saldo_actual - amount;
        ta.nombre = name;
        window['cajaState'+caja] = ta;
        document.getElementById('amount'+caja).value='';
        showToast(fmt(amount)+' descontado de Mesa '+ta.mesa+' — Saldo restante: '+fmt(ta.saldo_actual));
        await loadData();
        if (window['cajaState'+caja]) window['cajaState'+caja].saldo_actual = ta.saldo_actual;
        renderCajaInner(caja);
        setModo(caja,'tarjeta');
      } else {
        const err = await res.json().catch(()=>({error:'Error del servidor'}));
        showToast(err.error||'Error del servidor', true);
      }
    } catch(e){ showToast('Error de conexion',true); }
  } else {
    name = document.getElementById('mname'+caja).value.trim();
    amount = parseFloat(document.getElementById('mamount'+caja).value);
    mesa = document.getElementById('mmesa'+caja).value.trim();
    tarjeta_codigo = '';
    if (!name) { document.getElementById('mname'+caja).focus(); showToast('Falta el nombre',true); return; }
    if (!amount||amount<=0) { document.getElementById('mamount'+caja).focus(); showToast('Falta el monto',true); return; }
    try {
      const res = await fetch('/api/tx',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,amount,caja,mesa,tarjeta_codigo,client_time:new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',hour12:false})})});
      if (res.ok) {
        document.getElementById('mamount'+caja).value='';
        showToast('Registrado — '+fmt(amount)+' para '+name);
        await loadData(); renderCajaInner(caja);
        setModo(caja,'manual');
      }
    } catch(e){ showToast('Error de conexion',true); }
  }
}

async function deleteTx(id) {
  await fetch('/api/tx/'+id,{method:'DELETE'});
  await loadData();
}
function editTx(id, amount, name) {
  document.getElementById('edit-tx-id').value = id;
  document.getElementById('edit-tx-amount').value = amount;
  document.getElementById('edit-tx-name').value = name;
  document.getElementById('edit-tx-modal').style.display = 'flex';
  setTimeout(()=>document.getElementById('edit-tx-amount').select(), 50);
}
async function confirmarEditTx() {
  const id     = parseInt(document.getElementById('edit-tx-id').value);
  const amount = parseFloat(document.getElementById('edit-tx-amount').value);
  const name   = document.getElementById('edit-tx-name').value.trim();
  if (!amount || amount <= 0) { showToast('Monto inválido', true); return; }
  try {
    const r = await fetch('/api/tx/'+id, {method:'PUT', headers:{'Content-Type':'application/json'}, body:JSON.stringify({amount, name})});
    const d = await r.json();
    if (d.ok) {
      document.getElementById('edit-tx-modal').style.display = 'none';
      await loadData();
      showToast('Operación actualizada');
    } else { showToast(d.error||'Error', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

let lastWinnerTs = 0;
let lastCartelTs = 0;

async function loadData() {
  try {
    const [r1,r2,r3] = await Promise.all([fetch('/api/tx'),fetch('/api/tarjetas'),fetch('/api/state')]);
    txData = await r1.json(); tarjetasData = await r2.json();
    const st = await r3.json();
    setStatus(true); render();
    // Sincronizar estado compartido en todas las pantallas
    if (st.hora_fin && st.hora_fin !== horaFin) {
      horaFin = st.hora_fin;
      const hfi = document.getElementById('hora-fin-input');
      if (hfi) hfi.value = horaFin;
      const hfe = document.getElementById('clock-fin');
      if (hfe) hfe.textContent = horaFin;
    }
    if (st.premio !== undefined) {
      const msgEl = document.getElementById('msg-input');
      if (msgEl && msgEl.value !== st.premio) msgEl.value = st.premio;
      const pb = document.getElementById('premio-box');
      if (pb) pb.textContent = st.premio;
    }
    // Mostrar ganador en todas las pantallas si el servidor lo indica
    if (st.winner_show && st.winner_ts && st.winner_ts !== lastWinnerTs) {
      lastWinnerTs = st.winner_ts;
      ganadorMostrado = true;
      mostrarGanador();
    } else if (!st.winner_show && document.getElementById('winner-overlay').classList.contains('show')) {
      document.getElementById('winner-overlay').classList.remove('show');
    }
    // Sincronizar cartel en todas las pantallas
    if (st.cartel_show && st.cartel_ts && st.cartel_ts !== lastCartelTs) {
      lastCartelTs = st.cartel_ts;
      const cd = st.cartel_data || {};
      // Mostrar cartel directamente sin abrir el modal
      const nd = document.getElementById('cartel-nombre-display');
      const md = document.getElementById('cartel-mesa-display');
      const fd = document.getElementById('cartel-frase-display');
      const ed = document.getElementById('cartel-emoji-big');
      if (nd) { nd.textContent = cd.nombre||''; nd.style.display = cd.nombre ? 'block' : 'none'; }
      if (md) md.textContent = cd.mesa ? 'MESA '+cd.mesa : '';
      if (fd) fd.textContent = cd.frase||'';
      if (ed) ed.textContent = cd.emoji||'🍾';
      // Regenerar rayos en todas las pantallas
      const rays = document.getElementById('cartel-rays');
      if (rays) {
        rays.innerHTML = '';
        const rayColor = temaActual === 'jagger12boxeo' ? 'rgba(255,34,34,0.12)' : temaActual === 'touchofpink' ? 'rgba(244,114,182,0.12)' : 'rgba(201,162,39,0.12)';
        for (let i = 0; i < 12; i++) {
          const angle = i * 30;
          const r = document.createElement('div');
          r.style.cssText = `position:absolute;left:50%;top:50%;width:1px;height:55vh;background:linear-gradient(to bottom,${rayColor},transparent);transform-origin:0% 0%;transform:rotate(${angle}deg);opacity:0.5;animation:rayPulse2 ${2+i*0.15}s ease-in-out ${i*0.1}s infinite alternate;`;
          r.style.setProperty('--r', (i*30)+'deg');
          rays.appendChild(r);
        }
        if (!document.getElementById('kf-ray2')) {
          const s = document.createElement('style'); s.id = 'kf-ray2';
          s.textContent = '@keyframes rayPulse2{0%{opacity:0.2;transform:rotate(var(--r,0deg)) scaleY(0.5)}100%{opacity:0.7;transform:rotate(var(--r,0deg)) scaleY(1)}}';
          document.head.appendChild(s);
        }
      }
      // Regenerar fondo del cartel
      const bg = document.getElementById('cartel-tema-bg');
      if (bg) {
        bg.innerHTML = temaActual === 'jagger12boxeo' ?
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#1a0000 0%,#050000 100%);"></div>
           <div style="position:absolute;bottom:0;left:0;font-size:160px;opacity:0.06;transform:rotate(-15deg);">🥊</div>
           <div style="position:absolute;bottom:0;right:0;font-size:160px;opacity:0.06;transform:rotate(15deg) scaleX(-1);">🥊</div>` :
          temaActual === 'jagger12' ?
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#111 0%,#000 100%);"></div>
           <div style="position:absolute;inset:0;background:radial-gradient(ellipse at 50% 50%,rgba(201,162,39,0.05) 0%,transparent 70%);"></div>` :
          temaActual === 'touchofpink' ?
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#1e0018 0%,#080005 100%);"></div>
           <div style="position:absolute;inset:0;background:radial-gradient(ellipse at 50% 50%,rgba(244,114,182,0.07) 0%,transparent 70%);"></div>` :
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#0a0a0a 0%,#000 100%);"></div>`;
      }
      const overlay = document.getElementById('cartel-overlay');
      if (overlay) {
        overlay.style.display = 'flex';
        const c = document.getElementById('cartel-content');
        if (c) {
          c.style.animation = 'none';
          c.style.opacity = '0';
          requestAnimationFrame(() => {
            requestAnimationFrame(() => {
              c.style.animation = 'winnerEntrada 0.8s cubic-bezier(.22,1,.36,1) forwards';
              c.style.opacity = '';
            });
          });
        }
      }
    } else if (!st.cartel_show) {
      const overlay = document.getElementById('cartel-overlay');
      if (overlay && overlay.style.display === 'flex') overlay.style.display = 'none';
    }
    for (let c = 1; c <= 3; c++) {
      const ta = window['cajaState'+c];
      if (ta && ta.codigo && tarjetasData[ta.codigo] !== undefined) {
        ta.saldo_actual = tarjetasData[ta.codigo].saldo_actual;
        window['cajaState'+c] = ta;
      }
    }
  } catch(e){ setStatus(false); }
}

async function sincronizarConfTarjetas() {
  const tabAbierta = document.getElementById('tab-config') && document.getElementById('tab-config').classList.contains('active');
  if (tabAbierta) return;
  try {
    const res = await fetch('/api/tarjetas/config');
    if (res.ok) { const d = await res.json(); if (d && d.length) confTarjetas = d; }
  } catch(e){}
}

function setStatus(ok) {
  const el = document.getElementById('status-txt');
  if (ok){el.textContent='Conectado';el.className='status-ok';document.getElementById('last-update').textContent=new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',second:'2-digit',hour12:false});}
  else{el.textContent='Sin conexion...';el.className='status-err';}
}

function render() {
  renderPantalla();
  for(let c=1;c<=3;c++) renderCajaList(c);
  updateNamelists();
  renderStats();
}

let knownNames = new Set();
let prevRankOrder = [];
let koAnimEnCurso = false;

function renderPantalla() {
  if (koAnimEnCurso) return;

  const header = document.getElementById('rank-header');
  const rows   = document.getElementById('rank-rows');
  const empty  = document.getElementById('empty-msg');

  const totals = {}, mesas = {};
  txData.forEach(t => {
    totals[t.name] = (totals[t.name] || 0) + t.amount;
    if (t.mesa && !mesas[t.name]) mesas[t.name] = t.mesa;
  });
  const names = Object.keys(totals).sort((a,b) => totals[b] - totals[a]).slice(0, 5);

  if (!names.length) {
    header.style.display = 'none';
    rows.innerHTML = '';
    empty.style.display = 'block';
    knownNames.clear();
    prevRankOrder = [];
    return;
  }
  header.style.display = 'grid';
  empty.style.display  = 'none';

  const esc = s => String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');

  // ── Detectar KO animación (solo tema boxeo) ──
  if (temaActual === 'jagger12boxeo' && koAnimActiva && prevRankOrder.length > 0) {
    const suben = names.filter((n,i) => { const p = prevRankOrder.indexOf(n); return p !== -1 && p > i; });
    const bajan = names.filter((n,i) => { const p = prevRankOrder.indexOf(n); return p !== -1 && p < i; });
    if (suben.length >= 1 && bajan.length >= 1) {
      const nameUp = suben[0], nameKO = bajan[0];
      const cardUp = document.querySelector(`.rank-row[data-name="${CSS.escape(nameUp)}"]`);
      const cardKO = document.querySelector(`.rank-row[data-name="${CSS.escape(nameKO)}"]`);
      if (cardUp && cardKO) {
        koAnimEnCurso = true;
        const onComplete = () => {
          rows.innerHTML = names.map((name,i) => {
            const pos = i+1, rc = pos<=3?'rank-'+pos:'';
            return `<div class="rank-row ${rc}" data-name="${esc(name)}" data-pos="${pos}"><div class="col-puesto">#${pos}</div><div class="col-nombre">${esc(name)}</div><div class="col-mesa">${esc(mesas[name]||'—')}</div><div class="col-total">${fmt(totals[name])}</div></div>`;
          }).join('');
          prevRankOrder = [...names];
          knownNames    = new Set(names);
          setTimeout(() => { koAnimEnCurso = false; }, 700);
        };
        animateRankSwap(cardUp, cardKO, onComplete);
        return;
      }
    }
  }

  // ══════════════════════════════════════════
  //  FLIP: animación orgánica para todos los modos
  // ══════════════════════════════════════════

  // 1. FIRST — capturar posiciones actuales de cada tarjeta existente
  const firstRects = {};
  rows.querySelectorAll('.rank-row[data-name]').forEach(el => {
    firstRects[el.dataset.name] = el.getBoundingClientRect();
  });

  const prevNames = new Set(prevRankOrder);
  const hasMoves  = names.some((n,i) => prevRankOrder[i] !== n && prevNames.has(n));

  // 2. LAST — actualizar el DOM al nuevo orden (números ya correctos)
  rows.innerHTML = names.map((name, i) => {
    const pos    = i + 1;
    const rc     = pos <= 3 ? 'rank-' + pos : '';
    const esNuevo = !knownNames.has(name);
    return `<div class="rank-row ${rc}${esNuevo?' nueva':''}" data-name="${esc(name)}" data-pos="${pos}" style="will-change:transform,opacity;">
      <div class="col-puesto">#${pos}</div>
      <div class="col-nombre">${esc(name)}</div>
      <div class="col-mesa">${esc(mesas[name]||'—')}</div>
      <div class="col-total">${fmt(totals[name])}</div>
    </div>`;
  }).join('');

  prevRankOrder = [...names];
  knownNames    = new Set(names);

  // 3. INVERT + PLAY — animar solo si hay movimientos reales
  if (hasMoves) {
    // Leer posiciones AFTER el re-render para calcular deltas reales
    rows.querySelectorAll('.rank-row[data-name]').forEach(el => {
      const name  = el.dataset.name;
      const first = firstRects[name];
      if (!first || el.classList.contains('nueva')) return;

      const last = el.getBoundingClientRect();
      const dy   = first.top - last.top;

      if (Math.abs(dy) < 2) return;

      const subeAlPrimero = el.classList.contains('rank-1') && dy > 0;

      // Invert: colocar la tarjeta en su posición VIEJA visualmente
      // El número ya muestra el valor nuevo (correcto) — solo el bloque se mueve
      el.style.transition = 'none';
      el.style.transform  = `translateY(${dy}px)`;
      // Opacidad inicial: la que sube arranca un poco más tenue
      el.style.opacity    = dy > 0 ? '0.6' : '0.9';

      void el.offsetHeight; // force reflow

      // Play: movimiento lento y orgánico hacia la posición final
      // Más lento cuanto mayor es el desplazamiento
      const dur   = Math.min(1.1, 0.65 + Math.abs(dy) / 900);
      const delay = dy > 0 ? 0 : 0.06; // las que bajan con mínimo delay

      el.style.transition = `transform ${dur}s cubic-bezier(.25,.46,.45,.94) ${delay}s,
                              opacity   ${dur * 0.6}s ease ${delay}s`;
      el.style.transform  = 'translateY(0)';
      el.style.opacity    = '1';

      // Glow dorado suave al llegar al #1
      if (subeAlPrimero) {
        setTimeout(() => {
          el.classList.add('ascendio');
          setTimeout(() => el.classList.remove('ascendio'), 950);
        }, (dur + delay) * 1000 - 80);
      }

      // Cleanup estilos inline
      setTimeout(() => {
        el.style.transition = '';
        el.style.transform  = '';
        el.style.opacity    = '';
      }, (dur + delay) * 1000 + 80);
    });
  }

  // Limpiar clase nueva después de la animación
  rows.querySelectorAll('.rank-row.nueva').forEach(el => {
    el.addEventListener('animationend', () => el.classList.remove('nueva'), { once: true });
  });
}

// ════════════════════════════════════════════
//  animateRankSwap — ANIMACIÓN KO COMPLETA
// ════════════════════════════════════════════
// Glove element (shared, created once)
(function() {
  if (!document.getElementById('ko-glove')) {
    const g = document.createElement('div');
    g.id = 'ko-glove';
    g.textContent = '🥊';
    g.style.cssText = `
      position:fixed;font-size:60px;z-index:9999;pointer-events:none;
      display:none;transform-origin:center center;transform:rotate(90deg);
      filter:drop-shadow(0 0 18px rgba(255,50,50,0.9));
      will-change:transform,opacity;
    `;
    document.body.appendChild(g);
  }
  // Keyframes for glove jab (in rotated space: translateY = horizontal movement)
  if (!document.getElementById('kf-ko-system')) {
    const s = document.createElement('style'); s.id = 'kf-ko-system';
    s.textContent = `
      @keyframes jabVibrate {
        0%   { transform: translateX(0); }
        20%  { transform: translateX(-9px) rotate(-1deg); }
        40%  { transform: translateX(7px) rotate(0.5deg); }
        60%  { transform: translateX(-5px); }
        80%  { transform: translateX(4px); }
        100% { transform: translateX(0); }
      }
      @keyframes cardFlyRight {
        0%   { transform: translateX(0); opacity:1; }
        100% { transform: translateX(115vw); opacity:0; }
      }
      @keyframes flashImpact {
        0%   { background:#3a0000; border-color:#ff2222; filter:brightness(2); }
        100% { background:inherit; border-color:inherit; filter:none; }
      }
    `;
    document.head.appendChild(s);
  }
})();

/**
 * animateRankSwap(cardOvertaking, cardBeingOvertaken, onComplete)
 * - cardOvertaking     : el elemento DOM de la tarjeta que SUBE
 * - cardBeingOvertaken : el elemento DOM de la tarjeta que BAJA (recibe el KO)
 * - onComplete         : callback que hace el swap de datos + re-render
 *
 * Coreografía:
 *  1. Guante → aparece horizontal a la izquierda de la víctima
 *  2. Jab 1 + Jab 2 (vibración)
 *  3. Wind-up (retroceso)
 *  4. Power Punch → tarjeta vuela por la derecha (position:fixed para libertad total)
 *  5. onComplete() → swap datos + renderRanking() → DOM actualizado
 *  6. Tarjeta que bajó: entra desde la izquierda a su NUEVO slot correcto
 *  7. Tarjeta que subió: se desliza verticalmente desde donde estaba
 */
function animateRankSwap(cardOvertaking, cardBeingOvertaken, onComplete) {
  const glove = document.getElementById('ko-glove');

  // ── Capturar posiciones ANTES de cualquier cambio ──
  const rectKO = cardBeingOvertaken.getBoundingClientRect();
  const rectUp = cardOvertaking.getBoundingClientRect();

  // Guardar data-names antes de que el DOM cambie
  const nameKO = cardBeingOvertaken.dataset.name;
  const nameUp = cardOvertaking.dataset.name;

  // Posicionar guante horizontal (→) a la izquierda de la víctima
  const gloveY = rectKO.top + rectKO.height / 2 - 34;
  const gloveX = rectKO.left - 90;

  glove.style.cssText = `
    position:fixed; display:block;
    top:${gloveY}px; left:${gloveX}px;
    font-size:60px; z-index:9999; pointer-events:none;
    transform:rotate(90deg) translateY(0px) scaleX(1);
    opacity:1; transition:none;
    filter:drop-shadow(0 0 18px rgba(255,50,50,0.9));
    transform-origin:center center;
    will-change:transform,opacity;
  `;

  // ── Timings ──
  const T_JAB1      = 80;
  const T_JAB1_RET  = 220;
  const T_JAB2      = 390;
  const T_JAB2_RET  = 530;
  const T_WINDUP    = 650;
  const T_PUNCH     = 840;
  const T_GLOVE_OUT = 1020;
  const T_RERENDER  = 1080;

  function jabCard() {
    cardBeingOvertaken.style.animation = 'none';
    void cardBeingOvertaken.offsetWidth;
    cardBeingOvertaken.style.animation = 'jabVibrate 0.16s ease-in-out';
  }

  // JAB 1
  setTimeout(() => {
    glove.style.transition = 'transform 0.10s ease-out';
    glove.style.transform  = 'rotate(90deg) translateY(-58px) scaleX(1.15)';
    jabCard();
  }, T_JAB1);
  setTimeout(() => {
    glove.style.transition = 'transform 0.09s ease-in';
    glove.style.transform  = 'rotate(90deg) translateY(-6px) scaleX(0.92)';
  }, T_JAB1_RET);

  // JAB 2
  setTimeout(() => {
    glove.style.transition = 'transform 0.10s ease-out';
    glove.style.transform  = 'rotate(90deg) translateY(-58px) scaleX(1.15)';
    jabCard();
  }, T_JAB2);
  setTimeout(() => {
    glove.style.transition = 'transform 0.09s ease-in';
    glove.style.transform  = 'rotate(90deg) translateY(-6px) scaleX(0.92)';
  }, T_JAB2_RET);

  // WIND-UP
  setTimeout(() => {
    glove.style.transition = 'transform 0.17s cubic-bezier(.4,0,.2,1)';
    glove.style.transform  = 'rotate(90deg) translateY(55px) scaleX(0.72)';
  }, T_WINDUP);

  // POWER PUNCH → tarjeta sale volando por la derecha
  setTimeout(() => {
    glove.style.transition = 'transform 0.14s cubic-bezier(.1,0,.5,1)';
    glove.style.transform  = 'rotate(90deg) translateY(-110px) scaleX(1.35)';

    // Flash impacto
    cardBeingOvertaken.style.animation = 'flashImpact 0.25s ease-out';

    // Sacar la tarjeta del flujo normal → position:fixed en su lugar exacto
    cardBeingOvertaken.style.position = 'fixed';
    cardBeingOvertaken.style.top      = rectKO.top + 'px';
    cardBeingOvertaken.style.left     = rectKO.left + 'px';
    cardBeingOvertaken.style.width    = rectKO.width + 'px';
    cardBeingOvertaken.style.zIndex   = '200';
    cardBeingOvertaken.style.margin   = '0';
    void cardBeingOvertaken.offsetWidth;

    cardBeingOvertaken.style.transition = 'transform 0.28s cubic-bezier(.4,0,1,1), opacity 0.2s 0.08s ease';
    cardBeingOvertaken.style.transform  = 'translateX(115vw)';
    cardBeingOvertaken.style.opacity    = '0';
  }, T_PUNCH);

  // Guante fade out
  setTimeout(() => {
    glove.style.transition = 'opacity 0.2s ease';
    glove.style.opacity    = '0';
  }, T_GLOVE_OUT);

  // ── RE-RENDER + animaciones de entrada ──
  setTimeout(() => {
    // 1. Ejecutar onComplete: actualiza datos y re-renderiza el DOM
    onComplete();

    // 2. Buscar las nuevas tarjetas en el DOM ya re-renderizado
    const newCardKO = document.querySelector(`.rank-row[data-name="${CSS.escape(nameKO)}"]`);
    const newCardUp = document.querySelector(`.rank-row[data-name="${CSS.escape(nameUp)}"]`);

    // 3. Preparar estado inicial ANTES de que el browser pinte
    //    — Tarjeta que bajó: invisible, fuera de pantalla a la izquierda
    //    — Tarjeta que subió: desplazada hacia abajo (donde estaba antes)
    if (newCardUp) {
      const destUp = newCardUp.getBoundingClientRect();
      const deltaY = rectUp.top - destUp.top;
      newCardUp.style.transition = 'none';
      newCardUp.style.transform  = `translateY(${deltaY}px)`;
      newCardUp.style.zIndex     = '80';
    }
    if (newCardKO) {
      newCardKO.style.transition = 'none';
      newCardKO.style.opacity    = '0';
      newCardKO.style.transform  = 'translateX(-115vw)';
    }

    // 4. Forzar un reflow para que el browser registre el estado inicial
    if (newCardUp) void newCardUp.offsetHeight;
    if (newCardKO) void newCardKO.offsetHeight;

    // 5. Activar transiciones en el siguiente frame de pintura
    requestAnimationFrame(() => {
      requestAnimationFrame(() => {
        // Tarjeta que SUBIÓ: desliza hacia arriba a su posición final
        if (newCardUp) {
          newCardUp.style.transition = 'transform 0.45s cubic-bezier(.22,1,.36,1)';
          newCardUp.style.transform  = 'translateY(0)';
        }
        // Tarjeta que BAJÓ: entra desde la izquierda a su slot correcto
        if (newCardKO) {
          newCardKO.style.transition = 'transform 0.38s cubic-bezier(.22,1,.36,1), opacity 0.15s ease';
          newCardKO.style.transform  = 'translateX(0)';
          newCardKO.style.opacity    = '1';
        }
      });
    });

    // 6. Cleanup: quitar todos los estilos inline cuando terminen las transiciones
    setTimeout(() => {
      [newCardUp, newCardKO].forEach(c => {
        if (!c) return;
        c.style.transition = '';
        c.style.transform  = '';
        c.style.opacity    = '';
        c.style.zIndex     = '';
      });
      glove.style.display    = 'none';
      glove.style.opacity    = '';
      glove.style.transform  = '';
      glove.style.transition = '';
      koAnimEnCurso = false;
    }, 550);

  }, T_RERENDER);
}

// ── Helper: vibraciónde jab ──
function jabCard(card) {
  card.style.animation = 'none';
  void card.offsetWidth;
  card.style.animation = 'jabVibrate 0.16s ease-in-out';
}

function renderCajaList(caja) {
  const txs=txData.filter(t=>t.caja===caja).slice().reverse();
  const total=txs.reduce((s,t)=>s+t.amount,0);
  const tel=document.getElementById('ctotal'+caja);if(tel)tel.innerHTML=fmtDisplay(total);
  const list=document.getElementById('txlist'+caja);if(!list)return;
  if(!txs.length){list.innerHTML='<div class="no-tx">Sin operaciones aun</div>';return;}
  list.innerHTML=txs.map(t=>`
    <div class="tx-item">
      <div class="tx-info">
        <div class="tx-name">${t.name}${t.mesa?' <span style="color:#3a3a3a;font-size:11px">M'+t.mesa+'</span>':''}</div>
        <div class="tx-meta">${t.time}</div>
      </div>
      <div class="tx-right"><span class="tx-amount">${fmtDisplay(t.amount)}</span><button class="btn-edit" onclick="editTx(${t.id},${t.amount},'${String(t.name).replace(/'/g,"\\'")}')">✎</button><button class="btn-del" onclick="deleteTx(${t.id})">✕</button></div>
    </div>`).join('');
}

function updateNamelists() {
  const names=[...new Set(txData.map(t=>t.name))];
  for(let c=1;c<=3;c++){
    ['nl','mnl'].forEach(prefix=>{
      const dl=document.getElementById(prefix+c);
      if(dl)dl.innerHTML=names.map(n=>`<option value="${n}">`).join('');
    });
  }
}

// CONFIGURACION TARJETAS
function renderConfigTarjetas() {
  const grid=document.getElementById('tarjetas-grid');
  grid.innerHTML=confTarjetas.map((t,i)=>{
    const configurada=t.codigo;
    const saldoInfo=tarjetasData[t.codigo];
    const saldoIni=parseFloat(t.saldo_inicial||0);
    const saldoAct=saldoInfo!==undefined?saldoInfo.saldo_actual:saldoIni;
    const pct=saldoIni>0?Math.max(0,Math.round((saldoAct/saldoIni)*100)):0;
    const hasPending = pendingCodigos[i] !== undefined;
    return `<div class="tarjeta-conf ${configurada?'configurada':''}" id="tc-${i}">
      <div class="tc-header">
        <span class="tc-num">Tarjeta ${t.slot} — Mesa ${t.slot}</span>
        <div class="tc-btns">
          <span class="tc-status" id="tc-status-${i}" style="${configurada?'color:#3a9a5a':''}">${configurada?'Vinculada':'Sin vincular'}</span>
          <button class="tc-scan-btn" id="scan-btn-${i}" onclick="iniciarScan(${i})">LEER</button>
          <button class="tc-confirm-btn" id="confirm-btn-${i}" onclick="confirmarVinculo(${i})" style="display:${hasPending?'inline-block':'none'}">CONFIRMAR</button>
          <button class="tc-clear-btn" onclick="clearSlot(${i})" title="Borrar tarjeta">✕</button>
        </div>
      </div>
      <div class="tc-field">
        <span class="tc-label">Nombre del cliente</span>
        <input class="tc-input" id="tc-nombre-${i}" type="text" placeholder="Nombre..." value="${t.nombre_cliente||''}"
          oninput="confTarjetas[${i}].nombre_cliente=this.value" />
      </div>
      <div class="tc-field">
        <span class="tc-label">Saldo inicial</span>
        <input class="tc-input" id="tc-saldo-${i}" type="number" placeholder="" value="${t.saldo_inicial}"
          oninput="confTarjetas[${i}].saldo_inicial=this.value"
          onkeydown="if(event.key==='Enter'){event.preventDefault();const v=this.value;if(v>0){showToast('Monto $'+fmt(v).replace('$','')+' agregado a Tarjeta ${t.slot}');}}" />
      </div>
      <div class="tc-code" id="tc-code-${i}">${hasPending?'Leida: '+pendingCodigos[i]:(t.codigo?'Cod: '+t.codigo:'Presiona LEER y pasa la tarjeta')}</div>
      ${saldoIni>0?`<div class="tc-saldo-bar">
        <div class="tc-saldo-info"><span class="tc-saldo-used">Gastado: ${fmt(saldoIni-saldoAct)}</span><span class="tc-saldo-left">Disponible: ${fmt(saldoAct)}</span></div>
        <div class="tc-bar-wrap"><div class="tc-bar-fill" style="width:${pct}%"></div></div>
      </div>`:''}
    </div>`;
  }).join('');
}

function iniciarScan(idx) {
  if(scanSlotActivo!==null){const pb=document.getElementById('scan-btn-'+scanSlotActivo);if(pb){pb.textContent='LEER';pb.classList.remove('activo');}}
  scanSlotActivo=idx;
  const btn=document.getElementById('scan-btn-'+idx);
  btn.textContent='ESPERANDO...';btn.classList.add('activo');
  showToast('Pasa la tarjeta por el lector');
  setTimeout(()=>{if(scanSlotActivo===idx){scanSlotActivo=null;btn.textContent='LEER';btn.classList.remove('activo');}},10000);
}
function clearSlot(idx){confTarjetas[idx].codigo='';renderConfigTarjetas();}

async function guardarTarjetas() {
  try {
    const res=await fetch('/api/tarjetas/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(confTarjetas)});
    if(res.ok){showToast('Configuracion guardada');await loadData();renderConfigTarjetas();}
    else{
      const err=await res.json();
      showToast(err.error||'Error al guardar',true);
    }
  } catch(e){showToast('Error al guardar',true);}
}
async function limpiarTarjetas() {
  if(!confirm('Limpiar toda la configuracion?'))return;
  confTarjetas=Array.from({length:30},(_,i)=>({slot:i+1,codigo:'',saldo_inicial:''}));
  await fetch('/api/tarjetas/limpiar',{method:'POST'});
  renderConfigTarjetas();showToast('Tarjetas limpiadas');
}

async function cargarConfTarjetas() {
  try{const res=await fetch('/api/tarjetas/config');if(res.ok){const d=await res.json();if(d&&d.length)confTarjetas=d;}}catch(e){}
}

// ══════════════════════════════════════════
//  STATS
// ══════════════════════════════════════════
function renderStats() {
  if (!document.getElementById('tab-stats').classList.contains('active')) return;
  if (!txData.length) {
    document.getElementById('stats-sub').textContent = 'Sin datos todavía — registrá consumos para ver las estadísticas.';
    document.getElementById('kpi-total').textContent = '$0';
    document.getElementById('kpi-ops').textContent = '0';
    document.getElementById('kpi-avg').textContent = '$0';
    document.getElementById('stats-top-clientes').innerHTML = '<div style="color:#333;font-size:13px;padding:14px 0;">Sin datos aún</div>';
    document.getElementById('cajas-detail').innerHTML = '';
    return;
  }

  const total = txData.reduce((s,t)=>s+t.amount,0);
  const ops = txData.length;
  const avg = Math.round(total/ops);

  // Consumo por hora
  const porHora = {};
  txData.forEach(t => {
    const h = t.time ? t.time.split(':')[0] : '??';
    porHora[h] = (porHora[h]||0) + t.amount;
  });
  const horaMax = Object.entries(porHora).sort((a,b)=>b[1]-a[1])[0];

  // Primer y último registro
  const horas = txData.map(t=>t.time).filter(Boolean).sort();
  const subTxt = horas.length ? 'Primera operación: ' + horas[0] + '  •  Última: ' + horas[horas.length-1] : '';
  document.getElementById('stats-sub').textContent = subTxt;
  document.getElementById('kpi-total').innerHTML = fmtDisplay(total);
  document.getElementById('kpi-ops').textContent = ops;
  document.getElementById('kpi-avg').innerHTML = fmtDisplay(avg);

  // ─── Gráfico por caja ───
  const porCaja = {1:0, 2:0, 3:0};
  const opsCaja = {1:0, 2:0, 3:0};
  txData.forEach(t=>{ if(t.caja>=1&&t.caja<=3){porCaja[t.caja]+=t.amount;opsCaja[t.caja]++;} });
  const cajasColors = ['#c9a227','#e8c84a','#7a6010'];
  const cajasLabels = ['Abajo','Extendido','VIP'];
  drawBarChart('chart-cajas', [porCaja[1],porCaja[2],porCaja[3]], cajasLabels, cajasColors);
  document.getElementById('legend-cajas').innerHTML = cajasLabels.map((l,i)=>
    `<div class="legend-item"><div class="legend-dot" style="background:${cajasColors[i]}"></div>${l}: ${fmt(porCaja[i+1])} (${opsCaja[i+1]} ops)</div>`
  ).join('');

  // ─── Gráfico por hora ───
  // Ordenar horas cronológicamente (noche: 20-06)
  const allHoras = Array.from({length:12},(_,i)=>String((20+i)%24).padStart(2,'0'));
  const valHoras = allHoras.map(h=>porHora[h]||0);
  const horasLabels = allHoras.map(h=>h+'h');
  drawBarChart('chart-horas', valHoras, horasLabels, valHoras.map((_,i)=>allHoras[i]===horaMax?.[0]?'#c9a227':'#2a2a2a'));

  // ─── Top clientes ───
  const totals={}, mesas={};
  txData.forEach(t=>{totals[t.name]=(totals[t.name]||0)+t.amount; if(t.mesa&&!mesas[t.name])mesas[t.name]=t.mesa;});
  const sorted = Object.entries(totals).sort((a,b)=>b[1]-a[1]).slice(0,8);
  const maxVal = sorted[0]?.[1]||1;
  const esc = s=>String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  document.getElementById('stats-top-clientes').innerHTML = '<div class="top-list">' +
    sorted.map(([name,val],i)=>`
      <div class="top-item pos-${i+1}">
        <div class="top-pos">#${i+1}</div>
        <div style="flex:1">
          <div class="top-name">${esc(name)}</div>
          ${mesas[name]?`<div class="top-mesa">Mesa ${esc(mesas[name])}</div>`:''}
        </div>
        <div class="top-bar-wrap"><div class="top-bar-fill" style="width:${Math.round(val/maxVal*100)}%"></div></div>
        <div class="top-amount">${fmtDisplay(val)}</div>
      </div>`).join('') + '</div>';

  // ─── Detalle por caja ───
  document.getElementById('cajas-detail').innerHTML = [1,2,3].map(c=>{
    const txs = txData.filter(t=>t.caja===c);
    const tot = txs.reduce((s,t)=>s+t.amount,0);
    const items = txs.slice().reverse().slice(0,6);
    return `<div class="caja-stat-card">
      <div class="caja-stat-badge">CAJA ${c}</div>
      <div class="caja-stat-total">${fmtDisplay(tot)}</div>
      <div class="caja-stat-ops">${txs.length} operaciones</div>
      <div class="caja-stat-list">${
        items.length
          ? items.map(t=>`<div class="caja-stat-item"><span>${esc(t.name)} <span style="color:#2a2a2a">${t.time||''}</span></span><span>${fmtDisplay(t.amount)}</span></div>`).join('')
          : '<div style="color:#222;font-size:12px;padding:8px 0;">Sin operaciones</div>'
      }</div>
    </div>`;
  }).join('');
}

function drawBarChart(canvasId, values, labels, colors) {
  const canvas = document.getElementById(canvasId);
  if (!canvas) return;
  const ctx = canvas.getContext('2d');
  const W = canvas.offsetWidth || 300;
  canvas.width = W;
  const H = canvas.height;
  ctx.clearRect(0,0,W,H);
  const max = Math.max(...values, 1);
  const n = values.length;
  const padL=6, padR=6, padT=10, padB=28;
  const barW = Math.floor((W - padL - padR) / n);
  const gap = Math.max(2, Math.floor(barW*0.15));
  const bw = barW - gap;
  values.forEach((v,i)=>{
    const x = padL + i*barW + Math.floor(gap/2);
    const barH = Math.round((v/max)*(H-padT-padB));
    const y = H - padB - barH;
    const col = Array.isArray(colors)?colors[i%colors.length]:colors;
    // Barra
    ctx.fillStyle = v>0 ? col : '#1a1a1a';
    ctx.beginPath();
    ctx.roundRect ? ctx.roundRect(x, y, bw, barH, [3,3,0,0]) : ctx.rect(x, y, bw, barH);
    ctx.fill();
    // Label abajo
    ctx.fillStyle = '#444';
    ctx.font = `${Math.min(10, Math.floor(barW*0.55))}px Arial`;
    ctx.textAlign = 'center';
    ctx.fillText(labels[i], x + bw/2, H-padB+14);
    // Valor encima si la barra es visible
    if (v>0 && barH>18) {
      ctx.fillStyle = '#888';
      ctx.font = '9px Arial';
      const label = v>=1000000?'$'+(v/1000000).toFixed(1)+'M':v>=1000?'$'+(v/1000).toFixed(0)+'k':'$'+v;
      ctx.fillText(label, x+bw/2, y-3);
    }
  });
}


// ══════════════════════════════════════════
//  TEMAS NUEVOS: Jagger 12 años y Velada Boxeo
// ══════════════════════════════════════════
const TEMAS_EXTRA = {
  jagger12: {
    colors: {'--black':'#000000','--surface':'#0a0a0a','--border':'#333333','--gold':'#ffffff','--gold-light':'#dddddd','--gold-dim':'#888888','--text':'#e8e8e8','--text-dim':'#666666','--white':'#ffffff'},
    bodyClass: 'tema-jagger12',
    particleLabel: 'Activar burbujas de champagne'
  },
  jagger12boxeo: {
    colors: {'--black':'#050000','--surface':'#0d0000','--border':'#3a0000','--gold':'#ff2222','--gold-light':'#ff5555','--gold-dim':'#880000','--text':'#f0d0d0','--text-dim':'#6a3333','--white':'#fff0f0'},
    bodyClass: 'tema-jagger12boxeo',
    particleLabel: 'Activar efectos de ring'
  },
  touchofpink: {
    colors: {'--black':'#2d0020','--surface':'#480035','--border':'#8a3070','--gold':'#f472b6','--gold-light':'#fbb6ce','--gold-dim':'#e896cc','--text':'#ffe8f5','--text-dim':'#ddaacc','--white':'#ffffff'},
    bodyClass: 'tema-touchofpink',
    particleLabel: 'Activar pétalos animados'
  }
};

let temaActual = 'default';
let decoActiva = true;
let punchAnimActiva = false;
let koAnimActiva = true;
let fallingGlovesActivos = true;
let mostrar12Fondo = true;
let svg12Opacity = 0.13;
let svg12Color = 'white';
let svg12GlowBlur = 18;
let confettiGanadorActivo = true;
let tipoParticula = 'confetti'; // 'confetti' | 'billetes' | 'ninguno'
let pinkPetalosActivos = true;
let pinkModoClaro = false;

function aplicarTema(nombre) {
  const temaExtra = TEMAS_EXTRA[nombre];
  document.body.classList.remove('tema-fullblack','tema-navidad','tema-anonuevo','tema-halloween','tema-jagger12','tema-jagger12boxeo','tema-touchofpink','pink-claro');
  pinkModoClaro = false;
  const overlay = document.getElementById('tema-overlay');
  overlay.innerHTML = ''; overlay.style.opacity = '0';
  const tl = document.getElementById('tema-tagline');

  if (temaExtra) {
    COLOR_DEFS.forEach(c => {
      const v = temaExtra.colors[c.key] || c.default;
      document.documentElement.style.setProperty(c.key, v);
      customColors[c.key] = v;
    });
    // surface-gold tinted by new surface
    document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(temaExtra.colors['--surface'] || '#111'));
    if (temaExtra.bodyClass) document.body.classList.add(temaExtra.bodyClass);
    temaActual = nombre;
    // Show/hide punch animation toggle
    const fallingToggle = document.getElementById('falling-gloves-toggle');
    if (fallingToggle) fallingToggle.style.display = nombre === 'jagger12boxeo' ? 'block' : 'none';
    const koToggle = document.getElementById('ko-anim-toggle');
    if (koToggle) koToggle.style.display = nombre === 'jagger12boxeo' ? 'block' : 'none';
    const show12Toggle = document.getElementById('show-12-toggle');
    if (show12Toggle) show12Toggle.style.display = (nombre === 'jagger12' || nombre === 'jagger12boxeo') ? 'block' : 'none';
    const pinkPetTog = document.getElementById('pink-petalos-toggle');
    if (pinkPetTog) pinkPetTog.style.display = nombre === 'touchofpink' ? 'block' : 'none';
    const pinkModoTog = document.getElementById('pink-modo-toggle');
    if (pinkModoTog) pinkModoTog.style.display = nombre === 'touchofpink' ? 'flex' : 'none';
    // Update logo
    const logoVip = document.getElementById('logo-vip');
    const savedVip = (()=>{ try { return localStorage.getItem('rankingVIP_vip') || 'VIP'; } catch(e){ return 'VIP'; } })();
    // Restore standard RANKING VIP logo (no club span)
    const mainLogo = document.getElementById('main-logo');
    if (mainLogo) mainLogo.innerHTML = `RANKING <span class="vip" id="logo-vip">${savedVip}</span>`;
    if (tl) {
      if (nombre === 'jagger12boxeo') {
        tl.innerHTML = `<div style="text-align:center;line-height:1.3;">
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(14px,2.2vw,24px);font-weight:700;letter-spacing:4px;color:#ff2222;text-shadow:0 0 14px rgba(255,34,34,0.9),0 0 30px rgba(255,34,34,0.5);">JAGGER CLUB · 12 AÑOS</div>
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(11px,1.6vw,17px);font-weight:600;letter-spacing:3px;color:#ff2222;opacity:0.75;text-shadow:0 0 10px rgba(255,34,34,0.6);margin-top:2px;">12 AÑOS DE HISTORIA NO SON PARA CUALQUIERA</div>
        </div>`;
      } else if (nombre === 'touchofpink') {
        tl.innerHTML = `<div style="text-align:center;line-height:1.3;">
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(18px,2.8vw,30px);font-weight:700;letter-spacing:6px;color:#ffffff;text-shadow:0 0 12px rgba(255,255,255,0.9),0 0 28px rgba(255,255,255,0.5),0 0 50px rgba(244,114,182,0.4);">JAGGER CLUB</div>
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(13px,2vw,22px);font-weight:600;letter-spacing:5px;color:#f472b6;margin-top:3px;text-shadow:0 0 10px rgba(244,114,182,1),0 0 24px rgba(244,114,182,0.7),0 0 50px rgba(244,114,182,0.4);">TURNS PINK</div>
        </div>`;
      } else {
        tl.innerHTML = `<div style="text-align:center;line-height:1.3;">
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(14px,2.2vw,24px);font-weight:700;letter-spacing:4px;color:#e8c84a;text-shadow:0 0 14px rgba(232,200,74,0.9),0 0 30px rgba(201,162,39,0.6);">JAGGER CLUB · 12 AÑOS</div>
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(11px,1.6vw,17px);font-weight:600;letter-spacing:3px;color:#e8c84a;opacity:0.75;text-shadow:0 0 10px rgba(201,162,39,0.5);margin-top:2px;">12 AÑOS DE HISTORIA NO SON PARA CUALQUIERA</div>
        </div>`;
      }
    }
    // Toggle deco
    const toggleWrap = document.getElementById('tema-deco-toggle');
    if (toggleWrap) {
      toggleWrap.style.display = 'flex';
      const decoLabel = nombre === 'jagger12' ? 'Activar burbujas de champagne' :
                        nombre === 'jagger12boxeo' ? 'Activar chispas y efectos del ring' :
                        'Activar decoraciones animadas';
      document.getElementById('toggle-deco-label').textContent = decoLabel;
      document.getElementById('toggle-deco').checked = decoActiva;
      const decoMainLabel = document.getElementById('deco-main-label');
      if (decoMainLabel) decoMainLabel.style.display = nombre === 'touchofpink' ? 'none' : '';
    }
    if (decoActiva) iniciarDecoTema(nombre);
    buildColorGrid();
    try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
    showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
    return;
  }

  if (nombre === 'default') {
    COLOR_DEFS.forEach(c => document.documentElement.style.setProperty(c.key, c.default));
    document.documentElement.style.setProperty('--surface-gold', '#0d0b00');
    customColors = {};
    COLOR_DEFS.forEach(c => customColors[c.key] = c.default);
    const lv = document.getElementById('logo-vip'); if(lv) lv.textContent='VIP';
    // Restore main logo without club
    const mainLogo2 = document.getElementById('main-logo');
    if (mainLogo2) mainLogo2.innerHTML = `RANKING <span class="vip" id="logo-vip">VIP</span>`;
    if (tl) { tl.textContent = 'JAGGER CLUB'; tl.style.fontSize='28px'; tl.style.fontWeight='600'; tl.style.letterSpacing='5px'; tl.style.color='#555'; }
  }
  temaActual = nombre;
  const toggleWrap = document.getElementById('tema-deco-toggle');
  if (toggleWrap) toggleWrap.style.display = 'none';
  buildColorGrid();
  try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
  showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
}

function toggleDecoActual(checked) {
  decoActiva = checked;
  if (temaActual === 'jagger12boxeo') {
    // En boxeo, el toggle controla chispas pero los guantes de fondo siempre están
    if (!checked) {
      const wrap = document.getElementById('boxing-particles');
      if (wrap) wrap.innerHTML = ''; // limpiar chispas/lluvia
    } else {
      iniciarChispasBoxeo();
      if (fallingGlovesActivos) iniciarLluviaGuantes();
    }
  } else if (temaActual === 'jagger12') {
    // En aniversario, solo controla las burbujas animadas (no el 12 de fondo)
    const wrap = document.getElementById('jagger12-particles');
    if (!checked) {
      if (wrap) wrap.innerHTML = ''; // detener burbujas
    } else {
      // Reiniciar solo burbujas
      if (wrap) {
        wrap.innerHTML = '';
        function lanzarBurbuja() {
          if (!decoActiva) return; // Stop when deactivated
          if (!document.getElementById('jagger12-particles')) return;
          const el = document.createElement('div');
          const sz = 4 + Math.random() * 12;
          const isGold = Math.random() > 0.7;
          el.style.cssText = `position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.35:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`;
          wrap.appendChild(el);
          setTimeout(()=>el.remove(), 11000);
          setTimeout(lanzarBurbuja, 200+Math.random()*600);
        }
        for(let i=0;i<8;i++) setTimeout(lanzarBurbuja, i*120);
      }
    }
  } else {
    const overlay = document.getElementById('tema-overlay');
    if (!checked) { overlay.innerHTML = ''; overlay.style.opacity = '0'; }
    else iniciarDecoTema(temaActual);
  }
}

// Actualizar solo el SVG del 12 sin tocar las burbujas ni las decoraciones
function actualizar12Overlay() {
  if (temaActual === 'jagger12') {
    const overlay = document.getElementById('tema-overlay');
    const existing12 = overlay ? overlay.querySelector('svg:not([id])') : null;
    // Remove any existing 12 SVGs (identified by having the "12" text)
    if (overlay) {
      overlay.querySelectorAll('svg').forEach(svg => {
        if (svg.textContent && svg.textContent.trim() === '12') svg.remove();
      });
    }
    if (mostrar12Fondo && overlay) {
      const svgNS = 'http://www.w3.org/2000/svg';
      const svg = document.createElementNS(svgNS, 'svg');
      svg.setAttribute('width','100%'); svg.setAttribute('height','100%');
      svg.setAttribute('viewBox','0 0 1000 600');
      svg.style.cssText = `position:absolute;inset:0;pointer-events:none;opacity:${svg12Opacity};`;
      svg.setAttribute('preserveAspectRatio','xMidYMid meet');
      const defs = document.createElementNS(svgNS,'defs');
      defs.innerHTML = `<filter id="blur12j"><feGaussianBlur stdDeviation="8"/></filter><filter id="glow12j"><feGaussianBlur stdDeviation="${svg12GlowBlur}" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>`;
      svg.appendChild(defs);
      const t1 = document.createElementNS(svgNS,'text');
      t1.setAttribute('x','500'); t1.setAttribute('y','380'); t1.setAttribute('text-anchor','middle');
      t1.setAttribute('dominant-baseline','middle'); t1.setAttribute('font-family','Oswald,Arial');
      t1.setAttribute('font-weight','700'); t1.setAttribute('font-size','560');
      t1.setAttribute('fill',`rgba(201,162,39,0.15)`); t1.setAttribute('letter-spacing','-10');
      t1.setAttribute('filter','url(#blur12j)'); t1.textContent = '12';
      const t2 = document.createElementNS(svgNS,'text');
      t2.setAttribute('x','500'); t2.setAttribute('y','380'); t2.setAttribute('text-anchor','middle');
      t2.setAttribute('dominant-baseline','middle'); t2.setAttribute('font-family','Oswald,Arial');
      t2.setAttribute('font-weight','700'); t2.setAttribute('font-size','560');
      t2.setAttribute('fill',svg12Color); t2.setAttribute('letter-spacing','-10');
      t2.setAttribute('filter','url(#glow12j)'); t2.textContent = '12';
      svg.appendChild(t1); svg.appendChild(t2);
      // Insert after particles wrap
      const particles = overlay.querySelector('#jagger12-particles');
      if (particles && particles.nextSibling) overlay.insertBefore(svg, particles.nextSibling);
      else overlay.appendChild(svg);
    }
  } else if (temaActual === 'jagger12boxeo') {
    reiniciarDeco12();
  }
}

function iniciarDecoTema(nombre) {
  const overlay = document.getElementById('tema-overlay');
  overlay.style.opacity = '1';
  if (nombre === 'jagger12') {
    overlay.innerHTML = '';
    iniciarJagger12Deco();
  }
  if (nombre === 'jagger12boxeo') {
    iniciarJagger12BoxeoDeco();
  }
  if (nombre === 'touchofpink') {
    overlay.innerHTML = `
      <!-- Detalles blancos y rosados: destellos -->
      <div style="position:absolute;top:12%;left:8%;font-size:14px;color:rgba(255,255,255,0.45);pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div>
      <div style="position:absolute;top:18%;right:10%;font-size:11px;color:rgba(255,255,255,0.38);pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div>
      <div style="position:absolute;top:50%;left:5%;font-size:12px;color:rgba(255,255,255,0.35);pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div>
      <div style="position:absolute;top:65%;right:6%;font-size:15px;color:rgba(255,255,255,0.4);pointer-events:none;animation:goldTwinkle 3.5s ease-in-out .5s infinite;">✦</div>
      <div style="position:absolute;top:33%;left:14%;font-size:9px;color:rgba(251,182,206,0.55);pointer-events:none;animation:goldTwinkle 4.5s ease-in-out 1.5s infinite;">✦</div>
      <div style="position:absolute;top:42%;right:16%;font-size:10px;color:rgba(251,182,206,0.5);pointer-events:none;animation:goldTwinkle 3.8s ease-in-out .8s infinite;">✦</div>
      <div style="position:absolute;top:75%;left:20%;font-size:8px;color:rgba(255,255,255,0.3);pointer-events:none;animation:goldTwinkle 6s ease-in-out 3s infinite;">✦</div>
      <div style="position:absolute;top:28%;right:22%;font-size:13px;color:rgba(244,114,182,0.4);pointer-events:none;animation:goldTwinkle 5.5s ease-in-out 2.5s infinite;">✦</div>
      <div id="petalos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>`;
    if (!document.getElementById('kf-bubble')) {
      const s = document.createElement('style'); s.id = 'kf-pink-twinkle';
      s.textContent = `@keyframes goldTwinkle{0%,100%{opacity:0.15;transform:scale(1)}50%{opacity:0.55;transform:scale(1.4)}}`;
      document.head.appendChild(s);
    }
    if (pinkPetalosActivos) iniciarPetalos();
  }
}

function iniciarJagger12Deco() {
  const overlay = document.getElementById('tema-overlay');
  const svg12 = mostrar12Fondo ? `
    <!-- 12 grande de fondo, centrado -->
    <svg width="100%" height="100%" viewBox="0 0 1000 600" style="position:absolute;inset:0;pointer-events:none;opacity:0.13;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet">
      <defs>
        <filter id="blur12j"><feGaussianBlur stdDeviation="8"/></filter>
        <filter id="glow12j">
          <feGaussianBlur stdDeviation="18" result="blur"/>
          <feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge>
        </filter>
      </defs>
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="rgba(201,162,39,0.15)" letter-spacing="-10" filter="url(#blur12j)">12</text>
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="white" letter-spacing="-10" filter="url(#glow12j)">12</text>
    </svg>` : '';
  overlay.innerHTML = `
    <div id="jagger12-particles" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
    ${svg12}
    <!-- Líneas doradas sutiles en esquinas -->
    <svg width="200" height="200" viewBox="0 0 200 200" style="position:absolute;top:0;left:0;opacity:0.18;pointer-events:none;" xmlns="http://www.w3.org/2000/svg">
      <line x1="0" y1="0" x2="120" y2="0" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="0" y2="120" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="60" y2="60" stroke="#c9a227" stroke-width="0.5"/>
      <circle cx="0" cy="0" r="3" fill="#c9a227"/>
    </svg>
    <svg width="200" height="200" viewBox="0 0 200 200" style="position:absolute;top:0;right:0;opacity:0.18;pointer-events:none;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
      <line x1="0" y1="0" x2="120" y2="0" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="0" y2="120" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="60" y2="60" stroke="#c9a227" stroke-width="0.5"/>
      <circle cx="0" cy="0" r="3" fill="#c9a227"/>
    </svg>
    <!-- Detalles dorados: pequeñas estrellas/destellos -->
    <div style="position:absolute;top:15%;left:5%;font-size:10px;color:#c9a227;opacity:0.2;pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div>
    <div style="position:absolute;top:25%;right:7%;font-size:8px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div>
    <div style="position:absolute;top:60%;left:3%;font-size:12px;color:#c9a227;opacity:0.15;pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div>
    <div style="position:absolute;top:70%;right:4%;font-size:9px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 3.5s ease-in-out 0.5s infinite;">✦</div>`;
  if (!document.getElementById('kf-bubble')) {
    const s = document.createElement('style'); s.id = 'kf-bubble';
    s.textContent = `@keyframes bubbleRise{0%{opacity:0.6;transform:translateY(0) scale(1)}50%{opacity:0.3}100%{opacity:0;transform:translateY(-100vh) scale(0.5)}}
    @keyframes goldTwinkle{0%,100%{opacity:0.1;transform:scale(1)}50%{opacity:0.25;transform:scale(1.4)}}`;
    document.head.appendChild(s);
  }
  if (decoActiva) {
    function lanzarBurbuja() {
      if (!decoActiva) return; // Stop when deactivated
      const wrap = document.getElementById('jagger12-particles'); if(!wrap) return;
      const el = document.createElement('div');
      const sz = 4 + Math.random() * 12;
      const isGold = Math.random() > 0.7;
      el.style.cssText = `position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.35:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`;
      wrap.appendChild(el);
      setTimeout(()=>el.remove(), 11000);
      setTimeout(lanzarBurbuja, 200+Math.random()*600);
    }
    for(let i=0;i<8;i++) setTimeout(lanzarBurbuja, i*120);
  }
}

function reiniciarDeco12() {
  if (temaActual === 'jagger12') iniciarJagger12Deco();
  else if (temaActual === 'jagger12boxeo') {
    // Update only the 12 SVG in boxeo overlay
    const overlay = document.getElementById('tema-overlay');
    const existing12 = overlay ? overlay.querySelector('#svg-12-boxeo') : null;
    if (!mostrar12Fondo && existing12) existing12.remove();
    else if (mostrar12Fondo && !existing12 && overlay) {
      const svg = document.createElementNS('http://www.w3.org/2000/svg','svg');
      svg.id = 'svg-12-boxeo';
      svg.setAttribute('width','100%'); svg.setAttribute('height','100%');
      svg.setAttribute('viewBox','0 0 1000 600');
      svg.style.cssText = 'position:absolute;inset:0;pointer-events:none;opacity:0.035;';
      svg.setAttribute('preserveAspectRatio','xMidYMid meet');
      svg.innerHTML = '<text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="#ff0000" letter-spacing="-10">12</text>';
      overlay.insertBefore(svg, overlay.firstChild);
    }
  }
}

function iniciarJagger12BoxeoDeco() {
  const overlay = document.getElementById('tema-overlay');
  overlay.innerHTML = `
    <div id="boxing-particles" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
    <!-- Cuerdas del ring horizontales -->
    <svg style="position:absolute;inset:0;width:100%;height:100%;pointer-events:none;opacity:0.07;" xmlns="http://www.w3.org/2000/svg">
      <line x1="0" y1="28%" x2="100%" y2="28%" stroke="#ff2222" stroke-width="3"/>
      <line x1="0" y1="48%" x2="100%" y2="48%" stroke="#cc0000" stroke-width="2"/>
      <line x1="0" y1="68%" x2="100%" y2="68%" stroke="#ff2222" stroke-width="3"/>
      <!-- Postes del ring en las esquinas -->
      <line x1="2%" y1="20%" x2="2%" y2="76%" stroke="#888" stroke-width="4"/>
      <line x1="98%" y1="20%" x2="98%" y2="76%" stroke="#888" stroke-width="4"/>
      <rect x="0.5%" y="18%" width="3%" height="4%" fill="#555" rx="2"/>
      <rect x="96.5%" y="18%" width="3%" height="4%" fill="#555" rx="2"/>
      <rect x="0.5%" y="74%" width="3%" height="4%" fill="#555" rx="2"/>
      <rect x="96.5%" y="74%" width="3%" height="4%" fill="#555" rx="2"/>
    </svg>
    <!-- 12 grande de fondo -->
    <svg id="svg-12-boxeo" width="100%" height="100%" viewBox="0 0 1000 600" style="position:absolute;inset:0;pointer-events:none;opacity:0.035;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet">
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="#ff0000" letter-spacing="-10">12</text>
    </svg>
    <!-- Guantes grandes en las esquinas inferiores — SIEMPRE VISIBLES -->
    <div style="position:absolute;bottom:0;left:0;font-size:130px;opacity:0.15;transform:rotate(-20deg) translateY(20px);pointer-events:none;filter:hue-rotate(0deg);">🥊</div>
    <div style="position:absolute;bottom:0;right:0;font-size:130px;opacity:0.15;transform:rotate(20deg) translateY(20px) scaleX(-1);pointer-events:none;">🥊</div>
    <!-- Guantes medianos en las esquinas superiores — SIEMPRE VISIBLES -->
    <div style="position:absolute;top:60px;left:20px;font-size:55px;opacity:0.1;transform:rotate(15deg);pointer-events:none;">🥊</div>
    <div style="position:absolute;top:60px;right:20px;font-size:55px;opacity:0.1;transform:rotate(-15deg) scaleX(-1);pointer-events:none;">🥊</div>
    <!-- Guantes adicionales en el medio — SIEMPRE VISIBLES -->
    <div style="position:absolute;top:40%;left:5%;font-size:70px;opacity:0.06;transform:rotate(-10deg);pointer-events:none;animation:guanteFloat 6s ease-in-out infinite;">🥊</div>
    <div style="position:absolute;top:35%;right:5%;font-size:70px;opacity:0.06;transform:rotate(10deg) scaleX(-1);pointer-events:none;animation:guanteFloat 7s ease-in-out 1s infinite;">🥊</div>
    <div style="position:absolute;bottom:20%;left:15%;font-size:45px;opacity:0.07;transform:rotate(5deg);pointer-events:none;animation:guanteFloat 5s ease-in-out 2s infinite;">🥊</div>
    <div style="position:absolute;bottom:20%;right:15%;font-size:45px;opacity:0.07;transform:rotate(-5deg) scaleX(-1);pointer-events:none;animation:guanteFloat 8s ease-in-out 0.5s infinite;">🥊</div>
    <!-- Texto ROUND en la parte superior -->
    <div id="boxing-round-text" style="position:absolute;top:8px;left:50%;transform:translateX(-50%);font-family:'Oswald',sans-serif;font-size:13px;color:#ff0000;opacity:0.18;letter-spacing:8px;text-transform:uppercase;white-space:nowrap;pointer-events:none;animation:roundPulse 3s ease-in-out infinite;">⚔ ROUND 1 — FIGHT! ⚔</div>
    <!-- Campana del ring - icono -->
    <div style="position:absolute;top:8px;right:20px;font-size:22px;opacity:0.1;pointer-events:none;animation:bellShake 4s ease-in-out infinite;">🔔</div>`;
  if (!document.getElementById('kf-spark-boxing')) {
    const s = document.createElement('style'); s.id = 'kf-spark-boxing';
    s.textContent = `@keyframes sparkBox{0%{opacity:1;transform:scale(1)}100%{opacity:0;transform:translate(var(--bx),var(--by)) scale(0)}}
    @keyframes roundPulse{0%,100%{opacity:0.18}50%{opacity:0.32}}
    @keyframes bellShake{0%,90%,100%{transform:rotate(0deg)}92%{transform:rotate(-15deg)}96%{transform:rotate(15deg)}}
    @keyframes guanteFloat{0%,100%{transform:rotate(-10deg) translateY(0)}50%{transform:rotate(-10deg) translateY(-12px)}}
    @keyframes guanteFall{0%{opacity:0.8;transform:translateY(-40px) rotate(0deg)}100%{opacity:0;transform:translateY(105vh) rotate(360deg)}}`;
    document.head.appendChild(s);
  }
  // Chispas animadas (controladas por toggle-deco)
  if (decoActiva) iniciarChispasBoxeo();
  // Lluvia de guantes (controlada por falling-gloves toggle)
  if (fallingGlovesActivos) iniciarLluviaGuantes();
}

function iniciarChispasBoxeo() {
  function lanzarChispa() {
    if (!decoActiva) return; // Stop when deactivated
    const wrap = document.getElementById('boxing-particles'); if(!wrap) return;
    const el = document.createElement('div');
    const bx = (Math.random()-0.5)*100, by = (Math.random()-0.5)*100;
    const sz = 3+Math.random()*6;
    el.style.cssText = `position:absolute;left:${20+Math.random()*60}%;top:${20+Math.random()*60}%;width:${sz}px;height:${sz}px;border-radius:50%;background:#ff3333;box-shadow:0 0 8px #ff0000,0 0 16px rgba(255,0,0,0.4);animation:sparkBox 0.8s ease-out forwards;--bx:${bx}px;--by:${by}px;pointer-events:none;`;
    wrap.appendChild(el);
    setTimeout(()=>el.remove(), 900);
    setTimeout(lanzarChispa, 1200+Math.random()*2500);
  }
  lanzarChispa(); setTimeout(lanzarChispa,500); setTimeout(lanzarChispa,1000); setTimeout(lanzarChispa,1500);
}

function iniciarLluviaGuantes() {
  function lanzarGuante() {
    if (!fallingGlovesActivos) return; // Stop the recursive loop when deactivated
    const wrap = document.getElementById('boxing-particles'); if(!wrap) return;
    const el = document.createElement('div');
    const sz = 18 + Math.random() * 22;
    const dur = 5 + Math.random() * 6;
    el.style.cssText = `position:absolute;top:-50px;left:${Math.random()*100}%;font-size:${sz}px;animation:guanteFall ${dur}s linear forwards;pointer-events:none;opacity:0.55;`;
    el.textContent = '🥊';
    wrap.appendChild(el);
    setTimeout(()=>el.remove(), dur*1000+200);
    setTimeout(lanzarGuante, 600+Math.random()*1200);
  }
  lanzarGuante(); setTimeout(lanzarGuante,800); setTimeout(lanzarGuante,1600);
}

function reiniciarDecoBoxeo() {
  const wrap = document.getElementById('boxing-particles');
  if (!wrap) return;
  // Solo limpiamos las partículas volantes, no el overlay entero
  wrap.innerHTML = '';
  if (decoActiva) iniciarChispasBoxeo();
  if (fallingGlovesActivos) iniciarLluviaGuantes();
}

// ══════════════════════════════════════════
//  CARTEL
// ══════════════════════════════════════════
const FRASES_RAPIDAS = [
  '🍾 SACÓ UN NUVO CON BENGALAS',
  '🥂 PIDIÓ CHAMPAGNE',
  '🎉 SACÓ BOTELLA VIP',
  '🔥 ARRANCÓ EL SHOW',
  '💎 MODO VIP ACTIVADO',
  '🚀 NIVEL ÉLITE'
];

function contarEmojis(str) {
  return [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(str)].length;
}
function limitarEmojis(input) {
  const segs = [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(input.value)];
  if (segs.length > 3) input.value = segs.slice(0,3).map(s=>s.segment).join('');
}
function agregarEmojiCartel(emoji) {
  const inp = document.getElementById('cartel-emoji-input');
  const segs = [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(inp.value)];
  if (segs.length < 3) inp.value = segs.map(s=>s.segment).join('') + emoji;
  else showToast('Máximo 3 emojis', true);
}
function abrirCartelModal() {
  document.getElementById('cartel-modal').style.display = 'flex';
}
function cerrarCartelModal() { document.getElementById('cartel-modal').style.display = 'none'; }

function mostrarCartel() {
  const nombre = document.getElementById('cartel-nombre').value.trim().toUpperCase();
  const mesa = document.getElementById('cartel-mesa').value.trim();
  const frase = document.getElementById('cartel-frase').value.trim().toUpperCase();
  if (!frase) { showToast('Escribí una frase para el cartel', true); return; }
  document.getElementById('cartel-nombre-display').textContent = nombre;
  document.getElementById('cartel-nombre-display').style.display = nombre ? 'block' : 'none';
  document.getElementById('cartel-mesa-display').textContent = mesa ? 'MESA ' + mesa : '';
  document.getElementById('cartel-frase-display').textContent = frase;
  const emojiManual = document.getElementById('cartel-emoji-input').value.trim();
  const emojis = emojiManual || (
    frase.includes('NUVO') || frase.includes('BOTELLA') ? '🍾' :
    frase.includes('BENGALA') ? '🎆' :
    frase.includes('CHAMPAGNE') ? '🍾' :
    frase.includes('BOXEO') || frase.includes('FIGHT') ? '🥊' : '🍾'
  );
  const emojiEl = document.getElementById('cartel-emoji-big');
  emojiEl.textContent = emojis;
  const nEmojis = [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(emojis)].length;
  emojiEl.style.fontSize = nEmojis >= 3 ? '72px' : nEmojis === 2 ? '90px' : '110px';
  emojiEl.style.letterSpacing = nEmojis > 1 ? '8px' : '0';
  const bg = document.getElementById('cartel-tema-bg');
  bg.innerHTML = temaActual === 'jagger12boxeo' ?
    `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#1a0000 0%,#050000 100%);"></div>
     <div style="position:absolute;bottom:0;left:0;font-size:160px;opacity:0.06;transform:rotate(-15deg);">🥊</div>
     <div style="position:absolute;bottom:0;right:0;font-size:160px;opacity:0.06;transform:rotate(15deg) scaleX(-1);">🥊</div>` :
    temaActual === 'jagger12' ?
    `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#111 0%,#000 100%);"></div>
     <div style="position:absolute;inset:0;background:radial-gradient(ellipse at 50% 50%,rgba(201,162,39,0.05) 0%,transparent 70%);"></div>` :
    `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#0a0a0a 0%,#000 100%);"></div>`;

  // Rayos de fondo animados - FIXED (no yellow line bug)
  const rays = document.getElementById('cartel-rays');
  if (rays) {
    rays.innerHTML = '';
    const rayColor = temaActual === 'jagger12boxeo' ? 'rgba(255,34,34,0.12)' : 'rgba(201,162,39,0.12)';
    for (let i = 0; i < 12; i++) {
      const angle = i * 30;
      const r = document.createElement('div');
      r.style.cssText = `position:absolute;left:50%;top:50%;width:1px;height:55vh;background:linear-gradient(to bottom,${rayColor},transparent);transform-origin:0% 0%;transform:rotate(${angle}deg);opacity:0.5;animation:rayPulse2 ${2+i*0.15}s ease-in-out ${i*0.1}s infinite alternate;`;
      rays.appendChild(r);
    }
    if (!document.getElementById('kf-ray2')) {
      const s = document.createElement('style'); s.id = 'kf-ray2';
      s.textContent = '@keyframes rayPulse2{0%{opacity:0.2;transform:rotate(var(--r,0deg)) scaleY(0.5)}100%{opacity:0.7;transform:rotate(var(--r,0deg)) scaleY(1)}}';
      document.head.appendChild(s);
    }
    // Set --r custom property on each ray
    rays.querySelectorAll('div').forEach((r,i) => {
      r.style.setProperty('--r', (i*30)+'deg');
    });
  }

  cerrarCartelModal();
  const overlay = document.getElementById('cartel-overlay');
  overlay.style.display = 'flex';
  // Trigger animation reliably on ALL screens
  const c = document.getElementById('cartel-content');
  if (c) {
    c.style.animation = 'none';
    c.style.opacity = '0';
    c.style.transform = 'scale(0.5) translateY(60px)';
    requestAnimationFrame(() => {
      requestAnimationFrame(() => {
        c.style.animation = 'winnerEntrada 0.8s cubic-bezier(.22,1,.36,1) forwards';
        c.style.opacity = '';
        c.style.transform = '';
      });
    });
  }
  // Sincronizar con otras pantallas
  fetch('/api/cartel/show',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre,mesa,frase,emoji:emojis})}).catch(()=>{});
}

function cerrarCartel() {
  document.getElementById('cartel-overlay').style.display = 'none';
  fetch('/api/cartel/hide',{method:'POST'}).catch(()=>{});
}

// ══════════════════════════════════════════
//  EFECTOS DE FONDO
// ══════════════════════════════════════════
let efectoActual = 'ninguno';
let efectoInterval = null;

function aplicarEfecto(nombre) {
  efectoActual = nombre;
  const overlay = document.getElementById('efectos-overlay');
  if (overlay) { overlay.innerHTML = ''; }
  if (efectoInterval) { clearInterval(efectoInterval); efectoInterval = null; }
  // Highlight button activo
  document.querySelectorAll('[id^="efecto-btn-"]').forEach(b => {
    b.style.borderColor = '#2a2a2a'; b.style.color = '#555';
  });
  const activeBtn = document.getElementById('efecto-btn-' + nombre);
  if (activeBtn) { activeBtn.style.borderColor = '#c9a227'; activeBtn.style.color = '#c9a227'; }
  try { localStorage.setItem('rankingVIP_efecto', nombre); } catch(e){}
  if (nombre === 'ninguno') return;
  if (nombre === 'burbujas') iniciarEfectoBurbujas();
  if (nombre === 'estrellas') iniciarEfectoEstrellas();
  if (nombre === 'lluvia_dorada') iniciarEfectoLluviaDorada();
  if (nombre === 'confetti') iniciarEfectoConfetti();
  showToast('Efecto ' + nombre.replace('_',' ').toUpperCase() + ' activado');
}

function iniciarEfectoBurbujas() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-bubble')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-bubble';
    s.textContent = `@keyframes efectoBubble{0%{opacity:0.7;transform:translateY(0) scale(1)}50%{opacity:0.4}100%{opacity:0;transform:translateY(-100vh) scale(0.3)}}`;
    document.head.appendChild(s);
  }
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const sz = 5 + Math.random() * 20;
    const isGold = Math.random() > 0.6;
    el.style.cssText = `position:absolute;bottom:-30px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.4:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:efectoBubble ${5+Math.random()*8}s ease-in forwards;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 14000);
  }
  for(let i=0;i<10;i++) setTimeout(crear, i*200);
  efectoInterval = setInterval(crear, 350);
}

function iniciarEfectoEstrellas() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-star')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-star';
    s.textContent = `@keyframes starFloat{0%{opacity:0;transform:translateY(0) rotate(0deg) scale(0)}20%{opacity:0.8}80%{opacity:0.5}100%{opacity:0;transform:translateY(-80vh) rotate(360deg) scale(0.5)}}`;
    document.head.appendChild(s);
  }
  const syms = ['★','✦','✧','✶','✸','✺'];
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const sym = syms[Math.floor(Math.random()*syms.length)];
    const sz = 10 + Math.random()*18;
    const colors = ['#c9a227','#e8c84a','#ffffff','#f0ece0'];
    const color = colors[Math.floor(Math.random()*colors.length)];
    el.textContent = sym;
    el.style.cssText = `position:absolute;bottom:-30px;left:${Math.random()*100}%;font-size:${sz}px;color:${color};opacity:0;animation:starFloat ${6+Math.random()*6}s ease-in-out forwards;pointer-events:none;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 13000);
  }
  for(let i=0;i<8;i++) setTimeout(crear, i*300);
  efectoInterval = setInterval(crear, 500);
}

function iniciarEfectoLluviaDorada() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-gold')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-gold';
    s.textContent = `@keyframes goldFall{0%{opacity:0.9;transform:translateY(-10px) rotate(0deg)}100%{opacity:0;transform:translateY(100vh) rotate(180deg)}}`;
    document.head.appendChild(s);
  }
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const w = 2 + Math.random()*3, h = 12 + Math.random()*20;
    const gold = Math.random() > 0.3 ? '#c9a227' : '#e8c84a';
    el.style.cssText = `position:absolute;top:-30px;left:${Math.random()*100}%;width:${w}px;height:${h}px;background:${gold};border-radius:1px;opacity:0;animation:goldFall ${3+Math.random()*4}s linear forwards;pointer-events:none;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 8000);
  }
  for(let i=0;i<15;i++) setTimeout(crear, i*100);
  efectoInterval = setInterval(crear, 120);
}

function iniciarEfectoConfetti() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-confetti')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-confetti';
    s.textContent = `@keyframes confettiFall{0%{opacity:1;transform:translateY(-10px) rotate(0deg)}100%{opacity:0.3;transform:translateY(100vh) rotate(720deg)}}`;
    document.head.appendChild(s);
  }
  const colors = ['#c9a227','#e8c84a','#ffffff','#ff4444','#44aaff','#44ff88','#ff44ff'];
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const color = colors[Math.floor(Math.random()*colors.length)];
    const w = 6 + Math.random()*10, h = 4 + Math.random()*8;
    el.style.cssText = `position:absolute;top:-20px;left:${Math.random()*100}%;width:${w}px;height:${h}px;background:${color};border-radius:${Math.random()>0.5?'50%':'2px'};animation:confettiFall ${4+Math.random()*5}s ease-in forwards;pointer-events:none;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 10000);
  }
  for(let i=0;i<20;i++) setTimeout(crear, i*80);
  efectoInterval = setInterval(crear, 200);
}

function cargarEfectoGuardado() {
  try {
    const e = localStorage.getItem('rankingVIP_efecto');
    if (e && e !== 'ninguno') { setTimeout(()=>aplicarEfecto(e), 500); }
  } catch(e) {}
}

// ══════════════════════════════════════════
//  CARTEL – Override cargarTemaGuardado to support new themes
function cargarTemaGuardado() {
  try {
    const t = localStorage.getItem('rankingVIP_tema');
    if (t) aplicarTema(t);
  } catch(e) {}
}

// ══════════════════════════════════════════
//  AUTH — Login y cambio de contraseña
// ══════════════════════════════════════════
let pinValue = '';
function pinUpdateDots() {
  for(let i=0;i<4;i++) {
    const d = document.getElementById('pin-d'+i);
    if(d) d.classList.toggle('filled', i < pinValue.length);
  }
}
function pinPress(digit) {
  const err = document.getElementById('login-error');
  if(err) err.textContent='';
  if(pinValue.length >= 4) return;
  pinValue += digit;
  pinUpdateDots();
  if(pinValue.length === 4) setTimeout(checkLogin, 120);
}
function pinBack() {
  if(pinValue.length > 0) { pinValue = pinValue.slice(0,-1); pinUpdateDots(); }
  const err = document.getElementById('login-error');
  if(err) err.textContent='';
}
document.addEventListener('keydown', e=>{
  if(document.getElementById('login-modal').style.display==='none') return;
  if(e.key>='0'&&e.key<='9') pinPress(e.key);
  else if(e.key==='Backspace') pinBack();
  else if(e.key==='Enter') checkLogin();
});
async function checkLogin() {
  const err = document.getElementById('login-error');
  if(pinValue.length < 4) { if(err) err.textContent='Ingresá los 4 dígitos'; return; }
  try {
    const r = await fetch('/api/auth',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({password:pinValue})});
    const d = await r.json();
    if (d.ok) {
      sessionStorage.setItem('jagger_auth','ok');
      document.getElementById('login-modal').style.display='none';
    } else {
      if(err) err.textContent='PIN incorrecto';
      pinValue=''; pinUpdateDots();
    }
  } catch(e) { if(err) err.textContent='Error de conexión'; pinValue=''; pinUpdateDots(); }
}
if (sessionStorage.getItem('jagger_auth')==='ok') {
  document.getElementById('login-modal').style.display='none';
}

async function cambiarPassword() {
  const actual = document.getElementById('pwd-actual').value;
  const nueva  = document.getElementById('pwd-nueva').value;
  const conf   = document.getElementById('pwd-confirm').value;
  const msg    = document.getElementById('pwd-msg');
  if (!actual||!nueva) { msg.style.color='#a83030'; msg.textContent='Completá todos los campos'; return; }
  if (!/^\d{4}$/.test(nueva)) { msg.style.color='#a83030'; msg.textContent='El nuevo PIN debe tener exactamente 4 dígitos'; return; }
  if (nueva !== conf)  { msg.style.color='#a83030'; msg.textContent='Los PINs no coinciden'; return; }
  try {
    const r = await fetch('/api/auth/change',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({current:actual,new:nueva})});
    const d = await r.json();
    if (d.ok) {
      msg.style.color='#2ecc71'; msg.textContent='Contraseña actualizada correctamente';
      document.getElementById('pwd-actual').value='';
      document.getElementById('pwd-nueva').value='';
      document.getElementById('pwd-confirm').value='';
    } else {
      msg.style.color='#a83030'; msg.textContent=d.error||'Error al cambiar contraseña';
    }
  } catch(e) { msg.style.color='#a83030'; msg.textContent='Error de conexión'; }
}

// Init
buildColorGrid();
cargarPersonalizacionGuardada();
cargarTemaGuardado();
cargarEfectoGuardado();
for(let c=1;c<=3;c++) renderCajaInner(c);
cargarConfTarjetas().then(()=>{loadData();setInterval(loadData,2000);setInterval(sincronizarConfTarjetas,120000);});
</script>
</body>
</html>
"""

HISTORIAL_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Historial — Jagger VIP</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Oswald:wght@400;600;700&family=Rajdhani:wght@400;600&display=swap');
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#080808;color:#f0ece0;font-family:'Rajdhani',sans-serif;min-height:100vh;}
.top-bar{background:#0a0a0a;border-bottom:1px solid #1a1a1a;padding:14px 20px;display:flex;align-items:center;justify-content:space-between;}
.logo{font-family:'Oswald',sans-serif;font-size:22px;color:#c9a227;letter-spacing:3px;}
.back-btn{color:#555;text-decoration:none;font-size:13px;letter-spacing:2px;border:1px solid #222;border-radius:6px;padding:7px 14px;transition:all .15s;}
.back-btn:hover{color:#c9a227;border-color:#c9a227;}
.nav{display:flex;gap:0;background:#0a0a0a;border-bottom:1px solid #1a1a1a;overflow-x:auto;}
.nav-btn{padding:12px 22px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#555;background:none;border:none;border-bottom:2px solid transparent;transition:all .2s;white-space:nowrap;}
.nav-btn.active{color:#c9a227;border-bottom-color:#c9a227;}
.page{display:none;padding:20px;max-width:1100px;margin:0 auto;}
.page.active{display:block;}
.kpi-row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px;}
.kpi{flex:1;min-width:140px;background:#111;border:1px solid #1a1a1a;border-radius:10px;padding:16px 18px;}
.kpi-lbl{font-size:9px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:6px;}
.kpi-val{font-family:'Oswald',sans-serif;font-size:26px;font-weight:700;color:#c9a227;}
.section-title{font-family:'Oswald',sans-serif;font-size:15px;color:#c9a227;letter-spacing:2px;margin:24px 0 12px;text-transform:uppercase;}
table{width:100%;border-collapse:collapse;font-size:14px;}
th{background:#0d0d0d;color:#aaa;font-size:10px;letter-spacing:2px;text-transform:uppercase;padding:10px 12px;text-align:left;border-bottom:1px solid #222;}
td{padding:10px 12px;border-bottom:1px solid #181818;color:#e0ddd0;}
tr:hover td{background:#111;}
.gold{color:#c9a227;font-weight:700;}
.rank-badge{display:inline-block;width:24px;height:24px;border-radius:50%;text-align:center;line-height:24px;font-size:11px;font-weight:700;}
.r1{background:#c9a227;color:#000;}
.r2{background:#888;color:#000;}
.r3{background:#6a3a00;color:#e8c84a;}
.noche-row{cursor:pointer;}
.noche-row:hover td{background:#111;}
.detail-panel{display:none;background:#0d0d0d;border:1px solid #222;border-radius:8px;padding:16px;margin:4px 0 16px;}
.detail-panel.open{display:block;}
.trim-tabs{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;}
.trim-tab{padding:7px 16px;border:1px solid #2a2a2a;border-radius:20px;cursor:pointer;font-size:13px;letter-spacing:1px;color:#aaa;transition:all .15s;}
.trim-tab.active{border-color:#c9a227;color:#c9a227;background:#0d0b00;}
.empty{color:#444;text-align:center;padding:60px 20px;font-size:14px;letter-spacing:2px;}
.dl-btn{display:inline-block;background:transparent;color:#c9a227;border:1px solid #c9a227;border-radius:7px;padding:8px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;text-decoration:none;transition:all .15s;}
.dl-btn:hover{background:#c9a227;color:#000;}
.charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:16px 0 24px;}
@media(max-width:640px){.charts-grid{grid-template-columns:1fr;}}
.chart-card{background:#0d0d0d;border:1px solid #1a1a1a;border-radius:10px;padding:16px;}
.chart-card-title{font-size:10px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:12px;}
.chart-card.wide{grid-column:1/-1;}
.hora-row{display:flex;align-items:center;gap:10px;padding:5px 0;border-bottom:1px solid #181818;}
.hora-lbl{width:36px;font-size:12px;color:#aaa;flex-shrink:0;}
.hora-bar-wrap{flex:1;background:#111;border-radius:3px;height:8px;overflow:hidden;}
.hora-bar{height:8px;background:#c9a227;border-radius:3px;transition:width .3s;}
.hora-val{width:80px;text-align:right;font-size:12px;color:#e0ddd0;flex-shrink:0;}
</style>
</head>
<body>
<div class="top-bar">
  <span class="logo">RANKING VIP — HISTORIAL</span>
  <div style="display:flex;gap:10px;align-items:center;">
    <a class="dl-btn" href="/api/export/excel" target="_blank">⬇ Excel</a>
    <button class="dl-btn" id="gs-btn" onclick="syncGSheets()" style="cursor:pointer;border:none;">☁ Google Sheets</button>
    <a class="back-btn" href="/">← Volver</a>
  </div>
</div>
<div class="nav">
  <button class="nav-btn active" onclick="showPage('general',this)">General</button>
  <button class="nav-btn" onclick="showPage('noches',this)">Noches</button>
  <button class="nav-btn" onclick="showPage('ranking',this)">Ranking</button>
</div>

<!-- GENERAL -->
<div id="page-general" class="page active">
  <div id="kpis" class="kpi-row"></div>
  <div class="charts-grid">
    <div class="chart-card wide">
      <div class="chart-card-title">Evolución noche a noche</div>
      <canvas id="chart-trend"></canvas>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">Facturación por mes</div>
      <canvas id="chart-meses"></canvas>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">Distribución por caja (acumulado)</div>
      <canvas id="chart-cajas-total"></canvas>
    </div>
  </div>
  <div class="section-title">Totales por mes</div>
  <table id="tbl-meses"><thead><tr><th>Mes</th><th>Noches</th><th>Total</th><th>Ops</th></tr></thead><tbody></tbody></table>
</div>

<!-- NOCHES -->
<div id="page-noches" class="page">
  <div class="section-title">Todas las noches</div>
  <table><thead><tr><th>Fecha</th><th>Cierre</th><th>Total</th><th>Ops</th><th>1°</th><th>2°</th><th>3°</th></tr></thead>
  <tbody id="tbl-noches"></tbody></table>
</div>

<!-- RANKING -->
<div id="page-ranking" class="page">
  <div class="section-title">Período</div>
  <div class="trim-tabs" id="trim-tabs"></div>
  <table><thead><tr><th>#</th><th>Nombre</th><th>Noches</th><th>Total</th><th>Promedio / noche</th></tr></thead>
  <tbody id="tbl-ranking"></tbody></table>
</div>

<script>
let historial = [];
let periodoActivo = 'all';
let chartTrend=null, chartMeses=null, chartCajas=null;

const GOLD = '#c9a227', GOLD2 = '#e8c84a', DARK = '#0d0d0d';
const CAJA_COLORS = ['#c9a227','#3a9a5a','#3a6ac9'];
Chart.defaults.color = '#999';
Chart.defaults.borderColor = '#222';

async function init() {
  const r = await fetch('/api/historial');
  historial = await r.json();
  renderGeneral();
  renderNoches();
  renderTrimTabs();
  renderRanking('all');
}

function fmt(n) { return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0}); }
function esc(s){ return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

function showPage(id, btn) {
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('page-'+id).classList.add('active');
  if(btn) btn.classList.add('active');
}

function renderGeneral() {
  const total = historial.reduce((s,n)=>s+n.total,0);
  const ops   = historial.reduce((s,n)=>s+n.operaciones,0);
  const rk = {}; historial.forEach(n=>(n.ranking||[]).forEach(r=>{ rk[r.name]=(rk[r.name]||0)+r.total; }));
  const best = Object.entries(rk).sort((a,b)=>b[1]-a[1])[0];
  const mejorNoche = historial.length ? historial.reduce((mx,n)=>n.total>mx.total?n:mx) : null;
  document.getElementById('kpis').innerHTML = [
    ['Total acumulado', fmt(total)],
    ['Noches', historial.length],
    ['Operaciones', ops],
    ['Cliente top', best ? best[0] : '—'],
    ['Mejor noche', mejorNoche ? fmt(mejorNoche.total) : '—'],
  ].map(([l,v])=>`<div class="kpi"><div class="kpi-lbl">${l}</div><div class="kpi-val" style="font-size:${String(v).length>10?'16px':'26px'}">${v}</div></div>`).join('');

  // Tabla mensual
  const por_mes = {};
  historial.forEach(n=>{ const m=n.fecha.slice(0,7); if(!por_mes[m]) por_mes[m]={total:0,ops:0,noches:0}; por_mes[m].total+=n.total; por_mes[m].ops+=n.operaciones; por_mes[m].noches++; });
  document.querySelector('#tbl-meses tbody').innerHTML = Object.entries(por_mes).sort((a,b)=>b[0].localeCompare(a[0]))
    .map(([m,v])=>`<tr><td class="gold">${m}</td><td>${v.noches}</td><td class="gold">${fmt(v.total)}</td><td>${v.ops}</td></tr>`).join('') || '<tr><td colspan="4" class="empty">Sin datos</td></tr>';

  // Gráfico tendencia noche a noche (line)
  const nochesSorted = [...historial].sort((a,b)=>a.fecha.localeCompare(b.fecha));
  const trendLabels = nochesSorted.map(n=>n.fecha.slice(5));
  const trendData   = nochesSorted.map(n=>n.total);
  if(chartTrend) chartTrend.destroy();
  const fmtTick = v => '$'+Number(v).toLocaleString('es-AR');
  const tooltipFmt = { callbacks:{ label: ctx => ' '+fmt(ctx.raw) } };
  const xStyle = { color:'#bbb', font:{size:11} };
  const yStyle = { color:'#bbb', font:{size:11}, callback: fmtTick };
  const axisTitle = txt => ({ display:true, text:txt, color:'#aaa', font:{size:11} });

  chartTrend = new Chart(document.getElementById('chart-trend'),{
    type:'line',
    data:{labels:trendLabels,datasets:[{label:'Total por noche',data:trendData,borderColor:GOLD,backgroundColor:'rgba(201,162,39,0.10)',tension:0.35,pointBackgroundColor:GOLD,pointRadius:5,pointHoverRadius:7,fill:true}]},
    options:{
      plugins:{
        legend:{display:true,labels:{color:'#e0ddd0',font:{size:12},boxWidth:14}},
        title:{display:true,text:'Evolución del total por noche',color:'#e0ddd0',font:{size:13},padding:{bottom:10}},
        tooltip:{callbacks:{label: ctx=>' '+fmt(ctx.raw)}}
      },
      scales:{
        x:{title:axisTitle('Fecha'),ticks:xStyle},
        y:{title:axisTitle('Total recaudado ($)'),ticks:yStyle}
      },
      maintainAspectRatio:true,aspectRatio:3.2
    }
  });

  // Gráfico facturación por mes (bar)
  const mesesSorted = Object.entries(por_mes).sort((a,b)=>a[0].localeCompare(b[0]));
  if(chartMeses) chartMeses.destroy();
  chartMeses = new Chart(document.getElementById('chart-meses'),{
    type:'bar',
    data:{labels:mesesSorted.map(([m])=>m),datasets:[{label:'Total mensual',data:mesesSorted.map(([,v])=>v.total),backgroundColor:GOLD+'aa',borderColor:GOLD,borderWidth:1,borderRadius:4}]},
    options:{
      plugins:{
        legend:{display:true,labels:{color:'#e0ddd0',font:{size:12},boxWidth:14}},
        title:{display:true,text:'Facturación total por mes',color:'#e0ddd0',font:{size:13},padding:{bottom:10}},
        tooltip:tooltipFmt
      },
      scales:{
        x:{title:axisTitle('Mes'),ticks:xStyle},
        y:{title:axisTitle('Total ($)'),ticks:yStyle}
      },
      maintainAspectRatio:true,aspectRatio:2
    }
  });

  // Gráfico cajas total (doughnut)
  const cajasTotals = [1,2,3].map(c=>historial.reduce((s,n)=>s+(n.por_caja?.[c]||n.por_caja?.[String(c)]||0),0));
  if(chartCajas) chartCajas.destroy();
  chartCajas = new Chart(document.getElementById('chart-cajas-total'),{
    type:'doughnut',
    data:{labels:['Abajo','Extendido','VIP'],datasets:[{data:cajasTotals,backgroundColor:CAJA_COLORS,borderColor:'#080808',borderWidth:3}]},
    options:{
      plugins:{
        legend:{position:'bottom',labels:{color:'#e0ddd0',font:{size:13},padding:16,boxWidth:14}},
        title:{display:true,text:'Distribución por caja (acumulado)',color:'#e0ddd0',font:{size:13},padding:{bottom:10}},
        tooltip:{callbacks:{label: ctx=>' '+ctx.label+': '+fmt(ctx.raw)+' ('+Math.round(ctx.parsed/cajasTotals.reduce((a,b)=>a+b,0)*100)+'%)'}}
      },
      maintainAspectRatio:true,aspectRatio:1.5
    }
  });
}

function horasHTML(txs) {
  if (!txs || !txs.length) return '';
  const byHour = {};
  txs.forEach(t => {
    const h = (t.time||'00:00').split(':')[0].padStart(2,'0');
    byHour[h] = (byHour[h]||0) + t.amount;
  });
  const horas = Object.entries(byHour).sort((a,b)=>a[0].localeCompare(b[0]));
  if (!horas.length) return '';
  const maxVal = Math.max(...horas.map(([,v])=>v));
  return `<div style="margin-top:12px;">
    <div style="font-size:10px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:8px;">Consumo por hora</div>
    ${horas.map(([h,v])=>`<div class="hora-row">
      <span class="hora-lbl">${h}h</span>
      <div class="hora-bar-wrap"><div class="hora-bar" style="width:${Math.round(v/maxVal*100)}%"></div></div>
      <span class="hora-val">${fmt(v)}</span>
    </div>`).join('')}
  </div>`;
}

function renderNoches() {
  document.getElementById('tbl-noches').innerHTML = [...historial].reverse().map((n,idx)=>{
    const rnk=n.ranking||[];
    const id='dp-'+idx;
    return `<tr class="noche-row" onclick="toggleDetail('${id}')">
      <td class="gold">${n.fecha}</td><td>${n.hora_cierre||'—'}</td>
      <td class="gold">${fmt(n.total)}</td><td>${n.operaciones}</td>
      <td>${esc(rnk[0]?.name||'—')}</td><td>${esc(rnk[1]?.name||'—')}</td><td>${esc(rnk[2]?.name||'—')}</td>
    </tr>
    <tr><td colspan="7" style="padding:0;border:none">
      <div class="detail-panel" id="${id}">
        <div style="font-size:11px;color:#aaa;letter-spacing:2px;margin-bottom:10px;">
          CAJA 1: <span class="gold">${fmt(n.por_caja?.[1]||n.por_caja?.['1']||0)}</span> &nbsp;·&nbsp;
          CAJA 2: <span class="gold">${fmt(n.por_caja?.[2]||n.por_caja?.['2']||0)}</span> &nbsp;·&nbsp;
          CAJA 3: <span class="gold">${fmt(n.por_caja?.[3]||n.por_caja?.['3']||0)}</span>
        </div>
        <div style="font-size:10px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:6px;">Ranking de la noche</div>
        ${rnk.slice(0,5).map((r,i)=>`<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #181818;font-size:14px;">
          <span><span class="rank-badge ${i<3?'r'+(i+1):''}" style="${i>=3?'background:#1a1a1a;color:#aaa;width:24px;height:24px;border-radius:50%;display:inline-block;text-align:center;line-height:24px;font-size:11px;':''}">${i+1}</span> &nbsp;${esc(r.name)}</span>
          <span class="gold">${fmt(r.total)}</span>
        </div>`).join('')}
        ${horasHTML(n.transactions)}
      </div>
    </td></tr>`;
  }).join('') || `<tr><td colspan="7" class="empty">Sin noches registradas</td></tr>`;
}

function toggleDetail(id) { document.getElementById(id).classList.toggle('open'); }

function renderTrimTabs() {
  const ahora = new Date();
  const tabs = [
    {id:'all', label:'Todo'},
    {id:'year', label:String(ahora.getFullYear())},
    {id:'q'+Math.ceil((ahora.getMonth()+1)/3), label:'Trimestre '+Math.ceil((ahora.getMonth()+1)/3)},
    {id:'month', label:ahora.toLocaleString('es',{month:'long'})},
  ];
  document.getElementById('trim-tabs').innerHTML = tabs.map(t=>
    `<div class="trim-tab${t.id===periodoActivo?' active':''}" onclick="selectPeriodo('${t.id}',this)">${t.label}</div>`
  ).join('');
}

function selectPeriodo(id, el) {
  periodoActivo = id;
  document.querySelectorAll('.trim-tab').forEach(x=>x.classList.remove('active'));
  el.classList.add('active');
  renderRanking(id);
}

function filtrarNoches(periodo) {
  const ahora = new Date();
  const año = ahora.getFullYear(), mes = ahora.getMonth()+1;
  return historial.filter(n=>{
    if (periodo==='all') return true;
    const [y,m] = n.fecha.split('-').map(Number);
    if (periodo==='year') return y===año;
    if (periodo==='month') return y===año && m===mes;
    if (periodo.startsWith('q')) { const q=parseInt(periodo[1]); return y===año && Math.ceil(m/3)===q; }
    return true;
  });
}

function renderRanking(periodo) {
  const noches = filtrarNoches(periodo);
  const rk = {};
  noches.forEach(n=>(n.ranking||[]).forEach(r=>{ if(!rk[r.name]) rk[r.name]={total:0,noches:0}; rk[r.name].total+=r.total; rk[r.name].noches++; }));
  const sorted = Object.entries(rk).sort((a,b)=>b[1].total-a[1].total);
  document.getElementById('tbl-ranking').innerHTML = sorted.map(([nm,v],i)=>{
    const badge = i<3 ? `<span class="rank-badge r${i+1}">${i+1}</span>` : `<span style="color:#555;padding:0 6px;">${i+1}</span>`;
    return `<tr><td>${badge}</td><td>${esc(nm)}</td><td>${v.noches}</td><td class="gold">${fmt(v.total)}</td><td>${fmt(Math.round(v.total/v.noches))}</td></tr>`;
  }).join('') || `<tr><td colspan="5" class="empty">Sin datos para este período</td></tr>`;
}

init();

async function syncGSheets() {
  const btn = document.getElementById('gs-btn');
  const orig = btn.textContent;
  btn.textContent = '⏳ Sincronizando...';
  btn.disabled = true;
  try {
    const r = await fetch('/api/export/gsheets', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({modo:'all'})});
    const d = await r.json();
    if (d.ok) {
      btn.textContent = '✓ Listo';
      setTimeout(()=>{ btn.textContent = orig; btn.disabled = false; }, 2000);
      if (confirm('¡Sincronizado! ¿Abrir la planilla en Google Sheets?')) window.open(d.url, '_blank');
    } else {
      btn.textContent = orig;
      btn.disabled = false;
      const setup = d.error && (d.error.includes('creds') || d.error.includes('GSHEETS')) ?
        '\n\nPara configurar:\n1. Creá una cuenta de servicio en Google Cloud Console\n2. Descargá el JSON y guardalo como gsheets_creds.json junto al app.py\n3. Creá una planilla en Google Sheets y copiá el ID de la URL\n4. Seteá la variable de entorno GSHEETS_ID con ese ID\n5. Compartí la planilla con el email de la cuenta de servicio' : '';
      alert('Error: ' + d.error + setup);
    }
  } catch(e) {
    btn.textContent = orig;
    btn.disabled = false;
    alert('Error de conexión: ' + e.message);
  }
}
</script>
</body>
</html>"""

@app.route('/historial')
def historial_page():
    return render_template_string(HISTORIAL_HTML)

@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/api/tx', methods=['GET'])
def get_tx():
    with lock:
        data = load_data()
    return jsonify(data['transactions'])

@app.route('/api/tx', methods=['POST'])
def add_tx():
    try:
        with lock:
            data = load_data()
            body = request.get_json()
            if not body or 'name' not in body or 'amount' not in body or 'caja' not in body:
                return jsonify({'ok': False, 'error': 'Faltan campos requeridos'}), 400
            amount = float(body['amount'])
            codigo = str(body.get('tarjeta_codigo',''))
            if codigo:
                if 'tarjetas' not in data: data['tarjetas'] = {}
                if codigo not in data['tarjetas']:
                    conf_list = data.get('tarjetas_conf', [])
                    conf = next((t for t in conf_list if t.get('codigo')==codigo), None)
                    try:
                        saldo_ini = float(conf['saldo_inicial']) if conf and conf.get('saldo_inicial') not in ('', None) else 0
                    except (ValueError, TypeError):
                        saldo_ini = 0
                    data['tarjetas'][codigo] = {'saldo_actual': saldo_ini, 'nombre': str(body['name'])}
                saldo_disponible = data['tarjetas'][codigo]['saldo_actual']
                if saldo_disponible < amount:
                    return jsonify({'ok': False, 'error': 'Saldo insuficiente. Disponible: $' + str(int(saldo_disponible))}), 400
            data['tx_id_counter'] += 1
            tx = {
                'id': data['tx_id_counter'],
                'name': str(body['name']),
                'amount': amount,
                'caja': int(body['caja']),
                'mesa': str(body.get('mesa','')),
                'tarjeta_codigo': codigo,
                'time': str(body.get('client_time','')) or datetime.now().strftime('%H:%M')
            }
            data['transactions'].append(tx)
            if codigo:
                data['tarjetas'][codigo]['saldo_actual'] -= amount
                data['tarjetas'][codigo]['nombre'] = tx['name']
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tx/<int:tid>', methods=['PUT'])
def edit_tx(tid):
    try:
        with lock:
            data = load_data()
            tx = next((t for t in data['transactions'] if t['id']==tid), None)
            if not tx:
                return jsonify({'ok': False, 'error': 'Transacción no encontrada'}), 404
            body = request.get_json() or {}
            new_amount = float(body.get('amount', tx['amount']))
            diff = new_amount - tx['amount']
            codigo = tx.get('tarjeta_codigo', '')
            if codigo and diff != 0 and 'tarjetas' in data and codigo in data['tarjetas']:
                nuevo_saldo = data['tarjetas'][codigo]['saldo_actual'] - diff
                if nuevo_saldo < 0:
                    return jsonify({'ok': False, 'error': 'Saldo insuficiente para este ajuste'}), 400
                data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo
            tx['amount'] = new_amount
            if 'name' in body and str(body['name']).strip():
                tx['name'] = str(body['name']).strip()
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tx/<int:tid>', methods=['DELETE'])
def delete_tx(tid):
    with lock:
        data = load_data()
        tx = next((t for t in data['transactions'] if t['id']==tid), None)
        if tx:
            codigo = tx.get('tarjeta_codigo','')
            if codigo and 'tarjetas' in data and codigo in data['tarjetas']:
                data['tarjetas'][codigo]['saldo_actual'] += tx['amount']
            data['transactions'] = [t for t in data['transactions'] if t['id']!=tid]
            save_data(data)
    return jsonify({'ok': True})

@app.route('/api/tarjetas', methods=['GET'])
def get_tarjetas():
    with lock:
        data = load_data()
    return jsonify(data.get('tarjetas', {}))

@app.route('/api/tarjetas/config', methods=['GET'])
def get_tarjetas_conf():
    with lock:
        data = load_data()
    return jsonify(data.get('tarjetas_conf', []))

@app.route('/api/tarjetas/config', methods=['POST'])
def set_tarjetas_conf():
    try:
        with lock:
            data = load_data()
            conf = request.get_json()
            if not conf:
                return jsonify({'ok': False, 'error': 'Configuracion invalida'}), 400
            codigos_vistos = {}
            for t in conf:
                codigo = t.get('codigo','')
                if codigo:
                    if codigo in codigos_vistos:
                        slot_anterior = codigos_vistos[codigo]
                        return jsonify({'ok': False, 'error': f'La tarjeta ya esta asignada a Mesa {slot_anterior}'}), 400
                    codigos_vistos[codigo] = t.get('slot', '?')
            conf_vieja = {t.get('codigo',''): t for t in data.get('tarjetas_conf', []) if t.get('codigo','')}
            data['tarjetas_conf'] = conf
            if 'tarjetas' not in data: data['tarjetas'] = {}
            for t in conf:
                codigo = t.get('codigo','')
                if codigo and t.get('saldo_inicial'):
                    nuevo_saldo_ini = float(t['saldo_inicial'])
                    if codigo not in data['tarjetas']:
                        data['tarjetas'][codigo] = {'saldo_actual': nuevo_saldo_ini, 'nombre':''}
                    else:
                        old_conf = conf_vieja.get(codigo)
                        old_ini = float(old_conf['saldo_inicial']) if old_conf and old_conf.get('saldo_inicial') else None
                        if old_ini is not None and old_ini != nuevo_saldo_ini:
                            gastado = old_ini - data['tarjetas'][codigo]['saldo_actual']
                            nuevo_saldo_act = max(0, nuevo_saldo_ini - gastado)
                            data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo_act
                        elif old_ini is None:
                            data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo_ini
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tarjetas/limpiar', methods=['POST'])
def limpiar_tarjetas():
    with lock:
        data = load_data()
        data['tarjetas_conf'] = []
        data['tarjetas'] = {}
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/reset', methods=['POST'])
def reset():
    with lock:
        data = load_data()
        data['transactions'] = []
        data['tx_id_counter'] = 0
        conf_list = data.get('tarjetas_conf', [])
        data['tarjetas'] = {}
        for t in conf_list:
            codigo = t.get('codigo','')
            if codigo and t.get('saldo_inicial'):
                data['tarjetas'][codigo] = {'saldo_actual': float(t['saldo_inicial']), 'nombre':''}
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/state', methods=['GET'])
def get_state():
    return jsonify(_state)

@app.route('/api/state', methods=['POST'])
def set_state():
    import time
    body = request.get_json() or {}
    with lock:
        if 'hora_fin' in body: _state['hora_fin'] = str(body['hora_fin'])
        if 'premio' in body: _state['premio'] = str(body['premio'])
    return jsonify({'ok': True})

@app.route('/api/winner/show', methods=['POST'])
def winner_show():
    import time
    with lock:
        _state['winner_show'] = True
        _state['winner_ts'] = int(time.time() * 1000)
    return jsonify({'ok': True})

@app.route('/api/winner/hide', methods=['POST'])
def winner_hide():
    with lock:
        _state['winner_show'] = False
    return jsonify({'ok': True})

@app.route('/api/cartel/show', methods=['POST'])
def cartel_show():
    import time
    body = request.get_json() or {}
    with lock:
        _state['cartel_show'] = True
        _state['cartel_ts'] = int(time.time() * 1000)
        _state['cartel_data'] = {
            'nombre': str(body.get('nombre','')),
            'mesa': str(body.get('mesa','')),
            'frase': str(body.get('frase','')),
            'emoji': str(body.get('emoji','🍾')),
        }
    return jsonify({'ok': True})

@app.route('/api/cartel/hide', methods=['POST'])
def cartel_hide():
    with lock:
        _state['cartel_show'] = False
    return jsonify({'ok': True})

@app.route('/api/cerrar_noche', methods=['POST'])
def cerrar_noche():
    with lock:
        data = load_data()
        txs = data['transactions']
        if not txs:
            return jsonify({'ok': False, 'error': 'No hay operaciones esta noche'}), 400
        totals, mesas, por_caja = {}, {}, {1:0, 2:0, 3:0}
        for t in txs:
            totals[t['name']] = totals.get(t['name'], 0) + t['amount']
            if t.get('mesa') and t['name'] not in mesas:
                mesas[t['name']] = t['mesa']
            c = int(t.get('caja', 1))
            if 1 <= c <= 3:
                por_caja[c] += t['amount']
        ranking = sorted(
            [{'name': n, 'total': v, 'mesa': mesas.get(n, '')} for n, v in totals.items()],
            key=lambda x: -x['total']
        )
        noche = {
            'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'hora_cierre': datetime.now().strftime('%H:%M'),
            'hora_fin': _state.get('hora_fin', '05:30'),
            'total': sum(t['amount'] for t in txs),
            'operaciones': len(txs),
            'por_caja': por_caja,
            'ranking': ranking,
            'transactions': list(txs),
        }
        historial = load_historial()
        historial.append(noche)
        save_historial(historial)
    return jsonify({'ok': True, 'noche_id': noche['id']})

@app.route('/api/historial', methods=['GET'])
def get_historial():
    return jsonify(load_historial())

@app.route('/api/export/excel')
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.utils import get_column_letter
        from io import BytesIO
    except ImportError:
        return jsonify({'error': 'Instalá openpyxl: pip install openpyxl'}), 500

    historial = load_historial()
    if not historial:
        return jsonify({'error': 'No hay noches registradas aún'}), 400

    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────────
    GOLD  = PatternFill(start_color='C9A227', fill_type='solid')
    DARK  = PatternFill(start_color='0D0D0D', fill_type='solid')
    DARK2 = PatternFill(start_color='1A1A1A', fill_type='solid')
    f_hdr  = Font(bold=True, color='000000', name='Calibri', size=11)
    f_ti   = Font(bold=True, color='C9A227', name='Calibri', size=14)
    f_gold = Font(bold=True, color='C9A227', name='Calibri', size=11)
    f_gs   = Font(bold=True, color='C9A227', name='Calibri', size=10)
    f_dim  = Font(color='CCCCCC', name='Calibri', size=10)
    f_gray = Font(color='666666', name='Calibri', size=10)
    f_kpi  = Font(bold=True, color='888888', name='Calibri', size=9)
    thin   = Side(style='thin', color='2A2A2A')
    BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR    = Alignment(horizontal='center', vertical='center')

    def set_hdr(ws, row, col_start, labels):
        for ci, h in enumerate(labels, col_start):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = f_hdr; c.fill = GOLD; c.alignment = CTR; c.border = BDR

    def set_row(ws, row, col_start, ncols, alt=False):
        fill = DARK2 if alt else DARK
        for ci in range(col_start, col_start + ncols):
            c = ws.cell(row=row, column=ci)
            c.fill = fill; c.font = f_dim; c.border = BDR; c.alignment = CTR

    def nfmt(cell, v):
        cell.value = v; cell.number_format = '#,##0'; return cell

    # ════════════════════════════════════════════════════
    # HOJA 1 — GENERAL
    # ════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = 'General'
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = 'C9A227'

    total_all = sum(n['total'] for n in historial)
    total_ops = sum(n['operaciones'] for n in historial)
    rk_all = {}
    for n in historial:
        for rv in n.get('ranking', []):
            rk_all[rv['name']] = rk_all.get(rv['name'], 0) + rv['total']
    rk_sorted = sorted(rk_all.items(), key=lambda x: -x[1])
    por_mes = {}
    for n in historial:
        mes = n['fecha'][:7]
        por_mes[mes] = por_mes.get(mes, 0) + n['total']
    meses = sorted(por_mes.keys())

    ws1['A1'] = 'JAGGER CLUB — HISTORIAL VIP'
    ws1['A1'].font = Font(bold=True, color='C9A227', name='Calibri', size=16)
    ws1['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  ·  {len(historial)} noches registradas'
    ws1['A2'].font = f_gray
    for ri, (lbl, val) in enumerate([
        ('TOTAL FACTURADO', f'${total_all:,.0f}'.replace(',', '.')),
        ('NOCHES',          len(historial)),
        ('OPERACIONES',     total_ops),
    ], 4):
        ws1.cell(row=ri, column=1, value=lbl).font = f_kpi
        ws1.cell(row=ri, column=2, value=val).font = f_gold

    ws1.cell(row=8, column=1, value='RANKING GENERAL').font = f_ti
    set_hdr(ws1, 9, 1, ['#', 'Nombre', 'Total ($)'])
    for i, (nm, tot) in enumerate(rk_sorted[:20], 1):
        r = 9 + i
        ws1.cell(row=r, column=1, value=i)
        ws1.cell(row=r, column=2, value=nm)
        nfmt(ws1.cell(row=r, column=3), tot)
        set_row(ws1, r, 1, 3, alt=(i % 2 == 0))
        if i == 1:
            for ci in range(1, 4): ws1.cell(row=r, column=ci).font = f_gs

    base = 11 + len(rk_sorted[:20])
    ws1.cell(row=base, column=1, value='TOTAL POR MES').font = f_ti
    set_hdr(ws1, base+1, 1, ['Mes', 'Total ($)'])
    cd_start = base + 2
    for i, mes in enumerate(meses):
        ws1.cell(row=cd_start+i, column=1, value=mes)
        nfmt(ws1.cell(row=cd_start+i, column=2), por_mes[mes])
        set_row(ws1, cd_start+i, 1, 2, alt=(i % 2 == 0))
    cd_end = cd_start + len(meses) - 1

    if cd_end >= cd_start:
        ch_bar = BarChart(); ch_bar.type='col'; ch_bar.title='Facturación por mes'
        ch_bar.style=10; ch_bar.width=22; ch_bar.height=13
        ch_bar.add_data(Reference(ws1, min_col=2, min_row=cd_start, max_row=cd_end))
        ch_bar.set_categories(Reference(ws1, min_col=1, min_row=cd_start, max_row=cd_end))
        ws1.add_chart(ch_bar, 'E8')

    # Evolución noche a noche (tabla + LineChart debajo de los datos mensuales)
    ev_start = cd_end + 3
    ws1.cell(row=ev_start, column=1, value='EVOLUCIÓN NOCHE A NOCHE').font = f_ti
    set_hdr(ws1, ev_start+1, 1, ['Fecha', 'Total ($)'])
    ev_d_start = ev_start + 2
    for i, n in enumerate(historial):
        rr = ev_d_start + i
        ws1.cell(row=rr, column=1, value=n['fecha'])
        nfmt(ws1.cell(row=rr, column=2), n['total'])
        set_row(ws1, rr, 1, 2, alt=(i % 2 == 0))
    ev_d_end = ev_d_start + len(historial) - 1

    if len(historial) >= 2:
        ch_line = LineChart()
        ch_line.title = 'Evolución noche a noche'
        ch_line.style = 10; ch_line.width = 22; ch_line.height = 13
        ch_line.add_data(Reference(ws1, min_col=2, min_row=ev_start+1, max_row=ev_d_end),
                         titles_from_data=True)
        ch_line.set_categories(Reference(ws1, min_col=1, min_row=ev_d_start, max_row=ev_d_end))
        ws1.add_chart(ch_line, 'E' + str(ev_start))

    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 26
    ws1.column_dimensions['C'].width = 16

    # ════════════════════════════════════════════════════
    # HOJA 2 — NOCHES (tabla + autofilter)
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Noches')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = 'C9A227'
    ws2['A1'] = 'HISTORIAL DE NOCHES'; ws2['A1'].font = f_ti

    COLS_N = ['Fecha', 'Cierre', 'Total ($)', 'Ops', '1°', '2°', '3°', 'Abajo', 'Extendido', 'VIP']
    set_hdr(ws2, 3, 1, COLS_N)
    r2 = 4
    for idx, n in enumerate(reversed(historial)):
        rnk = n.get('ranking', []); pc = n.get('por_caja', {})
        ws2.cell(row=r2, column=1, value=n['fecha'])
        ws2.cell(row=r2, column=2, value=n.get('hora_cierre', ''))
        nfmt(ws2.cell(row=r2, column=3), n['total'])
        ws2.cell(row=r2, column=4, value=n['operaciones'])
        for ci, ridx in zip(range(5, 8), range(3)):
            ws2.cell(row=r2, column=ci, value=rnk[ridx]['name'] if len(rnk) > ridx else '')
        for ci, key in zip(range(8, 11), [1, 2, 3]):
            nfmt(ws2.cell(row=r2, column=ci), pc.get(key, pc.get(str(key), 0)))
        set_row(ws2, r2, 1, 10, alt=(idx % 2 == 0))
        r2 += 1

    ws2.auto_filter.ref = f'A3:J{r2-1}'
    for ci, w in zip(range(1, 11), [13, 8, 14, 6, 22, 22, 22, 13, 13, 13]):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    if rk_sorted:
        top5 = rk_sorted[:5]; cc = 12
        ws2.cell(row=3, column=cc, value='TOP 5 (todas las noches)').font = f_gold
        set_hdr(ws2, 4, cc, ['Nombre', 'Total ($)'])
        for i, (nm, tot) in enumerate(top5):
            ws2.cell(row=5+i, column=cc, value=nm)
            nfmt(ws2.cell(row=5+i, column=cc+1), tot)
            set_row(ws2, 5+i, cc, 2, alt=(i % 2 == 0))
        ch2 = BarChart(); ch2.type='bar'; ch2.title='Top 5 Clientes'
        ch2.style=10; ch2.width=18; ch2.height=10
        ch2.add_data(Reference(ws2, min_col=cc+1, min_row=4, max_row=4+len(top5)),
                     titles_from_data=True)
        ch2.set_categories(Reference(ws2, min_col=cc, min_row=5, max_row=4+len(top5)))
        ws2.add_chart(ch2, get_column_letter(cc)+'11')

    # ════════════════════════════════════════════════════
    # HOJA 3 — ÚLTIMA NOCHE (detalle completo)
    # ════════════════════════════════════════════════════
    ul = historial[-1]
    ws3 = wb.create_sheet(f'Noche {ul["fecha"]}')
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = 'C9A227'
    ws3['A1'] = f'DETALLE — {ul["fecha"]}'; ws3['A1'].font = f_ti
    ws3['A2'] = f'Total: ${ul["total"]:,.0f}   ·   {ul["operaciones"]} operaciones'.replace(',', '.')
    ws3['A2'].font = f_gold

    ws3.cell(row=4, column=1, value='RANKING DE LA NOCHE').font = f_ti
    set_hdr(ws3, 5, 1, ['#', 'Nombre', 'Mesa', 'Total ($)'])
    r3 = 6
    for i, rv in enumerate(ul.get('ranking', []), 1):
        ws3.cell(row=r3, column=1, value=i)
        ws3.cell(row=r3, column=2, value=rv['name'])
        ws3.cell(row=r3, column=3, value=rv.get('mesa', ''))
        nfmt(ws3.cell(row=r3, column=4), rv['total'])
        set_row(ws3, r3, 1, 4, alt=(i % 2 == 0))
        if i == 1:
            for ci in range(1, 5): ws3.cell(row=r3, column=ci).font = f_gs
        r3 += 1

    # Por caja — barra + torta
    r3 += 1
    ws3.cell(row=r3, column=1, value='FACTURACIÓN POR CAJA').font = f_ti
    set_hdr(ws3, r3+1, 1, ['Caja', 'Total ($)'])
    pc3 = ul.get('por_caja', {})
    cs = r3 + 2
    for i, cn in enumerate([1, 2, 3]):
        ws3.cell(row=cs+i, column=1, value=f'Caja {cn}')
        nfmt(ws3.cell(row=cs+i, column=2), pc3.get(cn, pc3.get(str(cn), 0)))
        set_row(ws3, cs+i, 1, 2, alt=(i % 2 == 0))

    ch3_bar = BarChart(); ch3_bar.type='col'; ch3_bar.title='Por caja'
    ch3_bar.style=10; ch3_bar.width=14; ch3_bar.height=10
    ch3_bar.add_data(Reference(ws3, min_col=2, min_row=cs, max_row=cs+2))
    ch3_bar.set_categories(Reference(ws3, min_col=1, min_row=cs, max_row=cs+2))
    ws3.add_chart(ch3_bar, 'F4')

    ch3_pie = PieChart()
    ch3_pie.title = 'Distribución por caja'; ch3_pie.style = 10
    ch3_pie.width = 14; ch3_pie.height = 10
    ch3_pie.add_data(Reference(ws3, min_col=2, min_row=cs-1, max_row=cs+2), titles_from_data=True)
    ch3_pie.set_categories(Reference(ws3, min_col=1, min_row=cs, max_row=cs+2))
    ws3.add_chart(ch3_pie, 'F' + str(cs + 4))

    # Todas las operaciones
    r3 = cs + 5
    ws3.cell(row=r3, column=1, value='TODAS LAS OPERACIONES').font = f_ti
    set_hdr(ws3, r3+1, 1, ['Hora', 'Nombre', 'Mesa', 'Monto ($)', 'Caja'])
    r3 += 2
    for i, t in enumerate(ul.get('transactions', [])):
        ws3.cell(row=r3, column=1, value=t.get('time', ''))
        ws3.cell(row=r3, column=2, value=t['name'])
        ws3.cell(row=r3, column=3, value=t.get('mesa', ''))
        nfmt(ws3.cell(row=r3, column=4), t['amount'])
        ws3.cell(row=r3, column=5, value=f'Caja {t.get("caja", 1)}')
        set_row(ws3, r3, 1, 5, alt=(i % 2 == 0))
        r3 += 1

    for ci, w in enumerate([10, 26, 8, 14, 8], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    fecha_str = datetime.now().strftime('%Y%m%d')
    return send_file(buf, as_attachment=True,
                     download_name=f'jagger_{fecha_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/export/gsheets', methods=['POST'])
def export_gsheets():
    try:
        import gspread
        from google.oauth2.service_account import Credentials as GCreds
    except ImportError:
        return jsonify({'ok': False, 'error': 'Instalá: pip install gspread google-auth'}), 500

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    sheet_id = os.environ.get('GSHEETS_ID', '').strip()
    creds_json_env = os.environ.get('GSHEETS_CREDS_JSON', '').strip()
    creds_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'gsheets_creds.json')

    if not sheet_id:
        return jsonify({'ok': False, 'error': 'Falta variable de entorno GSHEETS_ID con el ID del spreadsheet'}), 400
    try:
        if creds_json_env:
            creds = GCreds.from_service_account_info(json.loads(creds_json_env), scopes=SCOPES)
        elif os.path.exists(creds_file):
            creds = GCreds.from_service_account_file(creds_file, scopes=SCOPES)
        else:
            return jsonify({'ok': False, 'error': 'Falta gsheets_creds.json o variable GSHEETS_CREDS_JSON'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error en credenciales: {e}'}), 500

    try:
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'No se pudo abrir el spreadsheet: {e}'}), 500

    body = request.get_json() or {}
    modo = body.get('modo', 'all')  # 'all' o 'last'
    historial = load_historial()
    if not historial:
        return jsonify({'ok': False, 'error': 'Sin noches registradas aún'}), 400

    noches_a_sincronizar = [historial[-1]] if modo == 'last' else historial

    def safe_sheet(nombre):
        try: return sh.worksheet(nombre)
        except gspread.WorksheetNotFound: return sh.add_worksheet(title=nombre, rows=200, cols=12)

    # ── Hoja RESUMEN ──────────────────────────────────────────
    try:
        ws_res = safe_sheet('Resumen')
        ws_res.clear()
        total_all = sum(n['total'] for n in historial)
        rk_all = {}
        for n in historial:
            for r in n.get('ranking', []):
                rk_all[r['name']] = rk_all.get(r['name'], 0) + r['total']
        rk_sorted = sorted(rk_all.items(), key=lambda x: -x[1])
        header_rows = [
            ['JAGGER CLUB — HISTORIAL VIP', '', datetime.now().strftime('%d/%m/%Y %H:%M')],
            [],
            ['TOTAL ACUMULADO', f'${total_all:,.0f}'.replace(',', '.')],
            ['NOCHES REGISTRADAS', len(historial)],
            ['OPERACIONES', sum(n['operaciones'] for n in historial)],
            [],
            ['RANKING GENERAL'],
            ['#', 'Nombre', 'Total ($)'],
        ] + [[i+1, nm, tot] for i, (nm, tot) in enumerate(rk_sorted[:20])]
        ws_res.update('A1', header_rows)
    except Exception:
        pass

    # ── Hoja NOCHES ───────────────────────────────────────────
    try:
        ws_noches = safe_sheet('Noches')
        ws_noches.clear()
        n_header = [['Fecha', 'Cierre', 'Total ($)', 'Ops', '1°', '2°', '3°', 'Abajo', 'Extendido', 'VIP']]
        n_rows = []
        for n in reversed(historial):
            rnk = n.get('ranking', [])
            pc  = n.get('por_caja', {})
            n_rows.append([
                n['fecha'], n.get('hora_cierre',''), n['total'], n['operaciones'],
                rnk[0]['name'] if len(rnk)>0 else '', rnk[1]['name'] if len(rnk)>1 else '', rnk[2]['name'] if len(rnk)>2 else '',
                pc.get(1, pc.get('1', 0)), pc.get(2, pc.get('2', 0)), pc.get(3, pc.get('3', 0)),
            ])
        ws_noches.update('A1', n_header + n_rows)
    except Exception:
        pass

    # ── Una hoja por noche ────────────────────────────────────
    errores = []
    for noche in noches_a_sincronizar:
        try:
            nombre_hoja = f"Noche {noche['fecha']}"
            ws = safe_sheet(nombre_hoja)
            ws.clear()
            pc = noche.get('por_caja', {})
            rows = [
                [f"NOCHE {noche['fecha']} — cierre {noche.get('hora_cierre','?')} — Total: ${noche['total']:,.0f}".replace(',','.')],
                [f"Abajo: ${pc.get(1,pc.get('1',0)):,.0f}  |  Extendido: ${pc.get(2,pc.get('2',0)):,.0f}  |  VIP: ${pc.get(3,pc.get('3',0)):,.0f}".replace(',','.')],
                [],
                ['RANKING', '', '', ''],
                ['#', 'Nombre', 'Mesa', 'Total ($)'],
            ] + [[i+1, r['name'], r.get('mesa',''), r['total']] for i, r in enumerate(noche.get('ranking',[]))] + [
                [],
                ['OPERACIONES', '', '', ''],
                ['Hora', 'Nombre', 'Mesa', 'Monto ($)'],
            ] + [[t.get('time',''), t['name'], t.get('mesa',''), t['amount']] for t in noche.get('transactions',[])]
            ws.update('A1', rows)
        except Exception as e:
            errores.append(str(e))

    url = f'https://docs.google.com/spreadsheets/d/{sheet_id}'
    msg = f'Sincronizado: {len(noches_a_sincronizar)} noche(s)'
    if errores: msg += f' (errores: {len(errores)})'
    return jsonify({'ok': True, 'url': url, 'msg': msg})

@app.route('/api/tarjetas/recargar', methods=['POST'])
def recargar_tarjeta():
    body = request.get_json() or {}
    codigo = str(body.get('codigo', ''))
    try:
        monto = float(body.get('monto', 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    if not codigo or monto <= 0:
        return jsonify({'ok': False, 'error': 'Datos inválidos'}), 400
    with lock:
        data = load_data()
        if codigo not in data.get('tarjetas', {}):
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        data['tarjetas'][codigo]['saldo_actual'] += monto
        nuevo_saldo = data['tarjetas'][codigo]['saldo_actual']
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo})

@app.route('/api/auth', methods=['POST'])
def auth():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('password', '')) == str(cfg.get('password', '1212')):
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Contraseña incorrecta'}), 401

@app.route('/api/auth/change', methods=['POST'])
def auth_change():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('current', '')) != str(cfg.get('password', '1212')):
        return jsonify({'ok': False, 'error': 'PIN actual incorrecto'}), 401
    nueva = str(body.get('new', '')).strip()
    if not nueva.isdigit() or len(nueva) != 4:
        return jsonify({'ok': False, 'error': 'El PIN debe tener exactamente 4 dígitos numéricos'}), 400
    cfg['password'] = nueva
    save_config(cfg)
    return jsonify({'ok': True})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
